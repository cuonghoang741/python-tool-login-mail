import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import os
from functools import partial
import re
import json
import time
import random
import threading
import urllib.request
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

# Import story tab
try:
    from tabs.flow_tab_story import StoryPromptGenerator
    _HAS_STORY_TAB = True
except ImportError:
    StoryPromptGenerator = None
    _HAS_STORY_TAB = False


# Optional modern theming with ttkbootstrap (for a significantly improved look)
try:
    from ttkbootstrap import Style as TtkbStyle
    from ttkbootstrap import Window as TtkbWindow
    _HAS_TTKBOOTSTRAP = True
except Exception:
    TtkbStyle = None
    TtkbWindow = None
    _HAS_TTKBOOTSTRAP = False

class FlowBrowserTool:
    def __init__(self, root: tk.Tk | None, use_tk_ui: bool = True, ui_callbacks: dict | None = None):
        self.root = root
        self.use_tk_ui = use_tk_ui
        # Optional callbacks for non-Tk UI adapters (PySide6)
        # ui_callbacks keys: on_log(str), on_status(text,color), on_exec_status(text,color), on_jobs_update()
        self.ui_callbacks = ui_callbacks or {}
        
        # Add callback for story tab to execute tab transfer
        self.ui_callbacks['import_excel_and_switch'] = self._import_excel_and_switch_tab
        self.root.title("🎬 Google Flow Tool")
        self.root.geometry("900x650")
        self.root.resizable(True, True)
        
        # Apply luxury theme
        self._apply_luxury_theme()

        # Runtime state
        self.driver = None
        self.current_email = None
        self.current_cache_dir = None
        self.current_user_agent = None
        self.login_success = False
        # Execution state & queue
        self.exec_driver = None
        self.stop_exec = False
        self.exec_queue = []
        self.queue_running = False
        self.exec_driver = None
        self.stop_exec = False
        # Synchronization for queue operations
        self.queue_lock = threading.Lock()
        # Per-account execution states and tracking
        self.account_states = {}
        self.exec_current_jobs = {}
        self.exec_drivers = {}
        # Job counter for unique indexing
        self.job_counter = 0

        # Error log storage
        self.error_events = []
        self.error_log_path = os.path.join(os.getcwd(), "error_events.json")
        try:
            self._load_error_events()
        except Exception:
            pass

        # Profiles (cache per email)
        self.flow_profiles_path = os.path.join(os.getcwd(), "chrome_cache", "flow_profiles.json")
        self.flow_profiles = {}
        self._load_profiles()

        # User agents
        self.user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:122.0) Gecko/20100101 Firefox/122.0",
        ]

        if self.use_tk_ui:
            self._build_ui()

        # Route error popups to error log card (no blocking alerts)
        try:
            self._orig_showerror = messagebox.showerror
            def _no_alert_showerror(title, msg):
                try:
                    self._log_error(f"{title}: {msg}")
                except Exception:
                    pass
            messagebox.showerror = _no_alert_showerror
        except Exception:
            pass

    def _apply_luxury_theme(self) -> None:
        """Apply a modern, luxurious ttk theme across widgets."""
        # Color palette (deep slate with gold accents)
        self.colors = {
            'bg': '#0F1115',
            'surface': '#171A21',
            'border': '#2A2F3A',
            'text': '#EAECEF',
            'subtle': '#9AA4AF',
            'accent': '#D4AF37',  # gold
            'accent_hover': '#B8902E',
            'success': '#2ECC71',
            'warning': '#F1C40F',
            'error': '#E74C3C',
            'info': '#58A6FF',
        }
        # Prefer ttkbootstrap themes if available for a modern look
        if _HAS_TTKBOOTSTRAP and TtkbStyle is not None:
            self.style = TtkbStyle(theme='superhero')
        else:
            self.style = ttk.Style()
            self.style.theme_use('clam')
        # Window background
        self.root.configure(bg=self.colors['bg'])
        
        # Base backgrounds
        for element in ('TFrame', 'TLabelframe', 'TLabelframe.Label'):
            self.style.configure(element, background=self.colors['bg'], foreground=self.colors['text'])
        self.style.configure('Card.TFrame', background=self.colors['surface'], bordercolor=self.colors['border'], borderwidth=1, relief='ridge')
        self.style.configure('CardInner.TFrame', background='#141923')
        self.style.configure('CardHover.TFrame', background='#1B2330', bordercolor=self.colors['border'], borderwidth=1, relief='ridge')
        self.style.configure('Card.TLabelframe', background=self.colors['surface'], bordercolor=self.colors['border'], borderwidth=1, relief='ridge')
        
        # Labels
        self.style.configure('TLabel', background=self.colors['bg'], foreground=self.colors['text'], font=('Segoe UI', 10))
        self.style.configure('Title.TLabel', background=self.colors['bg'], foreground=self.colors['accent'], font=('Segoe UI Semibold', 18))
        self.style.configure('Subtitle.TLabel', background=self.colors['bg'], foreground=self.colors['subtle'], font=('Segoe UI', 10, 'bold'))
        self.style.configure('Success.TLabel', background=self.colors['bg'], foreground=self.colors['success'], font=('Segoe UI', 10))
        self.style.configure('Error.TLabel', background=self.colors['bg'], foreground=self.colors['error'], font=('Segoe UI', 10))
        self.style.configure('Info.TLabel', background=self.colors['bg'], foreground=self.colors['info'], font=('Segoe UI', 10))
        self.style.configure('Warning.TLabel', background=self.colors['bg'], foreground=self.colors['warning'], font=('Segoe UI', 10))
        
        # Buttons
        base_button_opts = {
            'font': ('Segoe UI', 10),
            'borderwidth': 0,
        }
        self.style.configure('TButton', **base_button_opts, padding=(12, 8), background=self.colors['surface'], foreground=self.colors['text'])
        self.style.map('TButton', background=[('active', '#1E2530')])
        self.style.configure('Primary.TButton', **base_button_opts, padding=(12, 8), background=self.colors['accent'], foreground='#0B0C10')
        self.style.map('Primary.TButton', background=[('active', self.colors['accent_hover'])], foreground=[('active', '#000000')])
        self.style.configure('Secondary.TButton', **base_button_opts, padding=(12, 8), background='#1A1F27', foreground=self.colors['text'])
        self.style.map('Secondary.TButton', background=[('active', '#222938')], foreground=[('active', self.colors['text'])])
        self.style.configure('Accent.TButton', **base_button_opts, padding=(16, 10), background=self.colors['accent'], foreground='#0B0C10')
        self.style.map('Accent.TButton', background=[('active', self.colors['accent_hover'])], foreground=[('active', '#000000')])
        
        # Inputs: Entry / Combobox / Checkbutton / Radiobutton
        entry_like = {
            'fieldbackground': '#1B222C',
            'foreground': self.colors['text'],
            'background': self.colors['bg'],
            'bordercolor': self.colors['border'],
            'lightcolor': self.colors['border'],
            'darkcolor': self.colors['border'],
            'padding': (8, 6),
        }
        self.style.configure('TEntry', **entry_like)
        self.style.configure('TCombobox', **entry_like)
        self.style.map('TCombobox', fieldbackground=[('readonly', '#1B222C')], background=[('readonly', '#1B222C')])
        self.style.configure('TCheckbutton', background=self.colors['bg'], foreground=self.colors['text'])
        self.style.configure('TRadiobutton', background=self.colors['bg'], foreground=self.colors['text'])
        
        # Notebook
        self.style.configure('TNotebook', background=self.colors['bg'], borderwidth=0, tabmargins=[4, 6, 4, 0])
        self.style.configure('TNotebook.Tab', background='#1A1F27', foreground=self.colors['text'], padding=[16, 10], font=('Segoe UI', 10, 'bold'))
        self.style.map('TNotebook.Tab', background=[('selected', self.colors['surface']), ('active', '#1E2530')])
        
        # Treeview
        self.style.configure('Treeview', background=self.colors['surface'], foreground=self.colors['text'], fieldbackground=self.colors['surface'], rowheight=26, bordercolor=self.colors['border'], borderwidth=1)
        self.style.configure('Treeview.Heading', background='#1A1F27', foreground=self.colors['subtle'], font=('Segoe UI', 10, 'bold'))
        self.style.map('Treeview', background=[('selected', '#253044')], foreground=[('selected', self.colors['text'])])
        
        # Scrollbar
        self.style.configure('Vertical.TScrollbar', background='#1A1F27', troughcolor=self.colors['bg'], bordercolor=self.colors['bg'])
        self.style.configure('Horizontal.TScrollbar', background='#1A1F27', troughcolor=self.colors['bg'], bordercolor=self.colors['bg'])

        # Badges
        self.style.configure('Badge.Running.TLabel', background='#20314A', foreground=self.colors['info'], padding=(6, 2))
        self.style.configure('Badge.Queued.TLabel', background='#2B2B1A', foreground=self.colors['warning'], padding=(6, 2))
        self.style.configure('Badge.Success.TLabel', background='#153724', foreground=self.colors['success'], padding=(6, 2))
        self.style.configure('Badge.Error.TLabel', background='#3A1B1B', foreground=self.colors['error'], padding=(6, 2))

    def _create_rounded_card(self, parent: tk.Widget, padding: int = 10) -> tuple[tk.Canvas, int, ttk.Frame]:
        """Create a rounded-corner card using a Canvas and return (canvas, shape_id, content_frame)."""
        try:
            radius = 12
            canvas = tk.Canvas(parent, bd=0, highlightthickness=0, bg=self.colors['surface'])
            # Placeholder size; will expand to parent width
            width = max(parent.winfo_width(), 280)
            height = 10 + padding * 2
            x1, y1, x2, y2 = 4, 4, width - 4, height

            def rounded_rect_path(x1, y1, x2, y2, r):
                return [
                    x1+r, y1, x2-r, y1, x2, y1, x2, y1+r, x2, y2-r, x2, y2,
                    x2-r, y2, x1+r, y2, x1, y2, x1, y2-r, x1, y1+r, x1, y1
                ]

            pts = rounded_rect_path(x1, y1, x2, y2, radius)
            shape = canvas.create_polygon(
                pts, smooth=True, splinesteps=36,
                fill=self.colors['surface'], outline=self.colors['border']
            )

            content = ttk.Frame(canvas)
            window_id = canvas.create_window(12, 12, anchor='nw', window=content)

            # Resize behavior to fit parent width
            def _resize(_e=None):
                try:
                    w = parent.winfo_width() - 8
                    if w < 260:
                        w = 260
                    # Estimate height from content
                    content.update_idletasks()
                    h = content.winfo_reqheight() + 24
                    canvas.configure(width=w, height=h)
                    # Recompute rounded rect points
                    x1n, y1n, x2n, y2n = 4, 4, w - 4, h - 4
                    new_pts = rounded_rect_path(x1n, y1n, x2n, y2n, radius)
                    canvas.coords(shape, *new_pts)
                    canvas.itemconfigure(window_id)
                except Exception:
                    pass

            try:
                parent.bind('<Configure>', _resize)
            except Exception:
                pass
            # Initial size
            _resize()
            return canvas, shape, content
        except Exception:
            # Fallback to simple frame if anything fails
            fallback = ttk.Frame(parent, style='Card.TFrame', padding=padding)
            return fallback, -1, fallback

    # ===================== UX Helpers =====================
    def _attach_tooltip(self, widget: tk.Widget, text: str) -> None:
        """Attach a simple tooltip to a widget."""
        try:
            tip = {'win': None}

            def show_tip(_e=None):
                try:
                    if tip['win'] is not None:
                        return
                    x = widget.winfo_rootx() + 12
                    y = widget.winfo_rooty() + widget.winfo_height() + 8
                    win = tk.Toplevel(widget)
                    win.wm_overrideredirect(True)
                    win.configure(bg=self.colors['border'])
                    frame = tk.Frame(win, bg=self.colors['surface'], bd=0, highlightthickness=1, highlightbackground=self.colors['border'])
                    frame.pack()
                    lbl = tk.Label(frame, text=text, bg=self.colors['surface'], fg=self.colors['text'], font=('Segoe UI', 9), padx=8, pady=4)
                    lbl.pack()
                    win.wm_geometry(f"+{x}+{y}")
                    tip['win'] = win
                except Exception:
                    pass

            def hide_tip(_e=None):
                try:
                    if tip['win'] is not None:
                        tip['win'].destroy()
                        tip['win'] = None
                except Exception:
                    pass

            widget.bind('<Enter>', show_tip)
            widget.bind('<Leave>', hide_tip)
        except Exception:
            pass

    # ===================== UI =====================
    def _build_ui(self) -> None:
        self.notebook = ttk.Notebook(self.root)
        self.notebook.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        # Bind resize event để responsive
        self.root.bind('<Configure>', self._on_window_resize)

        # Tabs
        login_tab = ttk.Frame(self.notebook)
        exec_tab = ttk.Frame(self.notebook)
        story_tab = ttk.Frame(self.notebook)
        self.notebook.add(login_tab, text="🔐 Đăng nhập & Tài khoản")
        self.notebook.add(exec_tab, text="🎥 Execute Media")
        self.notebook.add(story_tab, text="📚 All Story Prompts")
        # Mặc định chọn tab Execute Media
        self.notebook.select(1)

        # ===== Login Tab =====
        frame = ttk.Frame(login_tab, padding="20")
        frame.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        login_tab.columnconfigure(0, weight=1)
        login_tab.rowconfigure(0, weight=1)
        for i in range(2):
            frame.columnconfigure(i, weight=1)
        frame.rowconfigure(6, weight=1)  # Profiles section expandable

        self.title = ttk.Label(frame, text="🎬 Google Flow Login", style='Title.TLabel')
        self.title.grid(row=0, column=0, columnspan=2, pady=(0, 20))

        method_group = ttk.LabelFrame(frame, text="🔑 Phương thức đăng nhập", padding="15")
        method_group.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        method_group.configure(style='Card.TLabelframe')
        self.login_method = tk.StringVar(value="password")
        # Chỉ giữ password login, ẩn nhóm chọn phương thức
        try:
            method_group.grid_remove()
        except Exception:
            pass

        creds = ttk.LabelFrame(frame, text="📝 Thông tin tài khoản", padding="15")
        creds.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        creds.configure(style='Card.TLabelframe')
        creds.columnconfigure(1, weight=1)

        ttk.Label(creds, text="📧 Email:", style='Subtitle.TLabel').grid(row=0, column=0, sticky=tk.W, pady=(0, 8))
        self.email_entry = ttk.Entry(creds)
        self.email_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=(0, 8), padx=(10, 0))

        self.password_label = ttk.Label(creds, text="🔒 Mật khẩu:", style='Subtitle.TLabel')
        self.password_entry = ttk.Entry(creds, show="*")
        # Luôn hiển thị trường mật khẩu
        self.password_label.grid(row=1, column=0, sticky=tk.W, pady=(0, 8))
        self.password_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=(0, 8), padx=(10, 0))

        self.note_label = ttk.Label(creds, text="💡 Đăng nhập bằng mật khẩu Google (có thể cần 2FA)", style='Info.TLabel')
        self.note_label.grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(8, 0))

        self.login_btn = ttk.Button(frame, text="🚀 Đăng nhập Google Flow", command=self._login_flow, style='Accent.TButton')
        self.login_btn.grid(row=3, column=0, columnspan=2, pady=(15, 0))

        self.status_label = ttk.Label(frame, text="⏳ Chưa đăng nhập", style='Warning.TLabel')
        self.status_label.grid(row=4, column=0, columnspan=2, pady=(10, 0))

        sep = ttk.Separator(frame, orient=tk.HORIZONTAL)
        sep.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=15)

        profiles_group = ttk.LabelFrame(frame, text="👥 Tài khoản đã đăng nhập (cache)", padding="15")
        profiles_group.grid(row=6, column=0, columnspan=2, sticky=(tk.N, tk.S, tk.W, tk.E))
        profiles_group.configure(style='Card.TLabelframe')
        profiles_group.columnconfigure(0, weight=1)
        profiles_group.rowconfigure(0, weight=1)

        self.profiles_list = tk.Listbox(profiles_group, height=6, selectmode=tk.SINGLE,
                                        bg=self.colors['surface'], fg=self.colors['text'],
                                        highlightthickness=1, highlightbackground=self.colors['border'],
                                        selectbackground='#253044', selectforeground=self.colors['text'])
        self.profiles_list.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        sb = ttk.Scrollbar(profiles_group, orient=tk.VERTICAL, command=self.profiles_list.yview)
        sb.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.profiles_list.configure(yscrollcommand=sb.set)

        actions = ttk.Frame(profiles_group)
        actions.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        ttk.Button(actions, text="👁️ Mở profile", command=self._open_selected_profile, style='Secondary.TButton').pack(side=tk.LEFT)
        ttk.Button(actions, text="🗑️ Xóa cache", command=self._delete_selected_profile, style='Secondary.TButton').pack(side=tk.LEFT, padx=8)

        self._refresh_profiles_list()

        # ===== Execute Tab (scrollable) =====
        # Simple container without custom canvas/scroll to avoid black screen
        exec_tab.columnconfigure(0, weight=1)
        exec_tab.rowconfigure(0, weight=1)
        ex = ttk.Frame(exec_tab, padding="20")
        ex.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        # Layout: left content (col 0-1) and right jobs panel (col 2)
        ex.columnconfigure(0, weight=3)
        ex.columnconfigure(1, weight=3)
        ex.columnconfigure(2, weight=2)
        ex.rowconfigure(2, weight=1)

        exec_title = ttk.Label(ex, text="🎥 Execute Media Workflow", style='Title.TLabel')
        # Constrain title to left content area (columns 0-1) so right panel can start at row 0 col 2
        exec_title.grid(row=0, column=0, columnspan=2, pady=(0, 20))

        # Email selection
        email_frame = ttk.LabelFrame(ex, text="👤 Chọn tài khoản", padding="15")
        email_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        email_frame.configure(style='Card.TLabelframe')
        email_frame.columnconfigure(1, weight=1)

        ttk.Label(email_frame, text="📧 Email:", style='Subtitle.TLabel').grid(row=0, column=0, sticky=tk.W, pady=(0, 8))
        self.exec_email = tk.StringVar()
        self.exec_email_combo = ttk.Combobox(email_frame, textvariable=self.exec_email, state="readonly")
        self.exec_email_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(10, 8), pady=(0, 8))
        ttk.Button(email_frame, text="🔄 Làm mới", command=self._refresh_exec_emails, style='Secondary.TButton').grid(row=0, column=2, pady=(0, 8))
        # Hide email input section in Execute task
        try:
            email_frame.grid_remove()
        except Exception:
            pass

        # Workflow selection
        workflow_frame = ttk.LabelFrame(ex, text="⚙️ Loại Workflow", padding="15")
        workflow_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        workflow_frame.configure(style='Card.TLabelframe')
        workflow_frame.columnconfigure(0, weight=1)
        workflow_frame.columnconfigure(1, weight=1)

        ttk.Label(workflow_frame, text="Workflow:", style='Subtitle.TLabel').grid(row=0, column=0, sticky=tk.W, pady=(0, 8))
        self.workflow = tk.StringVar(value="frames_to_video")
        ttk.Radiobutton(workflow_frame, text="📝 Text to Video", variable=self.workflow, value="text_to_video").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(workflow_frame, text="🖼️ Frames to Video", variable=self.workflow, value="frames_to_video").grid(row=1, column=1, sticky=tk.W, pady=5)
        # Hide workflow selection in Execute task
        try:
            workflow_frame.grid_remove()
        except Exception:
            pass

        # Prompt section
        prompt_frame = ttk.LabelFrame(ex, text="💬 Prompt (Text to Video)", padding="15")
        prompt_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        prompt_frame.configure(style='Card.TLabelframe')
        prompt_frame.columnconfigure(0, weight=1)

        ttk.Label(prompt_frame, text="Prompt (Text to Video):", style='Subtitle.TLabel').grid(row=0, column=0, sticky=tk.NW, pady=(0, 8))
        self.prompt_text = scrolledtext.ScrolledText(prompt_frame, height=6, wrap=tk.WORD, bg='#1B222C', fg=self.colors['text'], insertbackground=self.colors['text'], highlightthickness=1, highlightbackground=self.colors['border'])
        self.prompt_text.grid(row=1, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        # Hide prompt input in Execute task
        try:
            prompt_frame.grid_remove()
        except Exception:
            pass

        # Media upload section
        media_frame = ttk.LabelFrame(ex, text="📁 Upload Media (Frames to Video)", padding="15")
        media_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        media_frame.configure(style='Card.TLabelframe')

        ttk.Label(media_frame, text="Upload media (chỉ 1 file ảnh):", style='Subtitle.TLabel').grid(row=0, column=0, sticky=tk.W, pady=(0, 8))
        media_input_frame = ttk.Frame(media_frame)
        media_input_frame.grid(row=1, column=0, sticky=(tk.W, tk.E))
        media_input_frame.columnconfigure(0, weight=1)
        
        self.media_paths = tk.StringVar()
        self.media_entry = ttk.Entry(media_input_frame, textvariable=self.media_paths)
        self.media_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 8))
        ttk.Button(media_input_frame, text="🖼️ Chọn ảnh", command=self._choose_image_file, style='Secondary.TButton').grid(row=0, column=1)
        # Hide media upload in Execute task
        try:
            media_frame.grid_remove()
        except Exception:
            pass

        # Configuration section
        cfg = ttk.LabelFrame(ex, text="⚙️ Cấu hình", padding="15")
        cfg.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        cfg.configure(style='Card.TLabelframe')
        for i in range(6):
            cfg.columnconfigure(i, weight=1)

        # Browser options
        ttk.Label(cfg, text="Browser mode", style='Subtitle.TLabel').grid(row=0, column=0, sticky=tk.W, pady=(0, 8))
        self.headless_mode = tk.BooleanVar(value=True)
        ttk.Checkbutton(cfg, text="Headless (ẩn browser)", variable=self.headless_mode).grid(row=0, column=1, sticky=tk.W, pady=(0, 8))

        # Settings (popover) - Aspect ratio, Outputs per prompt, Model
        ttk.Label(cfg, text="Aspect ratio", style='Subtitle.TLabel').grid(row=1, column=0, sticky=tk.W, pady=(8, 0))
        self.aspect_ratio = ttk.Combobox(cfg, values=["16:9", "9:16"], state="readonly")
        self.aspect_ratio.set("16:9")
        self.aspect_ratio.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=(8, 0))

        ttk.Label(cfg, text="Outputs per prompt", style='Subtitle.TLabel').grid(row=1, column=2, sticky=tk.W, pady=(8, 0))
        self.outputs_per_prompt = ttk.Combobox(cfg, values=["1", "2", "3", "4"], state="readonly")
        self.outputs_per_prompt.set("2")
        self.outputs_per_prompt.grid(row=1, column=3, sticky=(tk.W, tk.E), pady=(8, 0))

        ttk.Label(cfg, text="Model", style='Subtitle.TLabel').grid(row=1, column=4, sticky=tk.W, pady=(8, 0))
        # Model order must follow popover order to keep index/text consistent
        self.model_choice = ttk.Combobox(
            cfg,
            values=[
                "Veo 3 - Fast",
                "Veo 2 - Fast",
                "Veo 3 - Quality",
                "Veo 2 - Quality",
            ],
            state="readonly"
        )
        self.model_choice.set("Veo 3 - Fast")
        self.model_choice.grid(row=1, column=5, sticky=(tk.W, tk.E), pady=(8, 0))
        # Hide configuration section in Execute task
        try:
            cfg.grid_remove()
        except Exception:
            pass

        # Action buttons
        action_frame = ttk.Frame(ex)
        # Move action buttons up to fill space of hidden inputs
        action_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Arrange buttons in 2 rows x 2 columns for more space
        for i in range(2):
            try:
                action_frame.columnconfigure(i, weight=1)
            except Exception:
                pass
        btn_import = ttk.Button(action_frame, text="📥 Import Excel", command=self._import_excel_and_dispatch, style='Secondary.TButton')
        btn_import.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=6, pady=(0, 8))
        btn_tpl = ttk.Button(action_frame, text="⬇️ Tải Template", command=self._download_excel_template, style='Secondary.TButton')
        btn_tpl.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=6, pady=(0, 8))
        btn_exec = ttk.Button(action_frame, text="▶️ Execute", command=self._execute_workflow, style='Accent.TButton')
        btn_exec.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=6)
        btn_stop = ttk.Button(action_frame, text="⏹️ Stop", command=self._stop_execution, style='Secondary.TButton')
        btn_stop.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=6)
        # Attach tooltips
        self._attach_tooltip(btn_exec, "Thực thi theo cấu hình và dữ liệu nhập")
        self._attach_tooltip(btn_stop, "Dừng quá trình đang chạy")
        self._attach_tooltip(btn_import, "Import Excel và đưa vào hàng đợi")
        self._attach_tooltip(btn_tpl, "Tải file template Excel")

        # Quick toggle: Headless mode (visible in Execute tab)
        try:
            self.headless_mode_chk = ttk.Checkbutton(action_frame, text="Headless (ẩn browser)", variable=self.headless_mode)
            self.headless_mode_chk.grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(6, 0))
            self._attach_tooltip(self.headless_mode_chk, "Chạy ẩn trình duyệt khi thực thi workflow")
        except Exception:
            pass

        self.exec_status = ttk.Label(ex, text="✅ Sẵn sàng", style='Success.TLabel')
        # Place status just below actions
        self.exec_status.grid(row=2, column=0, columnspan=2, sticky=tk.W)

        # Jobs view (right side) simplified with tables
        jobs_side = ttk.LabelFrame(ex, text="📋 Tiến trình", padding="10")
        jobs_side.grid(row=0, column=2, rowspan=4, sticky=(tk.N, tk.S, tk.W, tk.E), padx=(15, 0))
        jobs_side.configure(style='Card.TLabelframe')
        jobs_side.columnconfigure(0, weight=1)
        jobs_side.rowconfigure(1, weight=1)
        jobs_side.rowconfigure(3, weight=1)

        header_running = ttk.Frame(jobs_side)
        header_running.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 6))
        header_running.columnconfigure(0, weight=1)
        ttk.Label(header_running, text="Đang chạy", style='Subtitle.TLabel').grid(row=0, column=0, sticky=tk.W)
        ttk.Label(header_running, text="RUNNING", style='Badge.Running.TLabel').grid(row=0, column=1, sticky=tk.E)

        run_wrap = ttk.Frame(jobs_side)
        run_wrap.grid(row=1, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        run_wrap.columnconfigure(0, weight=1)
        run_wrap.rowconfigure(0, weight=1)
        self.running_table = ttk.Treeview(run_wrap, columns=("email","wf","img","prompt"), show="headings", height=6)
        for col, txt, w in (("email","Email",110),("wf","Workflow",90),("img","Image",100),("prompt","Prompt",240)):
            self.running_table.heading(col, text=txt)
            self.running_table.column(col, width=w, stretch=True)
        run_scroll = ttk.Scrollbar(run_wrap, orient=tk.VERTICAL, command=self.running_table.yview)
        self.running_table.configure(yscrollcommand=run_scroll.set)
        self.running_table.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        run_scroll.grid(row=0, column=1, sticky=(tk.N, tk.S))

        header_queue = ttk.Frame(jobs_side)
        header_queue.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(10, 6))
        header_queue.columnconfigure(0, weight=1)
        ttk.Label(header_queue, text="Đang đợi", style='Subtitle.TLabel').grid(row=0, column=0, sticky=tk.W)
        ttk.Label(header_queue, text="QUEUE", style='Badge.Queued.TLabel').grid(row=0, column=1, sticky=tk.E)

        queue_wrap = ttk.Frame(jobs_side)
        queue_wrap.grid(row=3, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        queue_wrap.columnconfigure(0, weight=1)
        queue_wrap.rowconfigure(0, weight=1)
        self.queue_table = ttk.Treeview(queue_wrap, columns=("email","wf","img","prompt"), show="headings", height=6)
        for col, txt, w in (("email","Email",110),("wf","Workflow",90),("img","Image",100),("prompt","Prompt",240)):
            self.queue_table.heading(col, text=txt)
            self.queue_table.column(col, width=w, stretch=True)
        queue_scroll = ttk.Scrollbar(queue_wrap, orient=tk.VERTICAL, command=self.queue_table.yview)
        self.queue_table.configure(yscrollcommand=queue_scroll.set)
        self.queue_table.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        queue_scroll.grid(row=0, column=1, sticky=(tk.N, tk.S))

        # Progress log card
        log_frame = ttk.LabelFrame(ex, text="📜 Log tiến trình", padding="10")
        # Move log up under status
        log_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.N, tk.S, tk.W, tk.E))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        self.exec_log = scrolledtext.ScrolledText(log_frame, height=8, wrap=tk.WORD, state='disabled',
                                                  bg='#10141B', fg=self.colors['text'],
                                                  insertbackground=self.colors['text'],
                                                  highlightthickness=1, highlightbackground=self.colors['border'])
        self.exec_log.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        # Make the log row expandable now that inputs are hidden
        ex.rowconfigure(3, weight=1)

        # Error log card (dedicated area to display failed processes/errors)
        err_frame = ttk.LabelFrame(ex, text="❌ Nhật ký lỗi", padding="10")
        err_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.N, tk.S, tk.W, tk.E), pady=(10, 0))
        err_frame.columnconfigure(0, weight=1)
        err_frame.rowconfigure(0, weight=1)
        self.error_log = scrolledtext.ScrolledText(err_frame, height=6, wrap=tk.WORD, state='disabled',
                                                   bg='#1A1416', fg=self.colors['error'],
                                                   insertbackground=self.colors['error'],
                                                   highlightthickness=1, highlightbackground=self.colors['border'])
        self.error_log.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        # Load existing errors into UI if any
        try:
            self._refresh_error_log_ui()
        except Exception:
            pass
        ex.rowconfigure(4, weight=1)

        self._refresh_exec_emails()
        # Initialize jobs view
        self.exec_current_job = None
        self._refresh_jobs_view()

        # ===== Story Tab =====
        if _HAS_STORY_TAB and StoryPromptGenerator is not None:
            try:
                self.story_generator = StoryPromptGenerator(story_tab, self.ui_callbacks)
            except Exception as e:
                print(f"Failed to initialize story tab: {e}")
                # Create a simple error message frame
                error_frame = ttk.Frame(story_tab, padding="20")
                error_frame.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
                ttk.Label(error_frame, text=f"❌ Lỗi khởi tạo Story Tab: {e}", 
                         style='Error.TLabel').pack()
        else:
            # Create a simple message frame if story tab is not available
            error_frame = ttk.Frame(story_tab, padding="20")
            error_frame.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
            ttk.Label(error_frame, text="❌ Story Tab không khả dụng", 
                     style='Error.TLabel').pack()

    # ===================== Responsive Helpers =====================
    def _on_window_resize(self, event):
        """Handle window resize for responsive design"""
        if event.widget == self.root:
            # Update window size info for responsive adjustments
            width = self.root.winfo_width()
            height = self.root.winfo_height()
            
            # Adjust font sizes based on window size
            if width < 800 or height < 600:
                # Small window - reduce font sizes
                self._adjust_font_sizes("small")
            elif width > 1200 and height > 800:
                # Large window - increase font sizes
                self._adjust_font_sizes("large")
            else:
                # Normal window - default font sizes
                self._adjust_font_sizes("normal")
    
    def _adjust_font_sizes(self, size_mode):
        """Adjust font sizes based on window size"""
        try:
            if size_mode == "small":
                font_size = 8
                title_size = 13
                tab_font = ("Segoe UI", 9, "bold")
                btn_font = ("Segoe UI", 9, "bold")
            elif size_mode == "large":
                font_size = 11
                title_size = 18
                tab_font = ("Segoe UI", 11, "bold")
                btn_font = ("Segoe UI", 11, "bold")
            else:  # normal
                font_size = 9
                title_size = 15
                tab_font = ("Segoe UI", 10, "bold")
                btn_font = ("Segoe UI", 10, "bold")
            
            # Update title font
            if hasattr(self, 'title'):
                self.title.config(font=("Arial", title_size, "bold"))
            
            # Update entry fonts
            if hasattr(self, 'email_entry'):
                self.email_entry.config(font=("Arial", font_size))
            if hasattr(self, 'password_entry'):
                self.password_entry.config(font=("Arial", font_size))
            if hasattr(self, 'media_entry'):
                self.media_entry.config(font=("Arial", font_size))
            if hasattr(self, 'exec_email_combo'):
                self.exec_email_combo.config(font=("Arial", font_size))
            
            # Update text area font
            if hasattr(self, 'prompt_text'):
                self.prompt_text.config(font=("Arial", font_size))

            # Update style fonts dynamically for notebook tabs and buttons
            try:
                self.style.configure('TNotebook.Tab', font=tab_font)
                self.style.configure('Accent.TButton', font=btn_font)
                self.style.configure('Primary.TButton', font=btn_font)
            except Exception:
                pass
                
        except Exception:
            pass  # Ignore font adjustment errors

    # ===================== Helpers =====================
    def _on_method_change(self) -> None:
        method = self.login_method.get()
        if method == "browser":
            try:
                self.password_label.grid_remove()
                self.password_entry.grid_remove()
                self.note_label.config(text="Browser Login sẽ mở Chrome để đăng nhập Google Flow, hỗ trợ 2FA")
            except Exception:
                pass
        else:
            row = 1
            try:
                self.password_label.grid(row=row, column=0, sticky=tk.W, pady=5)
                self.password_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5)
                self.note_label.config(text="Nhập mật khẩu Google của bạn (có thể cần 2FA)")
            except Exception:
                pass

    def _load_profiles(self) -> None:
        try:
            os.makedirs(os.path.dirname(self.flow_profiles_path), exist_ok=True)
            if os.path.exists(self.flow_profiles_path):
                with open(self.flow_profiles_path, 'r', encoding='utf-8') as f:
                    self.flow_profiles = json.load(f)
            else:
                self.flow_profiles = {}
        except Exception:
            self.flow_profiles = {}

    def _save_profiles(self) -> None:
        try:
            os.makedirs(os.path.dirname(self.flow_profiles_path), exist_ok=True)
            with open(self.flow_profiles_path, 'w', encoding='utf-8') as f:
                json.dump(self.flow_profiles, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _refresh_profiles_list(self) -> None:
        try:
            self.profiles_list.delete(0, tk.END)
            items = sorted(self.flow_profiles.items(), key=lambda kv: kv[1].get("last_login", 0), reverse=True)
            for email_addr, meta in items:
                ts = meta.get("last_login")
                time_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts)) if ts else ""
                self.profiles_list.insert(tk.END, f"{email_addr}  |  {time_str}")
        except Exception:
            pass

    def _refresh_exec_emails(self) -> None:
        try:
            emails = list(self.flow_profiles.keys())
            self.exec_email_combo['values'] = emails
            if emails and not self.exec_email.get():
                self.exec_email.set(emails[0])
        except Exception:
            pass

    # ===================== Login Flow =====================
    def _login_flow(self) -> None:
        email = (self.email_entry.get() or "").strip()
        if not email:
            messagebox.showerror("Lỗi", "Vui lòng nhập email!")
            return
        self.current_email = email
        password = (self.password_entry.get() or "").strip()
        if not password:
            messagebox.showerror("Lỗi", "Vui lòng nhập mật khẩu!")
            return
        threading.Thread(target=self._login_flow_password_thread, args=(email, password), daemon=True).start()

    def _build_chrome(self, cache_key: str, existing_cache_dir: str = None) -> webdriver.Chrome:
        chrome_options = Options()
        # Ưu tiên dùng cache cũ nếu có
        cache_dir = None
        if existing_cache_dir and os.path.isdir(existing_cache_dir):
            cache_dir = existing_cache_dir
        else:
            safe_key = re.sub(r'[^a-zA-Z0-9_.-]', '_', cache_key) or "default"
            cache_dir = os.path.join(os.getcwd(), "chrome_cache", f"flow_{safe_key}")
            os.makedirs(cache_dir, exist_ok=True)
        self.current_cache_dir = cache_dir
        # Nếu profile đang bị một Chrome khác giữ, báo lỗi rõ ràng
        try:
            lock_file = os.path.join(cache_dir, "SingletonLock")
            if os.path.exists(lock_file):
                raise Exception("Profile cache đang được sử dụng bởi một phiên Chrome khác. Hãy đóng trình duyệt đang mở bằng cache này rồi thử lại.")
        except Exception as _:
            # Nếu raise ở trên thì sẽ bị catch ở try/catch phía trên caller
            pass
        chrome_options.add_argument(f"--user-data-dir={cache_dir}")
        chrome_options.add_argument("--profile-directory=Default")
        ua = random.choice(self.user_agents)
        self.current_user_agent = ua
        chrome_options.add_argument(f"--user-agent={ua}")
        chrome_options.add_argument("--lang=vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7")
        chrome_options.add_argument("--start-maximized")
        # Đặt window-size để giảm lỗi DevToolsActivePort khi khởi tạo
        try:
            chrome_options.add_argument("--window-size=1920,1080")
        except Exception:
            pass
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_experimental_option("detach", True)
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        
        # Luôn chạy ở chế độ hiển thị (headless = False)
        # Bỏ qua mọi cấu hình headless trước đây

        # Bật performance logging để đọc Network logs
        try:
            chrome_options.set_capability('goog:loggingPrefs', {'performance': 'ALL'})
        except Exception:
            pass
        service = Service(ChromeDriverManager().install())

        # Retry logic khi Chrome crash/không tạo được session
        max_retries = 3
        for attempt in range(max_retries):
            try:
                driver = webdriver.Chrome(service=service, options=chrome_options)
                try:
                    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
                except Exception:
                    pass
                # Bật Network domain để có thể lấy response body qua CDP
                try:
                    driver.execute_cdp_cmd('Network.enable', {})
                except Exception:
                    pass
                return driver
            except Exception as e:
                msg = str(e).lower()
                crash_like = (
                    "chrome failed to start" in msg or
                    "crashed" in msg or
                    "session not created" in msg or
                    "devtoolsactiveport" in msg
                )
                if crash_like and attempt < max_retries - 1:
                    try:
                        self._log_exec(f"Chrome crash detected, thử lại... (lần {attempt + 1}/{max_retries})")
                    except Exception:
                        pass
                    # Chỉ dọn dẹp lock trong cache profile hiện tại, tránh đóng các Chrome khác
                    try:
                        self._cleanup_profile_locks(self.current_cache_dir)
                    except Exception:
                        pass
                    time.sleep(2)
                    continue
                # Hết retry hoặc lỗi không thuộc dạng crash -> ném lại lỗi
                raise

    def _login_flow_browser_thread(self, email_addr: str) -> None:
        try:
            self._set_status("Đang mở trình duyệt...", "orange")
            self.login_btn.config(state="disabled")
            meta = self.flow_profiles.get(email_addr)
            exist_dir = meta.get("cache_dir") if meta else None
            self.driver = self._build_chrome(email_addr, existing_cache_dir=exist_dir)

            # Go to Flow
            self.driver.get("https://labs.google/fx/vi/tools/flow")
            self._human_delay(2, 4)

            # If not signed in, go through Google sign-in
            if self._flow_requires_login(self.driver):
                self._set_status("Yêu cầu đăng nhập Google - đang chuyển hướng...", "orange")
                self._trigger_flow_login(self.driver)
                self._human_delay(2, 4)
                self._handle_google_login(self.driver, email_addr)

            # Wait for Flow home after sign-in
            ok = self._wait_until(lambda: "labs.google" in self.driver.current_url, timeout=120)
            if not ok:
                raise Exception("Hết thời gian chờ vào Google Flow")

            # If Flow presents a final "Sign in with Google" gate, click it
            try:
                self._click_flow_google_signin(self.driver)
            except Exception:
                pass

            self._remember_profile(email_addr, self.current_cache_dir, self.current_user_agent)
            self.login_success = True
            self._set_status("Đăng nhập thành công Google Flow", "green")
            messagebox.showinfo("Thành công", "Đăng nhập Google Flow thành công!")
            # Đóng trình duyệt sau khi đăng nhập thành công
            try:
                if self.driver is not None:
                    self.driver.quit()
            except Exception:
                pass
            self.driver = None
        except Exception as ex:
            self.login_success = False
            self._set_status("Đăng nhập thất bại", "red")
            messagebox.showerror("Lỗi", f"Lỗi đăng nhập: {ex}")
        finally:
            self.login_btn.config(state="normal")

    def _login_flow_password_thread(self, email_addr: str, password: str) -> None:
        try:
            # Password flow still uses browser to complete 2FA if needed
            self._set_status("Đang mở trình duyệt...", "orange")
            self.login_btn.config(state="disabled")
            meta = self.flow_profiles.get(email_addr)
            exist_dir = meta.get("cache_dir") if meta else None
            self.driver = self._build_chrome(email_addr, existing_cache_dir=exist_dir)
            self.driver.get("https://accounts.google.com/signin")
            self._human_delay(2, 4)
            try:
                self._human_warm_up_page(self.driver)
            except Exception:
                pass
            self._google_type_email_then_password(self.driver, email_addr, password)

            ok = self._wait_signin_success(self.driver, timeout=180)
            if not ok:
                # Retry via AccountChooser (helps with "This browser or app may not be secure")
                try:
                    self._set_status("Thử lại đăng nhập qua AccountChooser...", "orange")
                    self.driver.get("https://accounts.google.com/AccountChooser?continue=https://labs.google/fx/vi/tools/flow")
                    self._human_delay(1, 2)
                    try:
                        self._human_warm_up_page(self.driver)
                    except Exception:
                        pass
                    self._google_type_email_then_password(self.driver, email_addr, password)
                    ok = self._wait_signin_success(self.driver, timeout=180)
                except Exception:
                    ok = False
                if not ok:
                    raise Exception("Không thể đăng nhập tài khoản Google")

            # After Google is signed in, open Flow
            self.driver.get("https://labs.google/fx/vi/tools/flow")
            self._wait_until(lambda: "labs.google" in self.driver.current_url, timeout=120)

            # If Flow presents a final "Sign in with Google" gate, click it
            try:
                self._click_flow_google_signin(self.driver)
            except Exception:
                pass

            self._remember_profile(email_addr, self.current_cache_dir, self.current_user_agent)
            self.login_success = True
            self._set_status("Đăng nhập thành công Google Flow", "green")
            messagebox.showinfo("Thành công", "Đăng nhập Google Flow thành công!")
            # Đóng trình duyệt sau khi đăng nhập thành công
            try:
                if self.driver is not None:
                    self.driver.quit()
            except Exception:
                pass
            self.driver = None
        except Exception as ex:
            self.login_success = False
            self._set_status("Đăng nhập thất bại", "red")
            messagebox.showerror("Lỗi", f"Lỗi đăng nhập: {ex}")
        finally:
            self.login_btn.config(state="normal")

    # ===================== Execute Media =====================
    def _import_excel_and_dispatch(self) -> None:
        try:
            path = filedialog.askopenfilename(title="Chọn file Excel", filetypes=[("Excel", "*.xlsx")])
            if not path:
                return
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            ws = wb.active
            rows = []
            # Safely convert any Excel cell value to trimmed string
            def _cell_to_str(val):
                try:
                    if val is None:
                        return ''
                    return str(val).strip()
                except Exception:
                    return ''
            # Expected columns now: workflow, prompt, media, aspect_ratio, outputs_per_prompt, model (header optional)
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if row is None:
                    continue
                # Skip header if first row contains strings like 'workflow'
                if i == 0 and row and isinstance(row[0], str) and 'workflow' in row[0].lower():
                    continue
                wf = _cell_to_str(row[0]) if len(row) > 0 else ''
                prompt = _cell_to_str(row[1]) if len(row) > 1 else ''
                media = _cell_to_str(row[2]) if len(row) > 2 else ''
                aspect_ratio = _cell_to_str(row[3]) if len(row) > 3 else ''
                outputs = _cell_to_str(row[4]) if len(row) > 4 else ''
                model = _cell_to_str(row[5]) if len(row) > 5 else ''
                # Skip rows with empty prompts (filter out empty records)
                if wf and prompt.strip():
                    rows.append({"wf": wf, "prompt": prompt, "media": media, "aspect_ratio": aspect_ratio, "outputs": outputs, "model": model})
            if not rows:
                messagebox.showerror("Lỗi", "Không có dữ liệu hợp lệ trong file Excel!")
                return
            # Determine available accounts from profiles
            available_emails = [e for e in self.flow_profiles.keys()]
            if not available_emails:
                messagebox.showerror("Lỗi", "Chưa có account nào trong cache!")
                return
            # Round-robin distribute tasks across accounts
            idx = 0
            for r in rows:
                target_email = available_emails[idx % len(available_emails)]
                idx += 1
                meta = self.flow_profiles.get(target_email)
                if not meta:
                    continue
                wf = r.get('wf') or 'frames_to_video'
                # Debug log để kiểm tra workflow value
                self._log_exec(f"Processing workflow: '{wf}' from import file")
                prompt = r.get('prompt') or ''
                media = r.get('media') or ''
                settings = {
                    'aspect_ratio': r.get('aspect_ratio') or '',
                    'outputs': r.get('outputs') or '',
                    'model': r.get('model') or ''
                }
                # Reuse existing enqueue/start logic per-account, passing settings
                self._enqueue_or_start_account_job(target_email, meta, wf, prompt, media, settings)
            self._log_exec(f"Imported {len(rows)} row(s) from Excel and dispatched to accounts.")
            self._refresh_jobs_view()
        except Exception as ex:
            messagebox.showerror("Lỗi", f"Không thể import Excel: {ex}")

    def _import_excel_and_switch_tab(self, excel_path: str) -> None:
        """Import Excel file and switch to execute tab (called from story tab)"""
        try:
            # Import the Excel file using existing logic
            wb = load_workbook(filename=excel_path, read_only=True, data_only=True)
            ws = wb.active
            rows = []
            # Safely convert any Excel cell value to trimmed string
            def _cell_to_str(val):
                try:
                    if val is None:
                        return ''
                    return str(val).strip()
                except Exception:
                    return ''
            # Expected columns now: workflow, prompt, media, aspect_ratio, outputs_per_prompt, model (header optional)
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if row is None:
                    continue
                # Skip header if first row contains strings like 'workflow'
                if i == 0 and row and isinstance(row[0], str) and 'workflow' in row[0].lower():
                    continue
                wf = _cell_to_str(row[0]) if len(row) > 0 else ''
                prompt = _cell_to_str(row[1]) if len(row) > 1 else ''
                media = _cell_to_str(row[2]) if len(row) > 2 else ''
                aspect_ratio = _cell_to_str(row[3]) if len(row) > 3 else ''
                outputs = _cell_to_str(row[4]) if len(row) > 4 else ''
                model = _cell_to_str(row[5]) if len(row) > 5 else ''
                # Skip rows with empty prompts (filter out empty records)
                if wf and prompt.strip():
                    rows.append({"wf": wf, "prompt": prompt, "media": media, "aspect_ratio": aspect_ratio, "outputs": outputs, "model": model})
            if not rows:
                messagebox.showerror("Lỗi", "Không có dữ liệu hợp lệ trong file Excel!")
                return
            # Determine available accounts from profiles
            available_emails = [e for e in self.flow_profiles.keys()]
            if not available_emails:
                messagebox.showerror("Lỗi", "Chưa có account nào trong cache!")
                return
            # Round-robin distribute tasks across accounts
            idx = 0
            for r in rows:
                target_email = available_emails[idx % len(available_emails)]
                idx += 1
                meta = self.flow_profiles.get(target_email)
                if not meta:
                    continue
                wf = r.get('wf') or 'frames_to_video'
                # Debug log để kiểm tra workflow value
                self._log_exec(f"Processing workflow: '{wf}' from import file")
                prompt = r.get('prompt') or ''
                media = r.get('media') or ''
                settings = {
                    'aspect_ratio': r.get('aspect_ratio') or '',
                    'outputs': r.get('outputs') or '',
                    'model': r.get('model') or ''
                }
                # Reuse existing enqueue/start logic per-account, passing settings
                self._enqueue_or_start_account_job(target_email, meta, wf, prompt, media, settings)
            
            self._log_exec(f"Imported {len(rows)} row(s) from Excel and dispatched to accounts.")
            self._refresh_jobs_view()
            
            # Switch to execute tab
            if hasattr(self, 'notebook'):
                self.notebook.select(1)  # Switch to execute tab (index 1)
            
            # Clean up temporary file
            try:
                os.unlink(excel_path)
            except Exception:
                pass
                
        except Exception as ex:
            messagebox.showerror("Lỗi", f"Không thể import Excel: {ex}")

    def _download_excel_template(self) -> None:
        try:
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile="flow_template.xlsx",
                title="Lưu template Excel"
            )
            if not path:
                return
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Tasks"
            # Pull options from current UI to ensure valid values
            ar_values = list(self.aspect_ratio.cget('values')) if hasattr(self, 'aspect_ratio') else ["16:9", "9:16"]
            op_values = list(self.outputs_per_prompt.cget('values')) if hasattr(self, 'outputs_per_prompt') else ["1", "2", "3", "4"]
            md_values = list(self.model_choice.cget('values')) if hasattr(self, 'model_choice') else ["Veo 3 - Fast", "Veo 2 - Fast", "Veo 3 - Quality", "Veo 2 - Quality"]

            ws.append(["workflow", "prompt", "media", "aspect_ratio", "outputs_per_prompt", "model"])  # header
            # three sample rows using valid options
            ws.append(["frames_to_video", "Running", "C:\\Users\\admin\\Downloads\\Shop-quan-ao-nu-quan-9-Fs-store.jpg", ar_values[0] if len(ar_values) > 1 else ar_values[0], op_values[0], md_values[-1]])
            ws.append(["text_to_video", "A cinematic sunset over mountains", "", ar_values[0], op_values[-1], md_values[0]])
            ws.append(["text_to_video", "A neon-lit cyberpunk city at night", "", ar_values[2] if len(ar_values) > 2 else ar_values[0], op_values[1] if len(op_values) > 1 else op_values[0], md_values[1] if len(md_values) > 1 else md_values[0]])
            wb.save(path)
            try:
                self._log_exec(f"Đã lưu template Excel: {path}")
            except Exception:
                pass
        except Exception as ex:
            messagebox.showerror("Lỗi", f"Không thể tạo template: {ex}")

    def _enqueue_or_start_account_job(self, email_addr: str, meta: dict, wf: str, prompt: str, media: str, settings: dict = None) -> None:
        job = {"email": email_addr, "meta": meta, "wf": wf, "prompt": prompt, "media": media, "settings": settings or {}}
        st = self.account_states.get(email_addr)
        if st is None:
            st = {'queue': [], 'running': False, 'lock': threading.Lock()}
            self.account_states[email_addr] = st
        with st['lock']:
            if st['running']:
                st['queue'].append(job)
                action = "queue"
            else:
                st['running'] = True
                action = "start"
        if action == "queue":
            self._log_exec(f"Queued job for {email_addr} ({wf}) from Excel")
        else:
            threading.Thread(target=self._execute_thread, args=(email_addr, meta, wf, prompt, media, job['settings']), daemon=True).start()
    def _open_flow_for_exec(self) -> None:
        email_addr = self.exec_email.get()
        if not email_addr:
            messagebox.showerror("Lỗi", "Vui lòng chọn email đã có cache!")
            return
        meta = self.flow_profiles.get(email_addr)
        if not meta:
            messagebox.showerror("Lỗi", "Không tìm thấy cache cho email đã chọn!")
            return
        threading.Thread(target=self._open_profile_thread, args=(email_addr, meta, True), daemon=True).start()

    def _open_profile_thread(self, email_addr: str, meta: dict, go_flow: bool) -> None:
        try:
            drv = self._open_profile_driver(meta)
            if go_flow:
                drv.get("https://labs.google/fx/vi/tools/flow")
                self._wait_until(lambda: "labs.google" in drv.current_url, timeout=120)
            self._set_exec_status(f"Đã mở Flow cho {email_addr}", "green")
        except Exception as ex:
            self._set_exec_status(f"Lỗi mở Flow: {ex}", "red")

    def _execute_workflow(self) -> None:
        email_addr = self.exec_email.get()
        if not email_addr:
            messagebox.showerror("Lỗi", "Vui lòng chọn email!")
            return
        meta = self.flow_profiles.get(email_addr)
        if not meta:
            messagebox.showerror("Lỗi", "Không tìm thấy cache cho email đã chọn!")
            return
        prompt = self.prompt_text.get("1.0", tk.END).strip()
        media = (self.media_paths.get() or "").strip()
        wf = self.workflow.get()
        # Validate inputs
        if not prompt:
            messagebox.showerror("Lỗi", "Vui lòng nhập Prompt!")
            return
        if wf == "frames_to_video" and not media:
            messagebox.showerror("Lỗi", "Workflow 'Frames to Video' yêu cầu chọn 1 ảnh!")
            return
        # Build a job
        job = {"email": email_addr, "meta": meta, "wf": wf, "prompt": prompt, "media": media}
        # Per-account queueing and start decision
        st = self.account_states.get(email_addr)
        if st is None:
            st = {'queue': [], 'running': False, 'lock': threading.Lock()}
            self.account_states[email_addr] = st
        with st['lock']:
            if st['running']:
                st['queue'].append(job)
                action = "queue"
            else:
                st['running'] = True
                action = "start"
        if action == "queue":
            self._log_exec(f"Queued job for {email_addr} ({wf})")
            self._refresh_jobs_view()
        else:
            threading.Thread(target=self._execute_thread, args=(email_addr, meta, wf, prompt, media, {}), daemon=True).start()

    def _execute_thread(self, email_addr: str, meta: dict, wf: str, prompt: str, media: str, settings: dict) -> None:
        try:
            # Increment job counter for unique indexing
            self.job_counter += 1
            current_job_index = self.job_counter
            self._log_exec(f"Starting job #{current_job_index} for {email_addr}")
            
            self._log_exec("Opening Flow page...")
            drv = self._open_profile_driver(meta)
            # Track driver per account to allow concurrent runs
            self.exec_drivers[email_addr] = drv
            self.stop_exec = False
            # Store current prompt for folder naming
            self.current_prompt = prompt
            # Store current job index for download naming
            self.current_job_index = current_job_index
            # set current job and refresh view
            try:
                # Track current job per account for UI
                self.exec_current_jobs[email_addr] = {"email": email_addr, "wf": wf, "prompt": prompt, "media": media}
                self._refresh_jobs_view()
            except Exception:
                pass
            wait = WebDriverWait(drv, 30)
            drv.get("https://labs.google/fx/vi/tools/flow")
            loaded = self._wait_until(lambda: "labs.google" in drv.current_url, timeout=120)
            if not loaded:
                self._log_exec("Failed to load Flow page", error=True)
                return
            time.sleep(5)

            # Choose workflow
            self._log_exec("Click New project (if visible)...")
            try:
                ready = self._ensure_workspace_ready(drv, max_attempts=3, wait_seconds=10)
                if not ready:
                    self._log_exec("Workspace not detected after retries. Continuing best-effort...", error=False)
            except Exception:
                self._log_exec("New project button not found - continue")

            # Mở combobox chọn workflow và chọn theo wf TRƯỚC
            try:
                self._log_exec("Opening workflow combobox and selecting option...")
                self._select_workflow_via_combobox(drv, wf, media)
                self._log_exec("Workflow selected.")
            except Exception:
                self._log_exec("Failed to select workflow via combobox", error=True)

            # Sau khi chọn workflow, mới nhập prompt
            self._log_exec("Typing prompt into textarea...")
            # Gõ prompt vào textarea id=PINHOLE_TEXT_AREA_ELEMENT_ID với tốc độ tối ưu
            try:
                area = drv.find_element(By.ID, "PINHOLE_TEXT_AREA_ELEMENT_ID")
                self._human_click_el(drv, area)
                try:
                    area.clear()
                except Exception:
                    pass
                if prompt:
                    # Sử dụng typing nhanh nhất cho prompt
                    self._fast_type_prompt(area, prompt)
                    self._log_exec("Prompt typed successfully.")
                else:
                    self._log_exec("No prompt provided, skipping typing.")
            except Exception:
                # Fallback: bất kỳ textarea nào nếu id không có - sử dụng typing nhanh
                self._log_exec("PINHOLE_TEXT_AREA_ELEMENT_ID not found, trying fallback textarea...")
                self._fast_type_into_any(drv, [
                    (By.ID, "PINHOLE_TEXT_AREA_ELEMENT_ID"),
                    (By.CSS_SELECTOR, "textarea#PINHOLE_TEXT_AREA_ELEMENT_ID"),
                    (By.CSS_SELECTOR, "textarea")
                ], prompt)

            # Áp dụng các setting trong popover (Aspect ratio, Outputs per prompt, Model)
            try:
                self._log_exec("Opening settings popover and applying options...")
                # Prefer per-row settings if provided; fallback to UI selections
                ar = (settings.get('aspect_ratio') or '').strip() if isinstance(settings, dict) else ''
                op = (settings.get('outputs') or '').strip() if isinstance(settings, dict) else ''
                md = (settings.get('model') or '').strip() if isinstance(settings, dict) else ''
                aspect_to_use = ar if ar else self.aspect_ratio.get()
                outputs_to_use = op if op else self.outputs_per_prompt.get()
                model_to_use = md if md else self.model_choice.get()
                
                # Update UI comboboxes to reflect the values being used
                if ar:
                    self.aspect_ratio.set(aspect_to_use)
                if op:
                    self.outputs_per_prompt.set(outputs_to_use)
                if md:
                    self.model_choice.set(model_to_use)
                
                self._open_settings_and_apply(drv, aspect_to_use, outputs_to_use, model_to_use)
                self._log_exec("Settings applied.")
            except Exception as ex:
                self._log_exec(f"Failed to apply settings: {ex}", error=True)

            # Với workflow text_to_video: cần nhấn nút 'Tạo' (Create) để execute
            if wf == "text_to_video":
                try:
                    self._log_exec("Clicking 'Tạo' (Create) button for text_to_video...")
                    self._click_create_button(drv)
                    self._log_exec("Clicked 'Tạo' successfully.")
                    # Chờ một chút trước khi monitor
                    self._log_exec("Waiting 5s before monitoring processing...")
                    time.sleep(5)
                    self._log_exec("Monitoring processing and then reading API logs...")
                    self._monitor_and_fetch_api(drv, wf="text_to_video")
                    return
                except Exception as ex:
                    self._log_exec(f"Failed to click Create button for text_to_video: {ex}", error=True)

            # Apply config if UI exposes inputs (best-effort)
            # (Removed) basic config for resolution/duration/fps per user request

            # Upload media: chỉ áp dụng cho frames_to_video
            if wf == "frames_to_video" and media:
                try:
                    self._log_exec("Waiting 3s after workflow selection before opening add panel...")
                    time.sleep(3)
                    self._log_exec("Opening frames upload panel...")
                    self._open_frames_upload_panel(drv)
                except Exception:
                    self._log_exec("Could not open frames upload panel (will still try upload)")
                self._log_exec("Uploading media (frames)...")
                self._upload_media_any(drv, media)
                self._log_exec("Upload step finished (best-effort).")
                # Nhấn "Cắt và lưu" và đợi đến khi xuất hiện khung hình đầu tiên
                try:
                    self._log_exec("Clicking 'Cắt và lưu' and waiting for first frame...")
                    self._confirm_crop_and_wait_first_frame(drv)
                    self._log_exec("First frame detected.")
                    # Sau khi có khung hình đầu tiên, nhấn nút Tạo (Create)
                    self._log_exec("Clicking 'Tạo' (Create) button...")
                    self._click_create_button(drv)
                    self._log_exec("Clicked 'Tạo' successfully.")
                    # Chờ 10s rồi theo dõi tiến trình và tải kết quả
                    self._log_exec("Waiting 100s before monitoring processing...")
                    time.sleep(100)
                    self._log_exec("Monitoring processing and then reading API logs...")
                    self._monitor_and_fetch_api(drv, wf="frames_to_video")
                    # ĐÃ HOÀN TẤT Frames to Video: return ngay để không chạy các bước Submit/Monitor chung bên dưới
                    return
                except Exception:
                    self._log_exec("Could not confirm crop/save or detect first frame", error=True)

            # Execute/Run
            self._log_exec("Submitting job...")
            self._try_click_any(drv, [
                "//button[contains(., 'Run')]",
                "//button[contains(., 'Generate')]",
                "//button[contains(., 'Create')]",
                "//*[contains(text(), 'Run')]",
                "//*[contains(text(), 'Generate')]",
            ])

            self._log_exec("Request submitted. Monitoring for completion...", success=False)
            # For all workflows, monitor until outputs ready, then read API
            try:
                # small delay to allow rendering to start
                time.sleep(5)
                self._monitor_and_fetch_api(drv, wf=wf)
            except Exception as ex:
                self._log_exec(f"Monitor error: {ex}", error=True)
        except Exception as ex:
            self._log_exec(f"Execute error: {ex}", error=True)
        finally:
            try:
                local_drv = self.exec_drivers.get(email_addr)
                if local_drv is not None:
                    local_drv.quit()
            except Exception:
                pass
            try:
                if email_addr in self.exec_drivers:
                    del self.exec_drivers[email_addr]
            except Exception:
                pass
            self.stop_exec = False
            # clear current job and refresh
            try:
                if email_addr in self.exec_current_jobs:
                    del self.exec_current_jobs[email_addr]
                self._refresh_jobs_view()
            except Exception:
                pass

            # Auto-run next queued job if available
            try:
                st = self.account_states.get(email_addr)
                next_job = None
                if st is None:
                    st = {'queue': [], 'running': False, 'lock': threading.Lock()}
                    self.account_states[email_addr] = st
                with st['lock']:
                    if st['queue']:
                        next_job = st['queue'].pop(0)
                    else:
                        st['running'] = False
                if next_job is not None:
                    self._log_exec(f"Starting next queued job for {next_job['email']} ({next_job['wf']})")
                    threading.Thread(
                        target=self._execute_thread,
                        args=(next_job['email'], next_job['meta'], next_job['wf'], next_job['prompt'], next_job['media'], next_job.get('settings', {})),
                        daemon=True,
                    ).start()
                else:
                    self._log_exec(f"All jobs completed for {email_addr}", success=True)
            except Exception as ex:
                try:
                    st = self.account_states.get(email_addr)
                    if st:
                        with st['lock']:
                            st['running'] = False
                except Exception:
                    pass
                self._log_exec(f"Queue scheduling error: {ex}", error=True)

    def _refresh_jobs_view(self) -> None:
        try:
            # Helper to format row data
            def fmt_row(job):
                try:
                    img = os.path.basename(job.get('media') or '') if job.get('media') else ''
                except Exception:
                    img = ''
                prompt = (job.get('prompt') or '').replace('\n', ' ').strip()
                if len(prompt) > 120:
                    prompt = prompt[:117] + '...'
                return (job.get('email') or '', job.get('wf') or '', img, prompt)

            # Running table
            if hasattr(self, 'running_table'):
                try:
                    for i in self.running_table.get_children():
                        self.running_table.delete(i)
                    if self.exec_current_jobs:
                        jobs = list(self.exec_current_jobs.values())[::-1]
                        for job in jobs:
                            email, wf, img, prompt = fmt_row(job)
                            self.running_table.insert('', 'end', values=(email, wf, img, prompt))
                except Exception:
                    pass

            # Queue table
            if hasattr(self, 'queue_table'):
                try:
                    for i in self.queue_table.get_children():
                        self.queue_table.delete(i)
                    aggregate = []
                    for _, st in self.account_states.items():
                        with st['lock']:
                            aggregate.extend(list(st['queue']))
                    for job in aggregate:
                        email, wf, img, prompt = fmt_row(job)
                        self.queue_table.insert('', 'end', values=(email, wf, img, prompt))
                except Exception:
                    pass
        except Exception:
            pass

    # ===================== Selenium Utils =====================
    def _flow_requires_login(self, driver: webdriver.Chrome) -> bool:
        try:
            # If any sign-in indicator or redirect present
            texts = ["Sign in", "Đăng nhập", "Log in"]
            for xp in [f"//*[contains(text(), '{t}')]" for t in texts]:
                if driver.find_elements(By.XPATH, xp):
                    return True
            # Or Google accounts domain
            url = driver.current_url or ""
            if "accounts.google.com" in url:
                return True
        except Exception:
            pass
        return False

    def _click_new_project(self, driver: webdriver.Chrome) -> None:
        """Nhấn nút tạo dự án mới trên trang Flow (EN/VI)."""
        candidates = [
            "//button[normalize-space()='New project']",
            "//button[contains(., 'New project')]",
            "//a[normalize-space()='New project']",
            "//a[contains(., 'New project')]",
            "//button[normalize-space()='Dự án mới']",
            "//button[contains(., 'Dự án mới')]",
            "//a[normalize-space()='Dự án mới']",
            "//a[contains(., 'Dự án mới')]",
        ]
        for xp in candidates:
            try:
                el = driver.find_element(By.XPATH, xp)
                try:
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                except Exception:
                    pass
                try:
                    self._human_click_el(driver, el)
                except Exception:
                    try:
                        driver.execute_script("arguments[0].click();", el)
                    except Exception:
                        continue
                return
            except Exception:
                continue
    
    def _ensure_workspace_ready(self, driver: webdriver.Chrome, max_attempts: int = 3, wait_seconds: int = 10) -> bool:
        try:
            def page_ready(before_url: str) -> bool:
                try:
                    if driver.current_url != before_url:
                        return True
                    if driver.find_elements(By.ID, "PINHOLE_TEXT_AREA_ELEMENT_ID"):
                        return True
                    if driver.find_elements(By.CSS_SELECTOR, "button[role='combobox']"):
                        return True
                except Exception:
                    pass
                return False

            def new_project_exists() -> bool:
                xps = [
                    "//button[normalize-space()='New project']",
                    "//button[contains(., 'New project')]",
                    "//a[normalize-space()='New project']",
                    "//a[contains(., 'New project')]",
                    "//button[normalize-space()='Dự án mới']",
                    "//button[contains(., 'Dự án mới')]",
                    "//a[normalize-space()='Dự án mới']",
                    "//a[contains(., 'Dự án mới')]",
                ]
                for xp in xps:
                    try:
                        if driver.find_elements(By.XPATH, xp):
                            return True
                    except Exception:
                        continue
                return False

            attempt = 0
            while attempt < max_attempts:
                before_url = driver.current_url
                self._click_new_project(driver)
                self._log_exec(f"Waiting {wait_seconds}s for workspace to appear (attempt {attempt+1}/{max_attempts})...")
                time.sleep(wait_seconds)
                ready = self._wait_until(lambda: page_ready(before_url), timeout=10, interval=0.5)
                if ready:
                    return True
                if new_project_exists():
                    attempt += 1
                    continue
                try:
                    driver.get("https://labs.google/fx/vi/tools/flow")
                except Exception:
                    pass
                time.sleep(3)
                attempt += 1
            return False
        except Exception:
            return False

    def _select_workflow_via_combobox(self, driver: webdriver.Chrome, wf: str, media: str = None) -> None:
        """Mở combobox (button[role="combobox"]) và chọn mục theo wf.
        Luôn tìm đúng text đã mapping rồi click vào đó thay vì dựa vào vị trí.
        Hỗ trợ text tiếng Việt và tiếng Anh trong dropdown Radix.
        Nếu không tìm được text phù hợp:
        - Nếu có media (hình ảnh): chọn option 2
        - Nếu không có media: chọn option 1
        """
        # Debug log để kiểm tra workflow value được truyền vào
        self._log_exec(f"Selecting workflow: '{wf}' in combobox")
        # Click trigger combobox
        trigger = driver.find_element(By.CSS_SELECTOR, "button[role='combobox']")
        self._human_click_el(driver, trigger)
        
        # Đợi 2s sau khi mở combobox như yêu cầu
        self._log_exec("Waiting 2s after opening combobox...")
        time.sleep(2)

        # Đợi content xuất hiện (role=listbox)
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//*[@role='listbox']"))
            )
        except Exception:
            pass

        # Xác định text cần tìm dựa trên wf
        if wf == "text_to_video":
            target_texts = [
                "Từ văn bản sang video",
                "Text to Video",
                "text_to_video",  # Fallback cho trường hợp text khác
            ]
        else:  # frames_to_video hoặc bất kỳ giá trị nào khác
            target_texts = [
                "Tạo video từ các khung hình", 
                "Frames to Video",
                "frames_to_video",  # Fallback cho trường hợp text khác
            ]

        # Tìm và click vào option có text phù hợp
        for target_text in target_texts:
            try:
                # Thử tìm element chứa text chính xác
                el = driver.find_element(By.XPATH, f"//*[@role='listbox']//*[@role='option'][contains(., '{target_text}')]")
                self._human_click_el(driver, el)
                self._log_exec(f"Selected workflow: {target_text}")
                return
            except Exception:
                continue

        # Nếu không tìm thấy, log lỗi và thử click option dựa trên media
        self._log_exec(f"Could not find workflow text for '{wf}', using fallback logic", error=True)
        try:
            options = driver.find_elements(By.XPATH, "//*[@role='listbox']//*[@role='option']")
            if options:
                # Kiểm tra xem có media (hình ảnh) hay không
                has_media = media and media.strip() and os.path.isfile(media)
                
                if has_media:
                    # Nếu có media: chọn option 2 (index 1)
                    if len(options) > 1:
                        self._human_click_el(driver, options[1])
                        self._log_exec("Selected option 2 (index 1) as fallback - has media")
                    else:
                        # Nếu chỉ có 1 option, chọn option đó
                        self._human_click_el(driver, options[0])
                        self._log_exec("Selected only available option as fallback - has media")
                else:
                    # Nếu không có media: chọn option 1 (index 0)
                    self._human_click_el(driver, options[0])
                    self._log_exec("Selected option 1 (index 0) as fallback - no media")
        except Exception:
            self._log_exec("Failed to select any workflow option", error=True)

    def _trigger_flow_login(self, driver: webdriver.Chrome) -> None:
        # Try clicking login button on Flow if available
        self._try_click_any(driver, [
            "//button[contains(., 'Sign in')]",
            "//a[contains(., 'Sign in')]",
            "//*[contains(text(), 'Sign in')]",
        ])

    def _click_flow_google_signin(self, driver: webdriver.Chrome) -> None:
        """Nhấn nút "Sign in with Google" trên trang Flow nếu có.
        Hỗ trợ cả tiếng Việt nếu UI được bản địa hóa.
        """
        candidates = [
            "//button[.//span[normalize-space()='Sign in with Google']]",
            "//button[contains(., 'Sign in with Google')]",
            "//button[contains(@class,'sc-') and contains(., 'Sign in with Google')]",
            "//button[.//span[normalize-space()='Đăng nhập bằng Google']]",
            "//button[contains(., 'Đăng nhập bằng Google')]",
        ]
        # Đợi trang ổn định một chút
        self._human_delay(0.5, 1.5)
        for xp in candidates:
            try:
                el = driver.find_element(By.XPATH, xp)
                # Đảm bảo phần tử có thể click
                try:
                    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xp)))
                except Exception:
                    pass
                self._human_click_el(driver, el)
                time.sleep(0.5)
                return
            except Exception:
                continue

    def _handle_google_login(self, driver: webdriver.Chrome, email_addr: str) -> None:
        try:
            wait = WebDriverWait(driver, 30)
            wait.until(EC.presence_of_element_located((By.NAME, "identifier")))
            email_input = driver.find_element(By.NAME, "identifier")
            self._human_click_el(driver, email_input)
            email_input.clear()
            self._human_type_el(email_input, email_addr)
            driver.find_element(By.ID, "identifierNext").click()
            self._human_delay(2, 4)
            # Let user finish password/2FA manually
            messagebox.showinfo("Hướng dẫn", "Vui lòng hoàn tất nhập mật khẩu và xác thực 2FA (nếu có) trong trình duyệt mở.")
            # Thử đóng hộp thoại Passkey nếu xuất hiện
            try:
                self._dismiss_passkey_prompt(driver)
            except Exception:
                pass
            self._wait_signin_success(driver, timeout=300)
        except Exception:
            pass

    def _google_type_email_then_password(self, driver: webdriver.Chrome, email_addr: str, password: str) -> None:
        wait = WebDriverWait(driver, 30)
        wait.until(EC.presence_of_element_located((By.NAME, "identifier")))
        email_input = driver.find_element(By.NAME, "identifier")
        self._human_click_el(driver, email_input)
        email_input.clear()
        self._human_type_el(email_input, email_addr)
        driver.find_element(By.ID, "identifierNext").click()
        self._human_delay(2, 4)
        try:
            wait.until(EC.presence_of_element_located((By.NAME, "Passwd")))
            pw_input = driver.find_element(By.NAME, "Passwd")
        except TimeoutException:
            pw_input = driver.find_element(By.CSS_SELECTOR, "input[type='password']")
        self._human_click_el(driver, pw_input)
        pw_input.clear()
        self._human_type_el(pw_input, password)
        self._human_delay(0.5, 1.5)
        try:
            driver.find_element(By.ID, "passwordNext").click()
        except Exception:
            pw_input.send_keys(Keys.ENTER)
        # Sau khi submit mật khẩu, cố gắng đóng Passkey prompt nếu có
        try:
            self._human_delay(1, 2)
            self._dismiss_passkey_prompt(driver)
        except Exception:
            pass

    def _wait_signin_success(self, driver: webdriver.Chrome, timeout: int = 180) -> bool:
        start = time.time()
        while time.time() - start < timeout:
            try:
                # Thử đóng Passkey prompt nếu xuất hiện trong lúc chờ
                try:
                    self._dismiss_passkey_prompt(driver)
                except Exception:
                    pass
                if self._is_google_signed_in(driver):
                    return True
            except Exception:
                pass
            time.sleep(1)
        return False

    def _dismiss_passkey_prompt(self, driver: webdriver.Chrome) -> None:
        """Đóng hộp thoại gợi ý thêm Passkey bằng cách bấm Not now/No thanks/Bỏ qua."""
        try:
            candidates = [
                "//button[normalize-space()='Not now']",
                "//button[contains(., 'Not now')]",
                "//div[@role='dialog']//button[normalize-space()='Not now']",
                "//span[normalize-space()='Not now']/parent::button",
                "//button[normalize-space()='No thanks']",
                "//button[contains(., 'No thanks')]",
                "//span[normalize-space()='No thanks']/parent::button",
                "//button[normalize-space()='Skip']",
                "//button[contains(., 'Skip')]",
                "//span[normalize-space()='Skip']/parent::button",
                # Vietnamese fallbacks
                "//button[contains(., 'Không phải bây giờ')]",
                "//button[contains(., 'Bỏ qua')]",
            ]
            for xp in candidates:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    self._human_click_el(driver, el)
                    time.sleep(0.5)
                    return
                except Exception:
                    continue
        except Exception:
            pass

    def _kill_chrome_processes(self) -> None:
        """Đóng mọi tiến trình Chrome/ChromeDriver còn treo để chuẩn bị retry."""
        try:
            import subprocess
            import platform
            system = platform.system().lower()
            if system == "windows":
                try:
                    subprocess.run(["taskkill", "/f", "/im", "chrome.exe"], capture_output=True, timeout=5)
                except Exception:
                    pass
                try:
                    subprocess.run(["taskkill", "/f", "/im", "chromedriver.exe"], capture_output=True, timeout=5)
                except Exception:
                    pass
            elif system in ["linux", "darwin"]:
                try:
                    subprocess.run(["pkill", "-f", "chrome"], capture_output=True, timeout=5)
                except Exception:
                    pass
                try:
                    subprocess.run(["pkill", "-f", "chromedriver"], capture_output=True, timeout=5)
                except Exception:
                    pass
        except Exception:
            pass

    def _cleanup_profile_locks(self, cache_dir: str | None) -> None:
        """Xóa các file lock trong profile cache để giải phóng session của đúng email."""
        try:
            if not cache_dir:
                return
            candidates = [
                os.path.join(cache_dir, "SingletonLock"),
                os.path.join(cache_dir, "SingletonCookie"),
                os.path.join(cache_dir, "SingletonSocket"),
            ]
            for fp in candidates:
                try:
                    if os.path.exists(fp):
                        os.remove(fp)
                except Exception:
                    continue
        except Exception:
            pass

    def _is_google_signed_in(self, driver: webdriver.Chrome) -> bool:
        try:
            url = driver.current_url or ""
            if url.startswith("https://myaccount.google.com/"):
                return True
            selectors = [
                "a[aria-label^='Google Account:' i]",
                "a[aria-label^='Tài khoản Google:' i]",
                "img.gb_P.gbii",
            ]
            for sel in selectors:
                if driver.find_elements(By.CSS_SELECTOR, sel):
                    return True
        except Exception:
            return False
        return False

    def _remember_profile(self, email_addr: str, cache_dir: str, user_agent: str) -> None:
        if not email_addr:
            return
        self.flow_profiles[email_addr] = {
            "cache_dir": cache_dir,
            "user_agent": user_agent,
            "last_login": int(time.time()),
        }
        self._save_profiles()
        self._refresh_profiles_list()
        self._refresh_exec_emails()

    def _open_profile_driver(self, meta: dict) -> webdriver.Chrome:
        chrome_options = Options()
        cache_dir = meta.get("cache_dir")
        if cache_dir and os.path.isdir(cache_dir):
            chrome_options.add_argument(f"--user-data-dir={cache_dir}")
            chrome_options.add_argument("--profile-directory=Default")
        if meta.get("user_agent"):
            chrome_options.add_argument(f"--user-agent={meta['user_agent']}")
        chrome_options.add_argument("--start-maximized")
        # Thiết lập kích thước cửa sổ cố định giúp ổn định khởi tạo
        try:
            chrome_options.add_argument("--window-size=1920,1080")
        except Exception:
            pass
        chrome_options.add_argument("--lang=vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7")
        try:
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        except Exception:
            pass
        
        # Cho phép tùy biến headless cho luồng execute theo self.headless_mode
        try:
            if self.headless_mode.get():
                chrome_options.add_argument("--headless")
                chrome_options.add_argument("--no-sandbox")
                chrome_options.add_argument("--disable-dev-shm-usage")
        except Exception:
            pass
            
        # Bật performance logging để đọc Network logs
        try:
            chrome_options.set_capability('goog:loggingPrefs', {'performance': 'ALL'})
        except Exception:
            pass
        service = Service(ChromeDriverManager().install())

        # Retry logic tương tự whisk khi Chrome crash
        max_retries = 3
        for attempt in range(max_retries):
            try:
                drv = webdriver.Chrome(service=service, options=chrome_options)
                try:
                    drv.execute_cdp_cmd('Network.enable', {})
                except Exception:
                    pass
                return drv
            except Exception as e:
                msg = str(e).lower()
                crash_like = (
                    "chrome failed to start" in msg or
                    "crashed" in msg or
                    "session not created" in msg or
                    "devtoolsactiveport" in msg
                )
                if crash_like and attempt < max_retries - 1:
                    try:
                        self._log_exec(f"Chrome crash detected, đóng browser và thử lại... (lần {attempt + 1}/{max_retries})")
                    except Exception:
                        pass
                    # Dọn dẹp lock trong cache profile này thay vì đóng toàn bộ Chrome
                    try:
                        self._cleanup_profile_locks(cache_dir)
                    except Exception:
                        pass
                    time.sleep(2)
                    continue
                raise
        

    def _open_selected_profile(self) -> None:
        try:
            sel = self.profiles_list.curselection()
            if not sel:
                messagebox.showinfo("Thông báo", "Vui lòng chọn một tài khoản trong danh sách!")
                return
            line = self.profiles_list.get(sel[0])
            email_addr = line.split("  |  ")[0].strip()
            meta = self.flow_profiles.get(email_addr)
            if not meta:
                messagebox.showerror("Lỗi", "Không tìm thấy cache!")
                return
            threading.Thread(target=self._open_profile_thread, args=(email_addr, meta, False), daemon=True).start()
        except Exception as ex:
            messagebox.showerror("Lỗi", f"Không thể mở profile: {ex}")

    def _delete_selected_profile(self) -> None:
        try:
            sel = self.profiles_list.curselection()
            if not sel:
                messagebox.showinfo("Thông báo", "Vui lòng chọn một tài khoản trong danh sách!")
                return
            line = self.profiles_list.get(sel[0])
            email_addr = line.split("  |  ")[0].strip()
            meta = self.flow_profiles.get(email_addr)
            if not meta:
                return
            cache_dir = meta.get("cache_dir")
            try:
                if cache_dir and os.path.isdir(cache_dir):
                    import shutil
                    shutil.rmtree(cache_dir, ignore_errors=True)
            except Exception:
                pass
            self.flow_profiles.pop(email_addr, None)
            self._save_profiles()
            self._refresh_profiles_list()
            self._refresh_exec_emails()
            self._set_status(f"Đã xóa cache của {email_addr}", "green")
        except Exception as ex:
            messagebox.showerror("Lỗi", f"Không thể xóa cache: {ex}")

    # ===================== Generic Actions =====================
    def _choose_image_file(self) -> None:
        """Choose single image file for upload"""
        try:
            file = filedialog.askopenfilename(
                title="Chọn file ảnh",
                filetypes=[
                    ("Image files", "*.jpg *.jpeg *.png *.gif *.bmp *.tiff *.webp"),
                    ("JPEG files", "*.jpg *.jpeg"),
                    ("PNG files", "*.png"),
                    ("All files", "*.*")
                ]
            )
            if file:
                self.media_paths.set(file)
        except Exception as ex:
            messagebox.showerror("Lỗi", f"Không thể chọn file ảnh: {ex}")

    def _apply_basic_config(self, driver: webdriver.Chrome, resolution: str, duration: str, fps: str) -> None:
        # Best-effort: try to locate simple inputs/selects
        try:
            # Resolution dropdown/select
            for xp in [
                "//select[contains(@name, 'resolution')]",
                "//select[contains(@id, 'resolution')]",
                "//label[contains(., 'Resolution')]/following::select[1]",
            ]:
                els = driver.find_elements(By.XPATH, xp)
                if els:
                    try:
                        els[0].click()
                        time.sleep(0.2)
                        opt = driver.find_elements(By.XPATH, f"//option[contains(., '{resolution}')]")
                        if opt:
                            opt[0].click()
                    except Exception:
                        pass

            # Duration
            self._type_into_any(driver, [
                (By.XPATH, "//input[@type='number' and contains(@name, 'duration')]"),
                (By.XPATH, "//input[contains(@placeholder, 'Duration')]")
            ], duration)

            # FPS
            self._type_into_any(driver, [
                (By.XPATH, "//input[@type='number' and contains(@name, 'fps')]"),
                (By.XPATH, "//input[contains(@placeholder, 'FPS')]")
            ], fps)
        except Exception:
            pass

    def _upload_media_any(self, driver: webdriver.Chrome, media: str) -> None:
        # Chỉ upload 1 file ảnh duy nhất
        if not media or not os.path.isfile(media):
            self._log_exec("No valid image file provided", error=True)
            return
            
        # Kiểm tra đuôi file có phải ảnh không
        image_extensions = [".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif", ".tiff"]
        if not any(media.lower().endswith(ext) for ext in image_extensions):
            self._log_exec("File must be an image (png, jpg, jpeg, webp, bmp, gif, tiff)", error=True)
            return
            
        self._log_exec(f"Uploading single image: {os.path.basename(media)}")
        
        # Try visible inputs first
        inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")
        if inputs:
            try:
                inputs[0].send_keys(media)
                self._log_exec("File uploaded successfully via file input")
                return
            except Exception as ex:
                self._log_exec(f"Failed to upload via file input: {ex}", error=True)
                pass

        # Try clicking upload buttons then re-scan inputs
        self._try_click_any(driver, [
            "//button[contains(., 'Upload')]",
            "//button[contains(., 'Add')]",
            "//div[contains(., 'Upload')]",
            "//*[contains(@class, 'upload')]",
        ])
        time.sleep(0.5)
        inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")
        if inputs:
            try:
                if len(files) == 1:
                    inputs[0].send_keys(files[0])
                else:
                    for f in files:
                        inputs[0].send_keys(f)
                return
            except Exception:
                pass

    def _fast_type_into_any(self, driver: webdriver.Chrome, locators, text: str) -> None:
        """Fast version of _type_into_any optimized for prompt input."""
        if not text:
            return
        for by, sel in locators:
            try:
                el = driver.find_element(by, sel)
                self._human_click_el(driver, el)
                try:
                    el.clear()
                except Exception:
                    pass
                self._fast_type_prompt(el, text)
                break
            except Exception:
                continue

    def _type_into_any(self, driver: webdriver.Chrome, locators, text: str) -> None:
        if not text:
            return
        for by, sel in locators:
            try:
                el = driver.find_element(by, sel)
                self._human_click_el(driver, el)
                try:
                    el.clear()
                except Exception:
                    pass
                self._human_type_el(el, text)
                break
            except Exception:
                continue

    def _try_click_any(self, driver: webdriver.Chrome, xpaths) -> None:
        for xp in xpaths:
            try:
                el = driver.find_element(By.XPATH, xp)
                self._human_click_el(driver, el)
                time.sleep(0.3)
                return
            except Exception:
                continue

    def _log_exec(self, message: str, success: bool = False, error: bool = False) -> None:
        """Print to console, update status label, and append to on-screen log. Supports UI callbacks."""
        prefix = "[EXEC]"
        line = f"{prefix} {time.strftime('%H:%M:%S')} | {message}\n"
        print(line, end="")
        # Callback-based logging (for PySide6 adapter)
        try:
            cb = self.ui_callbacks.get('on_log') if hasattr(self, 'ui_callbacks') else None
            if cb:
                try:
                    cb(line)
                except Exception:
                    pass
        except Exception:
            pass
        # Tk UI log append
        try:
            if getattr(self, 'use_tk_ui', True):
                self._append_exec_log(line)
        except Exception:
            pass
        # Status update
        color = "orange"
        if error:
            color = "red"
        elif success:
            color = "green"
        try:
            # Callback for status line (non-Tk UI)
            cb_status = self.ui_callbacks.get('on_exec_status') if hasattr(self, 'ui_callbacks') else None
            if cb_status:
                try:
                    cb_status(message, color)
                except Exception:
                    pass
        except Exception:
            pass

    def _append_exec_log(self, text: str) -> None:
        """Append text to the progress log textbox and auto-scroll to bottom."""
        try:
            if hasattr(self, 'exec_log'):
                self.exec_log.configure(state='normal')
                self.exec_log.insert(tk.END, text)
                self.exec_log.see(tk.END)
                self.exec_log.configure(state='disabled')
        except Exception:
            pass

    # ===================== Error Log Helpers =====================
    def _append_error_log(self, text: str) -> None:
        try:
            if hasattr(self, 'error_log'):
                self.error_log.configure(state='normal')
                self.error_log.insert(tk.END, text)
                self.error_log.see(tk.END)
                self.error_log.configure(state='disabled')
        except Exception:
            pass

    def _load_error_events(self) -> None:
        try:
            if os.path.exists(self.error_log_path):
                with open(self.error_log_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        self.error_events = data
        except Exception:
            self.error_events = []

    def _save_error_events(self) -> None:
        try:
            with open(self.error_log_path, 'w', encoding='utf-8') as f:
                json.dump(self.error_events[-1000:], f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _refresh_error_log_ui(self) -> None:
        try:
            if hasattr(self, 'error_log'):
                self.error_log.configure(state='normal')
                self.error_log.delete('1.0', tk.END)
                for e in self.error_events[-300:]:
                    self.error_log.insert(tk.END, e)
                self.error_log.see(tk.END)
                self.error_log.configure(state='disabled')
        except Exception:
            pass

    def _log_error(self, message: str) -> None:
        ts = time.strftime('%H:%M:%S')
        line = f"[ERROR] {ts} | {message}\n"
        print(line, end="")
        try:
            self.error_events.append(line)
            # Persist asynchronously
            threading.Thread(target=self._save_error_events, daemon=True).start()
        except Exception:
            pass
        try:
            # UI callbacks if any
            cb = self.ui_callbacks.get('on_log') if hasattr(self, 'ui_callbacks') else None
            if cb:
                try:
                    cb(line)
                except Exception:
                    pass
        except Exception:
            pass
        try:
            # Append to error card
            self._append_error_log(line)
            # Also reflect on exec status area in red (non-blocking)
        except Exception:
            pass

    def _open_frames_upload_panel(self, driver: webdriver.Chrome) -> None:
        """Mở panel thêm media cho workflow Frames to Video theo mô tả UI.
        1) Click nút có icon 'add' (có thể là button với overlay)
        2) Click nút 'Tải lên' (icon 'upload') để bắt đầu upload ảnh
        """
        # Bước 1: Nút add
        add_candidates = [
            "//button[.//i[contains(text(),'add')]]",
            "//button[contains(@class,'sc-d6df593a') and .//i[contains(text(),'add')]]",
            "//i[text()='add']/ancestor::button[1]",
            "//button[contains(., 'Add') or contains(., 'Thêm')]",
        ]
        for xp in add_candidates:
            try:
                el = driver.find_element(By.XPATH, xp)
                self._human_click_el(driver, el)
                time.sleep(0.5)
                break
            except Exception:
                continue

        # Bước 2: Nút "Tải lên" (VN) hoặc tương đương với icon 'upload'
        upload_candidates = [
            "//button[.//div[contains(., 'Tải lên')]]",
            "//button[contains(., 'Tải lên')]",
            "//button[.//i[normalize-space(text())='upload']]",
            "//i[normalize-space(text())='upload']/ancestor::button[1]",
            "//button[contains(., 'Upload')]",
        ]
        for xp in upload_candidates:
            try:
                el = driver.find_element(By.XPATH, xp)
                self._human_click_el(driver, el)
                time.sleep(0.5)
                return
            except Exception:
                continue

    def _confirm_crop_and_wait_first_frame(self, driver: webdriver.Chrome) -> None:
        """Nhấn nút 'Cắt và lưu' và đợi đến khi phần tử 'Khung hình đầu tiên' xuất hiện (đã có ảnh)."""
        # Wait a bit to ensure UI is ready before clicking crop button
        time.sleep(2)
        
        # Click nút Cắt và lưu (VI) hoặc 'Crop and save' (EN fallback)
        crop_candidates = [
            "//button[.//i[normalize-space(text())='crop'] and contains(., 'Cắt và lưu')]",
            "//button[contains(., 'Cắt và lưu')]",
            "//button[.//i[normalize-space(text())='crop'] and contains(., 'Crop')]",
            "//button[contains(., 'Crop and save')]",
        ]
        for xp in crop_candidates:
            try:
                el = driver.find_element(By.XPATH, xp)
                # Đợi thêm 1s sau khi tìm thấy nút để đảm bảo UI đã sẵn sàng
                time.sleep(1)
                self._human_click_el(driver, el)
                time.sleep(0.5)
                break
            except Exception:
                continue

        # Đợi xuất hiện nút chứa nội dung 'Khung hình đầu tiên'
        def first_frame_ready():
            try:
                return len(driver.find_elements(By.XPATH, "//button[.//span[normalize-space(text())='Khung hình đầu tiên']]") ) > 0
            except Exception:
                return False
        # Thời gian chờ tối đa cho upload + crop preview là 60s
        ok = self._wait_until(first_frame_ready, timeout=60, interval=0.5)
        if not ok:
            raise Exception("First frame not detected after crop/save")

    def _click_create_button(self, driver: webdriver.Chrome) -> None:
        """Nhấn nút 'Tạo' (Create) với icon arrow_forward. Đợi nút enabled trước khi click."""
        # Chờ nút xuất hiện và enabled
        def get_button():
            candidates = [
                "//button[.//span[normalize-space(text())='Tạo']]",
                "//button[.//i[normalize-space(text())='arrow_forward']]",
                "//button[contains(., 'Tạo') or contains(., 'Create')]",
            ]
            for xp in candidates:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    return el
                except Exception:
                    continue
            return None

        start = time.time()
        btn = None
        while time.time() - start < 30:
            btn = get_button()
            if btn is not None:
                try:
                    disabled = btn.get_attribute('disabled')
                    aria_disabled = btn.get_attribute('aria-disabled')
                    if not disabled and (aria_disabled in (None, '', 'false')):
                        break
                except Exception:
                    break
            time.sleep(0.5)

        if btn is None:
            raise Exception("Create button not found")
        try:
            self._human_click_el(driver, btn)
            time.sleep(0.3)
        except Exception:
            # Fallback: try executing click via JS
            try:
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(0.3)
            except Exception as ex:
                raise Exception(f"Failed to click Create button: {ex}")

    def _monitor_and_fetch_api(self, driver: webdriver.Chrome, wf: str = None) -> None:
        """Theo dõi xử lý đến khi đủ video hoàn tất theo cấu hình Outputs per prompt,
        sau đó reload trang (nếu cần) và đọc API project.searchProjectWorkflows để lấy fifeUri."""
        # Số video kỳ vọng theo cấu hình Outputs per prompt (mặc định 1)
        try:
            expected_videos = int((self.outputs_per_prompt.get() or "1").strip())
            if expected_videos <= 0:
                expected_videos = 1
        except Exception:
            expected_videos = 1

        def is_any_running_or_progress():
            """Kiểm tra còn trạng thái đang chạy không (nhãn 'running' hoặc %)."""
            try:
                # Nhãn 'running'
                running = driver.find_elements(By.XPATH, "//*[contains(translate(normalize-space(text()), 'RUNNING', 'running'), 'running') or contains(@class,'running')]")
            except Exception:
                running = []
            # Tiến trình phần trăm: chỉ bám theo text có ký tự '%', giới hạn chuỗi ngắn để tránh nhiễu
            try:
                percents = driver.find_elements(By.XPATH, "//*[contains(normalize-space(text()), '%') and string-length(normalize-space(text())) <= 6]")
            except Exception:
                percents = []
            # Nếu có container style="overflow-anchor: none", ưu tiên xem trong đó
            try:
                anchor_percents = driver.find_elements(By.XPATH, "//*[@style='overflow-anchor: none']//*[contains(normalize-space(text()), '%') and string-length(normalize-space(text())) <= 6]")
            except Exception:
                anchor_percents = []
            return (len(running) > 0) or (len(percents) > 0) or (len(anchor_percents) > 0)

        def list_videos():
            """Get actual result videos, not placeholders or loading videos."""
            try:
                # Get all video elements
                all_videos = driver.find_elements(By.TAG_NAME, 'video')
                result_videos = []
                
                for video in all_videos:
                    try:
                        # Check if this video is a real result (has duration > 0 and is not a placeholder)
                        duration = video.get_attribute('duration')
                        if duration and float(duration) > 0:
                            # Additional check: video should be in a result card
                            parent_card = video.find_element(By.XPATH, "./ancestor::*[contains(@class, 'card') or contains(@class, 'result') or contains(@class, 'video')]")
                            if parent_card:
                                result_videos.append(video)
                    except Exception:
                        # If we can't determine if it's a real video, include it but log
                        result_videos.append(video)
                # Only log detailed video count for non text_to_video to avoid confusion when only cards exist
                if wf != "text_to_video":
                    self._log_exec(f"Found {len(result_videos)} result videos out of {len(all_videos)} total video elements")
                return result_videos
            except Exception:
                return []

        def is_card_ready_from_video(video_el):
            """Card ready nếu tổ tiên có nút hành động đặc trưng (download/fullscreen/more/Thêm vào cảnh)."""
            try:
                node = video_el
                for _ in range(6):
                    try:
                        node = node.find_element(By.XPATH, "..")
                    except Exception:
                        break
                    actions = node.find_elements(By.XPATH, 
                        ".//button[.//i[normalize-space(text())='download'] or .//span[contains(., 'Tải xuống')] or .//i[normalize-space(text())='fullscreen'] or .//i[normalize-space(text())='more_vert'] or .//span[contains(., 'Thêm vào cảnh')]]"
                    )
                    if actions:
                        return True
                return False
            except Exception:
                return False

        # Bỏ phát hiện lỗi theo text; chỉ dựa vào timeout và API

        # Specialized card handling for text_to_video
        def list_t2v_cards():
            try:
                # Card container per provided HTML: sc-510f5a89-0 ...
                cards = driver.find_elements(By.XPATH, "//*[contains(@class,'sc-510f5a89-0')]")
                return cards
            except Exception:
                return []

        def t2v_card_has_progress(card_el):
            try:
                # Progress percent inside card: sc-dd6abb21-1 or any % text within the card
                prog = card_el.find_elements(By.XPATH, ".//*[contains(@class,'sc-dd6abb21-1') or contains(text(), '%')]")
                return len(prog) > 0
            except Exception:
                return False

        def is_t2v_card_ready(card_el):
            try:
                if t2v_card_has_progress(card_el):
                    return False
                # Ready if action buttons present or media loaded
                actions = card_el.find_elements(By.XPATH,
                    ".//button[.//i[normalize-space(text())='download'] or .//span[contains(., 'Tải xuống')] or .//i[normalize-space(text())='fullscreen'] or .//i[normalize-space(text())='more_vert'] or .//span[contains(., 'Thêm vào cảnh')]]"
                )
                if actions:
                    return True
                # Or a playable video inside
                vids = card_el.find_elements(By.TAG_NAME, 'video')
                for v in vids:
                    try:
                        duration = v.get_attribute('duration')
                        if duration and float(duration) > 0:
                            return True
                    except Exception:
                        continue
                # Or an image fully loaded
                imgs = card_el.find_elements(By.TAG_NAME, 'img')
                for im in imgs:
                    try:
                        loaded = driver.execute_script("return arguments[0].complete && arguments[0].naturalWidth > 0;", im)
                        if loaded:
                            return True
                    except Exception:
                        continue
                return False
            except Exception:
                return False

        def all_videos_ready():
            try:
                # For text_to_video, base readiness on cards rather than <video> elements
                if wf == "text_to_video":
                    cards = list_t2v_cards()
                    self._log_exec(f"Checking card readiness (text_to_video): found {len(cards)} cards, expected {expected_videos}")
                    if len(cards) < expected_videos:
                        self._log_exec(f"Not enough cards yet: {len(cards)} < {expected_videos}")
                        return False
                    if is_any_running_or_progress():
                        self._log_exec("Still processing - progress indicators present in cards")
                        return False
                    ready = 0
                    for i, c in enumerate(cards[:expected_videos]):
                        if is_t2v_card_ready(c):
                            ready += 1
                            self._log_exec(f"Card {i+1} is ready")
                        else:
                            self._log_exec(f"Card {i+1} is not ready yet")
                    if ready < expected_videos:
                        self._log_exec(f"Not all cards ready: {ready} < {expected_videos}")
                        return False
                    self._log_exec(f"All {expected_videos} card(s) are ready!")
                    return True

                vids = list_videos()
                # For frames_to_video, if no videos yet, derive from percent nodes to avoid confusing log
                if wf == "frames_to_video" and len(vids) == 0:
                    try:
                        anchor_nodes = driver.find_elements(By.XPATH, "//*[@style='overflow-anchor: none']//*[contains(normalize-space(text()), '%') and string-length(normalize-space(text())) <= 6]")
                    except Exception:
                        anchor_nodes = []
                    derived_total = len(anchor_nodes)
                    self._log_exec(f"Checking video readiness (frames_to_video): found 0 videos, derived {derived_total} processing cards from %")
                else:
                    self._log_exec(f"Checking video readiness: found {len(vids)} videos, expected {expected_videos}")
                
                # If we have fewer videos than expected, not ready yet
                if len(vids) < expected_videos:
                    self._log_exec(f"Not enough videos yet: {len(vids)} < {expected_videos}")
                    return False
                
                # When we have enough videos, ensure no running/progress states
                if is_any_running_or_progress():
                    self._log_exec("Still processing - found running/progress indicators")
                    return False
                
                # Each card must have action buttons (ready)
                ready = 0
                for i, v in enumerate(vids):
                    # Với text_to_video, card container có class sc-510f5a89-0 ...; nếu cần, xác nhận thêm
                    if is_card_ready_from_video(v):
                        ready += 1
                        self._log_exec(f"Video {i+1} is ready")
                    else:
                        self._log_exec(f"Video {i+1} is not ready yet")
                
                if ready < expected_videos:
                    self._log_exec(f"Not all videos ready: {ready} < {expected_videos}")
                    return False
                
                self._log_exec(f"All {expected_videos} videos are ready!")
                return True
            except Exception as e:
                self._log_exec(f"Error checking video readiness: {e}")
                return False

        # Poll API bằng cách reload trang mỗi 10s, tối đa 3 phút, đến khi đủ media URL(s)
        start = time.time()
        last_reload = 0.0
        # Map normalized_url -> original_url (keep original with query params for authorized download)
        collected_url_map = {}
        target_fragment = '/fx/api/trpc/project.searchProjectWorkflows'

        # Bật Network nếu cần (an toàn khi gọi nhiều lần)
        try:
            driver.execute_cdp_cmd('Network.enable', {})
        except Exception:
            pass

        # Delay first reload by 60s, then reload every ~10s
        initial_reload_delay = 60.0
        notified_wait = False
        retries_after_initial = 0
        max_retries_after_initial = 3

        while True:
            if self.stop_exec:
                self._log_exec("Stopped by user during monitoring")
                return

            now = time.time()
            if (now - start) < initial_reload_delay:
                if not notified_wait:
                    try:
                        self._log_exec("Waiting 60s before first reload to collect media...")
                    except Exception:
                        pass
                    notified_wait = True
            elif now - last_reload >= 10:
                try:
                    self._log_exec("Reloading page to check TRPC API responses...")
                    driver.refresh()
                    time.sleep(2)
                except Exception:
                    pass
                last_reload = now
                # Count retries only after the initial 60s window
                if (now - start) >= initial_reload_delay:
                    retries_after_initial += 1

            # Đọc performance logs và thu thập response bodies của API mục tiêu
            try:
                logs = driver.get_log('performance')
            except Exception:
                logs = []

            request_ids = []
            for item in logs:
                try:
                    msg = json.loads(item.get('message', '{}')).get('message', {})
                    method = msg.get('method')
                    params = msg.get('params', {})
                    if method == 'Network.responseReceived':
                        response = params.get('response', {})
                        url = response.get('url', '')
                        if target_fragment in url and response.get('mimeType', '').startswith('application/json'):
                            request_ids.append(params.get('requestId'))
                except Exception:
                    continue

            for rid in request_ids:
                try:
                    body = driver.execute_cdp_cmd('Network.getResponseBody', {'requestId': rid})
                    text = body.get('body', '')
                    try:
                        data = json.loads(text)
                    except Exception:
                        data = text
                    urls = self._extract_fife_uris_from_api_json(data)
                    for u in urls:
                        try:
                            norm = self._normalize_media_url(u)
                        except Exception:
                            norm = u
                        if norm not in collected_url_map:
                            collected_url_map[norm] = u
                except Exception:
                    continue

            self._log_exec(f"API media collected: {len(collected_url_map)}/{expected_videos}")

            if len(collected_url_map) >= expected_videos:
                break
            # If we've passed the initial delay and exhausted retry attempts, proceed with whatever we have
            if (time.time() - start) >= initial_reload_delay and retries_after_initial >= max_retries_after_initial:
                self._log_exec("Max retries after initial wait reached. Proceeding with available media.")
                break
            if time.time() - start >= 180:
                self._log_exec("Timeout 3 minutes reached. Proceeding to download available media.")
                break
            time.sleep(1)

        # Use original URLs (with query) for actual download, but de-duplicated by normalized key
        all_urls = list(collected_url_map.values())
        # Limit to expected amount to avoid over-downloading due to variant URLs
        try:
            expected_videos = int((self.outputs_per_prompt.get() or "1").strip())
            if expected_videos > 0:
                all_urls = all_urls[:expected_videos]
        except Exception:
            pass
        if all_urls:
            self._log_exec(f"Found {len(all_urls)} media URL(s) from API. Downloading...")
            prompt_text = getattr(self, 'current_prompt', '')
            self._download_files(all_urls, prompt_text)
            try:
                self._log_exec("Job completed. Closing browser now and continuing queue...", success=True)
                try:
                    if self.exec_driver is not None:
                        self.exec_driver.quit()
                except Exception:
                    pass
                return
            except Exception:
                pass
        else:
            self._log_exec("No API JSON or no media URLs extracted - marking as failed", error=True)
            try:
                messagebox.showerror("Thất bại", "Không tìm thấy kết quả để tải từ API. Tiến trình được đánh dấu thất bại.")
            except Exception:
                pass

    def _stop_execution(self) -> None:
        """Stop current execution and close browser if running."""
        try:
            self.stop_exec = True
            self._log_exec("Stopping execution and closing browser...")
            if self.exec_driver is not None:
                try:
                    self.exec_driver.quit()
                except Exception:
                    pass
                self.exec_driver = None
            self._set_exec_status("Stopped", "red")
        except Exception:
            pass

    def _extract_fife_uris_from_api_json(self, payload):
        urls = []
        try:
            if isinstance(payload, str):
                payload = json.loads(payload)
        except Exception:
            return urls
        try:
            result = payload.get('result', {})
            d2 = result.get('data', {})
            j = d2.get('json', {})
            res = j.get('result', {})
            workflows = res.get('workflows', [])
            for wf in workflows:
                for st in wf.get('workflowSteps', []):
                    for g in st.get('mediaGenerations', []):
                        try:
                            vd = g.get('mediaData', {}).get('videoData', {})
                            fife = vd.get('fifeUri')
                            if fife:
                                urls.append(fife)
                        except Exception:
                            continue
        except Exception:
            return urls
        return urls

    def _sanitize_folder_name(self, text, max_length=50):
        """Sanitize text for use as folder name, keeping it safe and readable."""
        if not text:
            return "no_prompt"
        
        # Remove or replace problematic characters
        sanitized = re.sub(r'[<>:"/\\|?*]', '_', text)
        # Replace multiple spaces with single space
        sanitized = re.sub(r'\s+', ' ', sanitized)
        # Remove leading/trailing spaces and dots
        sanitized = sanitized.strip(' .')
        # Truncate if too long
        if len(sanitized) > max_length:
            sanitized = sanitized[:max_length].rstrip()
        # Ensure it's not empty
        if not sanitized:
            sanitized = "prompt"
        
        return sanitized

    def _download_files(self, urls, prompt_text=""):
        try:
            # Save all files to common downloads folder (no per-prompt subfolder)
            out_dir = Path(os.getcwd()) / "downloads"
            out_dir.mkdir(parents=True, exist_ok=True)
            self._log_exec("Using common downloads folder")
            for i, url in enumerate(urls, 1):
                try:
                    ext = ".mp4"
                    if 'image' in url:
                        ext = ".png"
                    # Ensure unique filenames to avoid overwriting
                    # Format: index_array + media_number + timestamp
                    ts = time.strftime('%Y%m%d_%H%M%S')
                    # index_array là số thứ tự của job hiện tại
                    index_array = getattr(self, 'current_job_index', 1)
                    media_number = i
                    base_name = f"{index_array}_{media_number}_{ts}{ext}"
                    dest = out_dir / base_name
                    attempt = 1
                    while dest.exists() and attempt < 1000:
                        dest = out_dir / f"{index_array}_{media_number}_{ts}_{attempt}{ext}"
                        attempt += 1
                    self._log_exec(f"Downloading {dest.name}...")
                    # Build request with headers to avoid 403
                    req = urllib.request.Request(url, headers={
                        'User-Agent': self.current_user_agent or 'Mozilla/5.0',
                        'Referer': 'https://labs.google/fx/tools/flow',
                        'Accept': '*/*',
                        'Connection': 'keep-alive',
                    })
                    # Simple retries for transient errors
                    max_dl_retries = 3
                    for r in range(max_dl_retries):
                        try:
                            with urllib.request.urlopen(req, timeout=30) as resp, open(dest, 'wb') as f:
                                f.write(resp.read())
                            break
                        except Exception as e:
                            if r == max_dl_retries - 1:
                                raise e
                            time.sleep(1.5)
                    self._log_exec(f"Downloaded {dest.name}", success=True)
                except Exception as ex:
                    self._log_exec(f"Failed to download #{i}: {ex}", error=True)
            # No popup: silently finish; the outer finally will close browser and continue queue
        except Exception as ex:
            self._log_exec(f"Download error: {ex}", error=True)

    def _normalize_media_url(self, url: str) -> str:
        """Normalize media URL to avoid duplicates differing only by query/fragment.
        Keeps scheme, netloc, and path. Strips query and fragment.
        """
        try:
            parts = urlsplit(url)
            return urlunsplit((parts.scheme, parts.netloc, parts.path, '', ''))
        except Exception:
            return url

    def _open_settings_and_apply(self, driver: webdriver.Chrome, aspect: str, outputs: str, model: str) -> None:
        """Mở popover Cài đặt (tune) và đặt: Tỷ lệ khung hình, Outputs per prompt, Model."""
        # Mở popover Cài đặt
        settings_btn_xp = [
            "//button[.//i[normalize-space(text())='tune']]",
            "//button[.//span[normalize-space(text())='Cài đặt']]",
            "//button[contains(., 'Cài đặt') or contains(., 'Settings')]",
        ]
        clicked = False
        for xp in settings_btn_xp:
            try:
                el = driver.find_element(By.XPATH, xp)
                self._human_click_el(driver, el)
                clicked = True
                break
            except Exception:
                continue
        if not clicked:
            raise Exception("Settings button not found")

        # Đợi popover role=dialog mở
        try:
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//*[@role='dialog']")))
        except Exception:
            pass

        # Helper chọn từ combobox theo nhãn
        def select_from_combobox(label_texts, option_texts):
            # Tìm combobox có label tương ứng và thao tác trên listbox theo aria-controls của chính combobox đó
            for label_text in label_texts:
                try:
                    combo = driver.find_element(By.XPATH, f"//button[@role='combobox'][.//span[normalize-space(text())='{label_text}']]")
                    self._log_exec(f"Opening combobox: {label_text}...")
                    # Lấy id listbox mục tiêu từ aria-controls (Radix Select gán id động: radix-:xxx:)
                    aria_controls = combo.get_attribute("aria-controls") or ""
                    self._human_click_el(driver, combo)
                    # Đợi đúng listbox của combobox này mở ra
                    target_listbox = None
                    if aria_controls:
                        try:
                            target_listbox = WebDriverWait(driver, 7).until(
                                EC.presence_of_element_located((By.ID, aria_controls))
                            )
                        except Exception:
                            target_listbox = None
                    if target_listbox is None:
                        # Fallback: lấy listbox gần nhất sau khi mở
                        try:
                            target_listbox = WebDriverWait(driver, 5).until(
                                EC.presence_of_element_located((By.XPATH, "//*[@role='listbox']"))
                            )
                        except Exception:
                            target_listbox = None

                    # Chọn option theo text bên trong listbox mục tiêu
                    if target_listbox is not None:
                        try:
                            driver.execute_script("arguments[0].scrollTop = 0;", target_listbox)
                        except Exception:
                            pass
                        
                        # Log all available options for debugging
                        all_options = target_listbox.find_elements(By.XPATH, ".//*[@role='option']")
                        self._log_exec(f"Available options in listbox: {[opt.text.strip() for opt in all_options]}")
                        
                        for opt_text in option_texts:
                            self._log_exec(f"Trying to select option: '{opt_text}'")
                            try:
                                # Ưu tiên match exact theo span trong listbox mục tiêu
                                el = target_listbox.find_element(By.XPATH, f".//*[@role='option'][.//span[normalize-space(text())='{opt_text}']]")
                                self._log_exec(f"Found exact match for: '{opt_text}'")
                                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                                self._human_click_el(driver, el)
                                time.sleep(0.2)
                                
                                # Verify selection by checking if the option is now selected
                                try:
                                    selected_el = target_listbox.find_element(By.XPATH, f".//*[@role='option'][.//span[normalize-space(text())='{opt_text}']")
                                    if selected_el.get_attribute("data-state") == "checked":
                                        self._log_exec(f"Successfully selected: '{opt_text}'")
                                        return True
                                    else:
                                        self._log_exec(f"Option '{opt_text}' clicked but not selected")
                                except Exception:
                                    self._log_exec(f"Could not verify selection for: '{opt_text}'")
                                
                                return True
                            except Exception:
                                # Fallback: contains text
                                try:
                                    el = target_listbox.find_element(By.XPATH, f".//*[@role='option'][contains(., '{opt_text}')]")
                                    self._log_exec(f"Found partial match for: '{opt_text}'")
                                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                                    self._human_click_el(driver, el)
                                    time.sleep(0.2)
                                    
                                    # Verify selection
                                    try:
                                        selected_el = target_listbox.find_element(By.XPATH, f".//*[@role='option'][contains(., '{opt_text}')]")
                                        if selected_el.get_attribute("data-state") == "checked":
                                            self._log_exec(f"Successfully selected: '{opt_text}'")
                                            return True
                                    except Exception:
                                        pass
                                    
                                    return True
                                except Exception:
                                    self._log_exec(f"No match found for: '{opt_text}'")
                                    continue

                        # Nếu vẫn chưa tìm thấy, thử viewport nội bộ của Radix để scroll
                        try:
                            viewport = target_listbox.find_element(By.CSS_SELECTOR, "[data-radix-select-viewport]")
                            driver.execute_script("arguments[0].scrollTop = 0;", viewport)
                            time.sleep(0.1)
                            for opt_text in option_texts:
                                try:
                                    el = target_listbox.find_element(By.XPATH, f".//*[@role='option'][.//span[normalize-space(text())='{opt_text}']]")
                                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                                    self._human_click_el(driver, el)
                                    time.sleep(0.2)
                                    return True
                                except Exception:
                                    continue
                        except Exception:
                            pass
                except Exception:
                    continue
            return False

        # 1) Aspect ratio
        aspect_map = {
            "16:9": ["Khổ ngang (16:9)", "Landscape (16:9)"],
            "9:16": ["Khổ dọc (9:16)", "Portrait (9:16)"],
            "1:1": ["Vuông (1:1)", "Square (1:1)"],
        }
        select_from_combobox(["Tỷ lệ khung hình", "Aspect ratio"], aspect_map.get(aspect, [aspect]))
        time.sleep(1)

        # 2) Outputs per prompt
        self._log_exec(f"Setting outputs per prompt to: {outputs}")
        success = select_from_combobox(["Câu trả lời đầu ra cho mỗi câu lệnh", "Outputs per prompt"], [outputs])
        if not success:
            self._log_exec(f"Failed to select outputs per prompt: {outputs}")
            # Try alternative approach - select by index
            try:
                combo = driver.find_element(By.XPATH, "//button[@role='combobox'][.//span[contains(text(), 'Câu trả lời') or contains(text(), 'Outputs')]]")
                self._human_click_el(driver, combo)
                time.sleep(0.5)
                
                # Get the listbox
                listbox = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//*[@role='listbox']"))
                )
                
                # Try to select by index (outputs-1 since it's 0-based)
                try:
                    index = int(outputs) - 1
                    options = listbox.find_elements(By.XPATH, ".//*[@role='option']")
                    if 0 <= index < len(options):
                        self._log_exec(f"Selecting option by index: {index}")
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", options[index])
                        self._human_click_el(driver, options[index])
                        
                        # Verify selection using JavaScript
                        time.sleep(0.5)
                        selected_option = driver.execute_script("""
                            const listbox = arguments[0];
                            const options = listbox.querySelectorAll('[role="option"]');
                            for (let i = 0; i < options.length; i++) {
                                if (options[i].getAttribute('data-state') === 'checked') {
                                    return options[i].textContent.trim();
                                }
                            }
                            return null;
                        """, listbox)
                        
                        if selected_option == outputs:
                            self._log_exec(f"Successfully selected outputs per prompt: {selected_option}")
                            success = True
                        else:
                            self._log_exec(f"Selection verification failed. Expected: {outputs}, Got: {selected_option}")
                            
                except Exception as e:
                    self._log_exec(f"Failed to select by index: {e}")
                
            except Exception as e:
                self._log_exec(f"Alternative selection failed: {e}")
        
        # Final fallback: Use JavaScript to directly select the option
        if not success:
            self._log_exec("Trying JavaScript fallback method...")
            try:
                # Find the combobox and click it
                combo = driver.find_element(By.XPATH, "//button[@role='combobox'][.//span[contains(text(), 'Câu trả lời') or contains(text(), 'Outputs')]]")
                self._human_click_el(driver, combo)
                time.sleep(0.5)
                
                # Use JavaScript to select the option directly
                result = driver.execute_script("""
                    const listbox = document.querySelector('[role="listbox"]');
                    if (!listbox) return false;
                    
                    const options = listbox.querySelectorAll('[role="option"]');
                    for (let i = 0; i < options.length; i++) {
                        const span = options[i].querySelector('span');
                        if (span && span.textContent.trim() === arguments[0]) {
                            // First uncheck all options
                            options.forEach(opt => {
                                opt.setAttribute('data-state', 'unchecked');
                                opt.setAttribute('aria-selected', 'false');
                            });
                            
                            // Then check the target option
                            options[i].setAttribute('data-state', 'checked');
                            options[i].setAttribute('aria-selected', 'true');
                            
                            // Trigger click event
                            options[i].click();
                            return true;
                        }
                    }
                    return false;
                """, outputs)
                
                if result:
                    self._log_exec(f"JavaScript fallback successful: selected {outputs}")
                    success = True
                else:
                    self._log_exec(f"JavaScript fallback failed: could not find option {outputs}")
                    
            except Exception as e:
                self._log_exec(f"JavaScript fallback error: {e}")
        
        if success:
            self._log_exec(f"Successfully set outputs per prompt to: {outputs}")
        else:
            self._log_exec(f"Failed to set outputs per prompt to: {outputs}")
        time.sleep(1)

        # 3) Model
        # Cố gắng chọn chính xác model và xác thực kết quả hiển thị. Nếu lệch, thử theo index.
        ok_model = self._select_model_strict(driver, model)
        time.sleep(1)
        if not ok_model:
            # Fallback: thử helper tổng quát + chọn trực tiếp trong listbox
            tried = select_from_combobox(["Mô hình", "Model"], [model])
            if not tried:
                try:
                    self._select_radix_option_by_text(driver, model)
                except Exception:
                    pass

    def _wait_until(self, predicate, timeout: int = 60, interval: float = 1.0) -> bool:
        start = time.time()
        while time.time() - start < timeout:
            try:
                if predicate():
                    return True
            except Exception:
                pass
            time.sleep(interval)
        return False
    
    def _select_model_strict(self, driver: webdriver.Chrome, model_text: str) -> bool:
        """Chọn model với xác thực sau click: mở combobox 'Mô hình', chọn option theo text.
        """
        try:
            def get_model_combo():
                try:
                    return driver.find_element(By.XPATH, "//button[@role='combobox'][.//span[normalize-space(text())='Mô hình' or normalize-space(text())='Model']]")
                except Exception:
                    return None

            def read_combo_value(combo_el):
                try:
                    # Lấy span hiển thị giá trị (thường là span sau label)
                    spans = combo_el.find_elements(By.XPATH, ".//span")
                    if len(spans) >= 2:
                        return (spans[-1].text or "").strip()
                    return (combo_el.text or "").strip()
                except Exception:
                    return ""

            attempts = 0
            while attempts < 3:
                attempts += 1
                combo = get_model_combo()
                if combo is None:
                    return False
             

                # Mở listbox và lấy đúng listbox theo aria-controls
                self._human_click_el(driver, combo)
                
                # Đợi 2 giây sau khi mở popup trước khi tìm model
                time.sleep(2)
                
                aria_controls = combo.get_attribute("aria-controls") or ""
                listbox = None
                if aria_controls:
                    try:
                        listbox = WebDriverWait(driver, 7).until(EC.presence_of_element_located((By.ID, aria_controls)))
                    except Exception:
                        listbox = None
                if listbox is None:
                    try:
                        listbox = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//*[@role='listbox']")))
                    except Exception:
                        listbox = None
                if listbox is None:
                    continue

                picked = False
                # Phương pháp 1: Click theo text chính xác
                try:
                    el = listbox.find_element(By.XPATH, f".//*[@role='option'][.//span[normalize-space(text())='{model_text}']]")
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                    self._human_click_el(driver, el)
                    picked = True
                except Exception:
                    # Phương pháp 2: Click theo text chứa (fallback)
                    try:
                        el = listbox.find_element(By.XPATH, f".//*[@role='option'][contains(., '{model_text}')]")
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        self._human_click_el(driver, el)
                        picked = True
                    except Exception:
                        pass

                time.sleep(0.3)
                # Xác thực kết quả hiển thị trên combobox
                combo = get_model_combo()
                if combo is None:
                    continue
                shown = read_combo_value(combo)
                if (model_text or "").strip().lower() == (shown or "").strip().lower():
                    return True
                # Nếu chưa đúng, thử đóng listbox (ESC) và lặp lại
                try:
                    ActionChains(driver).send_keys(Keys.ESCAPE).perform()
                except Exception:
                    pass
                time.sleep(0.2)
            return False
        except Exception:
            return False

    def _select_radix_option_by_text(self, driver: webdriver.Chrome, text: str) -> None:
        """Chọn option trong popover Radix Select theo nội dung span hiển thị.
        Hỗ trợ cấu trúc như user cung cấp: role=listbox, items role=option, có span chứa text.
        """
        try:
            # Đợi listbox mở
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//*[@role='listbox']"))
                )
            except Exception:
                pass

            # Ưu tiên match exact theo span
            candidates = [
                f"//*[@role='listbox']//*[@role='option'][.//span[normalize-space(text())='{text}']]",
                f"//*[@role='listbox']//*[@role='option'][contains(., '{text}')]",
                f"//*[@role='listbox']//*[contains(@class,'sc-acb5d8f5-2') and @role='option' and .//span[normalize-space(text())='{text}']]",
            ]





            for xp in candidates:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    self._human_click_el(driver, el)
                    time.sleep(0.2)
                    return
                except Exception:
                    continue

            # Fallback: scroll viewport và thử lại
            try:
                viewport = driver.find_element(By.CSS_SELECTOR, "[data-radix-select-viewport]")
                # Scroll lên đỉnh và thử
                driver.execute_script("arguments[0].scrollTop = 0;", viewport)
                time.sleep(0.1)
                el = driver.find_element(By.XPATH, f"//*[@role='listbox']//*[@role='option'][.//span[normalize-space(text())='{text}']]")
                self._human_click_el(driver, el)
                time.sleep(0.2)
                return
            except Exception:
                pass

            # Fallback cuối: duyệt tất cả role=option và chọn item có aria-selected='true' gần text (nếu text rút gọn)
            try:
                
                
                # dsasa
                options = driver.find_elements(By.XPATH, "//*[@role='listbox']//*[@role='option']")
                for opt in options:
                    try:
                        label_span = opt.find_element(By.XPATH, ".//span")
                        label = label_span.text.strip()
                        if label == text or text.lower() in label.lower():
                            self._human_click_el(driver, opt)
                            time.sleep(0.2)
                            return
                    except Exception:
                        continue
            except Exception:
                pass
        except Exception:
            pass

    # ===================== Human-like helpers =====================
    def _human_delay(self, min_seconds: float = 1.0, max_seconds: float = 2.5) -> None:
        time.sleep(random.uniform(min_seconds, max_seconds))

    def _human_warm_up_page(self, driver: webdriver.Chrome) -> None:
        try:
            for _ in range(random.randint(1, 3)):
                dy = random.randint(100, 400) * (1 if random.random() < 0.5 else -1)
                driver.execute_script("window.scrollBy(0, arguments[0]);", dy)
                time.sleep(random.uniform(0.1, 0.4))
            try:
                body = driver.find_element(By.TAG_NAME, 'body')
                ActionChains(driver).move_to_element_with_offset(body, random.randint(5, 50), random.randint(5, 30)).pause(random.uniform(0.05, 0.2)).perform()
            except Exception:
                pass
        except Exception:
            pass

    def _human_click_el(self, driver: webdriver.Chrome, element) -> None:
        try:
            ActionChains(driver).move_to_element(element).pause(random.uniform(0.05, 0.2)).click().perform()
        except Exception:
            try:
                element.click()
            except Exception:
                pass

    def _fast_type_prompt(self, element, text: str) -> None:
        """Ultra-fast typing specifically optimized for prompt input.
        Uses maximum chunk sizes and minimal delays for fastest possible typing.
        """
        try:
            length = len(text or "")
            if length == 0:
                return
            
            # Try to send entire text at once first (fastest possible)
            try:
                element.send_keys(text)
                return
            except Exception:
                pass
            
            # If that fails, use very large chunks with minimal delay
            if length >= 1000:
                chunk_size = 500  # Huge chunks for very long prompts
            elif length >= 500:
                chunk_size = 250  # Very large chunks for long prompts
            elif length >= 200:
                chunk_size = 100  # Large chunks for medium prompts
            else:
                chunk_size = 50   # Medium chunks for shorter prompts
            
            for i in range(0, length, chunk_size):
                chunk = text[i:i+chunk_size]
                element.send_keys(chunk)
                # Ultra-minimal delay - only 0.0001s between chunks
                time.sleep(0.0001)
        except Exception:
            # Last resort: per-character with absolute minimal delay
            for ch in text:
                try:
                    element.send_keys(ch)
                except Exception:
                    continue
                time.sleep(0.0001)

    def _human_type_el(self, element, text: str) -> None:
        """Type text into an element with optimized speed for prompt input.
        Uses larger chunks and minimal delays for faster typing.
        """
        try:
            length = len(text or "")
            if length == 0:
                return
            
            # Optimized chunk sizes for maximum speed while maintaining compatibility
            if length >= 500:
                chunk_size = 200  # Very large chunks for long prompts
            elif length >= 200:
                chunk_size = 100  # Large chunks for medium prompts
            elif length >= 50:
                chunk_size = 50   # Medium chunks for shorter prompts
            else:
                chunk_size = 25   # Small chunks for very short prompts
            
            for i in range(0, length, chunk_size):
                chunk = text[i:i+chunk_size]
                element.send_keys(chunk)
                # Minimal delay for speed - only 0.001s between chunks
                time.sleep(0.001)
        except Exception:
            # Fallback: try to send entire text at once for maximum speed
            try:
                element.send_keys(text)
            except Exception:
                # Last resort: per-character with minimal delay
                for ch in text:
                    try:
                        element.send_keys(ch)
                    except Exception:
                        continue
                    time.sleep(0.001)

    # ===================== Status =====================
    def _set_status(self, text: str, color: str) -> None:
        try:
            # Callback for non-Tk UI
            if hasattr(self, 'ui_callbacks') and self.ui_callbacks.get('on_status'):
                try:
                    self.ui_callbacks['on_status'](text, color)
                except Exception:
                    pass
            # Tk fallback
            if getattr(self, 'use_tk_ui', True) and hasattr(self, 'status_label'):
                style_map = {
                    'blue': 'Info.TLabel',
                    'green': 'Success.TLabel', 
                    'red': 'Error.TLabel',
                    'orange': 'Warning.TLabel'
                }
                style = style_map.get(color, 'Info.TLabel')
                self.status_label.config(text=text, style=style)
        except Exception:
            pass

    def _set_exec_status(self, text: str, color: str) -> None:
        try:
            # Callback for non-Tk UI
            if hasattr(self, 'ui_callbacks') and self.ui_callbacks.get('on_exec_status'):
                try:
                    self.ui_callbacks['on_exec_status'](text, color)
                except Exception:
                    pass
            # Tk fallback
            if getattr(self, 'use_tk_ui', True) and hasattr(self, 'exec_status'):
                style_map = {
                    'blue': 'Info.TLabel',
                    'green': 'Success.TLabel', 
                    'red': 'Error.TLabel',
                    'orange': 'Warning.TLabel'
                }
                style = style_map.get(color, 'Info.TLabel')
                self.exec_status.config(text=text, style=style)
        except Exception:
            pass


def main() -> None:
    # Use ttkbootstrap Window when available for a dramatically nicer default UI
    if _HAS_TTKBOOTSTRAP and TtkbWindow is not None:
        root = TtkbWindow(themename='superhero')
    else:
        root = tk.Tk()
    app = FlowBrowserTool(root, use_tk_ui=True)
    # Đảm bảo đóng toàn bộ browser khi thoát ứng dụng
    def _on_app_close():
        try:
            # Đóng driver login
            try:
                if getattr(app, 'driver', None) is not None:
                    app.driver.quit()
            except Exception:
                pass
            # Đóng driver execute
            try:
                if getattr(app, 'exec_driver', None) is not None:
                    app.exec_driver.quit()
            except Exception:
                pass
        except Exception:
            pass
        try:
            root.destroy()
        except Exception:
            pass

    try:
        root.protocol("WM_DELETE_WINDOW", _on_app_close)
    except Exception:
        pass
    root.mainloop()


if __name__ == "__main__":
    main()


