import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
from tkinter import simpledialog
import os
import base64
import re
import json
import time
import random
import threading

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager

# Optional modern theming with ttkbootstrap (for a nicer look)
try:
    from ttkbootstrap import Style as TtkbStyle
    from ttkbootstrap import Window as TtkbWindow
    _HAS_TTKBOOTSTRAP = True
except Exception:
    TtkbStyle = None
    TtkbWindow = None
    _HAS_TTKBOOTSTRAP = False


WHISK_URL = "https://labs.google/fx/vi/tools/whisk"
WHISK_PROJECT_URL = "https://labs.google/fx/vi/tools/whisk/project"


class WhiskBrowserTool:
    def __init__(self, root: tk.Tk | None, use_tk_ui: bool = True, ui_callbacks: dict | None = None):
        self.root = root
        self.use_tk_ui = use_tk_ui
        self.ui_callbacks = ui_callbacks or {}

        self.root.title("🧪 Google Whisk Tool")
        self.root.geometry("800x560")
        self.root.resizable(True, True)

        self._apply_theme()

        # Runtime state
        self.driver = None
        self.current_email = None
        self.current_cache_dir = None
        self.current_user_agent = None
        self.login_success = False

        # Execution state & queue (per-account)
        self.stop_exec = False
        self.account_states = {}
        self.exec_drivers = {}
        self.exec_success_count = 0
        self.exec_error_count = 0
        self.job_counter = 0

        # Profiles (cache per email)
        self.whisk_profiles_path = os.path.join(os.getcwd(), "chrome_cache", "whisk_profiles.json")
        self.whisk_profiles = {}
        self._load_profiles()

        # User agents
        self.user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:122.0) Gecko/20100101 Firefox/122.0",
        ]

        if self.use_tk_ui:
            self._build_ui()

        # Route error popups to non-blocking log
        try:
            self._orig_showerror = messagebox.showerror
            def _no_alert_showerror(title, msg):
                try:
                    self._append_log(f"[ERROR] {title}: {msg}\n")
                except Exception:
                    pass
            messagebox.showerror = _no_alert_showerror
        except Exception:
            pass

    # ===================== Theming =====================
    def _apply_theme(self) -> None:
        self.colors = {
            'bg': '#0F1115',
            'surface': '#171A21',
            'border': '#2A2F3A',
            'text': '#EAECEF',
            'subtle': '#9AA4AF',
            'accent': '#58A6FF',
            'accent_hover': '#3E7FD8',
            'success': '#2ECC71',
            'warning': '#F1C40F',
            'error': '#E74C3C',
            'info': '#58A6FF',
        }
        if _HAS_TTKBOOTSTRAP and TtkbStyle is not None:
            self.style = TtkbStyle(theme='superhero')
        else:
            self.style = ttk.Style()
            self.style.theme_use('clam')
        self.root.configure(bg=self.colors['bg'])
        for element in ('TFrame', 'TLabelframe', 'TLabelframe.Label'):
            self.style.configure(element, background=self.colors['bg'], foreground=self.colors['text'])
        self.style.configure('Card.TLabelframe', background=self.colors['surface'], bordercolor=self.colors['border'], borderwidth=1, relief='ridge')
        self.style.configure('TLabel', background=self.colors['bg'], foreground=self.colors['text'], font=('Segoe UI', 10))
        self.style.configure('Title.TLabel', background=self.colors['bg'], foreground=self.colors['accent'], font=('Segoe UI Semibold', 18))
        self.style.configure('Subtitle.TLabel', background=self.colors['bg'], foreground=self.colors['subtle'], font=('Segoe UI', 10, 'bold'))
        self.style.configure('Success.TLabel', background=self.colors['bg'], foreground=self.colors['success'], font=('Segoe UI', 10))
        self.style.configure('Error.TLabel', background=self.colors['bg'], foreground=self.colors['error'], font=('Segoe UI', 10))
        self.style.configure('Info.TLabel', background=self.colors['bg'], foreground=self.colors['info'], font=('Segoe UI', 10))
        self.style.configure('Warning.TLabel', background=self.colors['bg'], foreground=self.colors['warning'], font=('Segoe UI', 10))
        self.style.configure('Accent.TButton', padding=(14, 9), background=self.colors['accent'], foreground='#0B0C10', font=('Segoe UI', 10, 'bold'), borderwidth=0)
        self.style.map('Accent.TButton', background=[('active', self.colors['accent_hover'])])
        self.style.configure('Secondary.TButton', padding=(12, 8), background='#1A1F27', foreground=self.colors['text'], font=('Segoe UI', 10), borderwidth=0)

    # ===================== UI =====================
    def _build_ui(self) -> None:
        notebook = ttk.Notebook(self.root)
        notebook.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        login_tab = ttk.Frame(notebook)
        exec_tab = ttk.Frame(notebook)
        notebook.add(login_tab, text="🔐 Đăng nhập Whisk")
        notebook.add(exec_tab, text="🎥 Execute")
        help_tab = ttk.Frame(notebook)
        notebook.add(help_tab, text="❓ Help")
        notebook.select(1)

        frame = ttk.Frame(login_tab, padding="20")
        frame.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        login_tab.columnconfigure(0, weight=1)
        login_tab.rowconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="🧪 Google Whisk Login", style='Title.TLabel').grid(row=0, column=0, columnspan=2, pady=(0, 16))

        creds = ttk.LabelFrame(frame, text="📝 Thông tin tài khoản", padding="15", style='Card.TLabelframe')
        creds.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 12))
        creds.columnconfigure(1, weight=1)

        ttk.Label(creds, text="📧 Email:", style='Subtitle.TLabel').grid(row=0, column=0, sticky=tk.W, pady=(0, 8))
        self.email_entry = ttk.Entry(creds)
        self.email_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=(0, 8))

        ttk.Label(creds, text="🔒 Mật khẩu:", style='Subtitle.TLabel').grid(row=1, column=0, sticky=tk.W, pady=(0, 8))
        self.password_entry = ttk.Entry(creds, show="*")
        self.password_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=(0, 8))

        self.note_label = ttk.Label(creds, text="💡 Đăng nhập Google (có thể cần 2FA)", style='Info.TLabel')
        self.note_label.grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(6, 0))

        self.login_btn = ttk.Button(frame, text="🚀 Đăng nhập Whisk", command=self._login_whisk, style='Accent.TButton')
        self.login_btn.grid(row=2, column=0, columnspan=2, pady=(14, 0))

        self.status_label = ttk.Label(frame, text="⏳ Chưa đăng nhập", style='Warning.TLabel')
        self.status_label.grid(row=3, column=0, columnspan=2, pady=(10, 0))

        profiles = ttk.LabelFrame(frame, text="👥 Tài khoản đã đăng nhập (cache)", padding="12", style='Card.TLabelframe')
        profiles.grid(row=4, column=0, columnspan=2, sticky=(tk.N, tk.S, tk.W, tk.E), pady=(12, 0))
        profiles.columnconfigure(0, weight=1)
        profiles.rowconfigure(0, weight=1)

        self.profiles_list = tk.Listbox(profiles, height=6, bg=self.colors['surface'], fg=self.colors['text'],
                                        highlightthickness=1, highlightbackground=self.colors['border'],
                                        selectbackground='#253044', selectforeground=self.colors['text'])
        self.profiles_list.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        sb = ttk.Scrollbar(profiles, orient=tk.VERTICAL, command=self.profiles_list.yview)
        sb.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.profiles_list.configure(yscrollcommand=sb.set)

        actions = ttk.Frame(profiles)
        actions.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(8, 0))
        ttk.Button(actions, text="👁️ Mở profile", command=self._open_selected_profile, style='Secondary.TButton').pack(side=tk.LEFT)
        ttk.Button(actions, text="📄 Nhân bản cache", command=self._duplicate_selected_profile, style='Secondary.TButton').pack(side=tk.LEFT, padx=8)
        ttk.Button(actions, text="🗑️ Xóa cache", command=self._delete_selected_profile, style='Secondary.TButton').pack(side=tk.LEFT)
        
        # ===== Help Tab =====
        try:
            help_tab.columnconfigure(0, weight=1)
            help_tab.rowconfigure(0, weight=1)
            help_frame = ttk.Frame(help_tab, padding="20")
            help_frame.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
            help_frame.columnconfigure(0, weight=1)
            ttk.Label(help_frame, text="❓ Trợ giúp", style='Title.TLabel').grid(row=0, column=0, sticky=tk.W, pady=(0, 16))
            desc = (
                "📄 Nhân bản cache: Tạo một profile dựa trên cache hiện có để chạy song song nhiều luồng cho cùng một tài khoản.\n"
                "- Khi bấm 'Nhân bản cache', công cụ sẽ sao chép thư mục cache hiện tại (bỏ qua các file lock)\n"
                "- Profile mới được đặt tên kèm timestamp và xuất hiện trong danh sách tài khoản\n"
                "- Có thể mở nhiều profile cùng lúc để thực thi song song\n\n"
                "🔐 Đăng nhập Google: Theo dõi tiến trình đăng nhập hiển thị trên giao diện.\n"
                "- Tool sẽ chủ động hỗ trợ tối đa khi Google yêu cầu CAPTCHA/OTP/2FA (nếu có thể tự động).\n"
                "- Trường hợp cần thao tác thủ công (nhập mã OTP/2FA), hãy hoàn tất trong trình duyệt đang mở; tool sẽ tự phát hiện và tiếp tục."
            )
            lbl = ttk.Label(help_frame, text=desc, style='Info.TLabel', justify='left')
            lbl.grid(row=1, column=0, sticky=(tk.W, tk.E))
        except Exception:
            pass

        log_box = ttk.LabelFrame(frame, text="📜 Log", padding="10", style='Card.TLabelframe')
        log_box.grid(row=5, column=0, columnspan=2, sticky=(tk.N, tk.S, tk.W, tk.E), pady=(12, 0))
        log_box.columnconfigure(0, weight=1)
        log_box.rowconfigure(0, weight=1)
        self.log = scrolledtext.ScrolledText(log_box, height=6, wrap=tk.WORD, state='disabled',
                                             bg='#10141B', fg=self.colors['text'])
        self.log.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))

        self._refresh_profiles_list()

        # ===== Execute Tab =====
        exec_tab.columnconfigure(0, weight=1)
        exec_tab.rowconfigure(0, weight=1)
        ex = ttk.Frame(exec_tab, padding="20")
        ex.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        ex.columnconfigure(0, weight=3)
        ex.columnconfigure(1, weight=2)
        ex.rowconfigure(5, weight=1)

        ttk.Label(ex, text="🎥 Execute Whisk", style='Title.TLabel').grid(row=0, column=0, pady=(0, 16), sticky=tk.W)

        # Config: Headless toggle
        cfg = ttk.LabelFrame(ex, text="⚙️ Cấu hình", padding="12", style='Card.TLabelframe')
        cfg.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 12))
        self.headless_mode = tk.BooleanVar(value=True)
        ttk.Checkbutton(cfg, text="Headless (ẩn browser)", variable=self.headless_mode).grid(row=0, column=0, sticky=tk.W)

        # Actions
        actions_ex = ttk.Frame(ex)
        actions_ex.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 8))
        actions_ex.columnconfigure(0, weight=1)
        actions_ex.columnconfigure(1, weight=1)
        actions_ex.columnconfigure(2, weight=1)
        ttk.Button(actions_ex, text="📥 Import Excel", command=self._import_excel_whisk, style='Secondary.TButton').grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 6))
        ttk.Button(actions_ex, text="⬇️ Tải Template", command=self._download_excel_template_whisk, style='Secondary.TButton').grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(6, 6))

        # Status + Log
        self.exec_status = ttk.Label(ex, text="✅ Sẵn sàng", style='Success.TLabel')
        self.exec_status.grid(row=3, column=0, sticky=tk.W)

        log_frame = ttk.LabelFrame(ex, text="📜 Log tiến trình", padding="10", style='Card.TLabelframe')
        log_frame.grid(row=4, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        self.exec_log = scrolledtext.ScrolledText(log_frame, height=8, wrap=tk.WORD, state='disabled',
                                                  bg='#10141B', fg=self.colors['text'])
        self.exec_log.grid(row=0, column=0, sticky=(tk.N, tk.S, tk.W, tk.E))

        # Right side: Progress counters
        side = ttk.LabelFrame(ex, text="📊 Tiến trình", padding="10", style='Card.TLabelframe')
        side.grid(row=0, column=1, rowspan=5, sticky=(tk.N, tk.S, tk.W, tk.E), padx=(12, 0))
        side.columnconfigure(0, weight=1)
        ttk.Label(side, text="In queue:", style='Subtitle.TLabel').grid(row=0, column=0, sticky=tk.W, pady=(0,4))
        self.lbl_queue = ttk.Label(side, text="0", style='Info.TLabel')
        self.lbl_queue.grid(row=1, column=0, sticky=tk.W)
        ttk.Separator(side, orient=tk.HORIZONTAL).grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(8,8))
        ttk.Label(side, text="Executing:", style='Subtitle.TLabel').grid(row=3, column=0, sticky=tk.W, pady=(0,4))
        self.lbl_exec = ttk.Label(side, text="0", style='Info.TLabel')
        self.lbl_exec.grid(row=4, column=0, sticky=tk.W)
        ttk.Separator(side, orient=tk.HORIZONTAL).grid(row=5, column=0, sticky=(tk.W, tk.E), pady=(8,8))
        ttk.Label(side, text="Success:", style='Subtitle.TLabel').grid(row=6, column=0, sticky=tk.W, pady=(0,4))
        self.lbl_ok = ttk.Label(side, text="0", style='Success.TLabel')
        self.lbl_ok.grid(row=7, column=0, sticky=tk.W)
        ttk.Separator(side, orient=tk.HORIZONTAL).grid(row=8, column=0, sticky=(tk.W, tk.E), pady=(8,8))
        ttk.Label(side, text="Error:", style='Subtitle.TLabel').grid(row=9, column=0, sticky=tk.W, pady=(0,4))
        self.lbl_err = ttk.Label(side, text="0", style='Error.TLabel')
        self.lbl_err.grid(row=10, column=0, sticky=tk.W)
        # Initialize counters
        try:
            self._update_exec_counters()
        except Exception:
            pass

    # ===================== Profiles =====================
    def _load_profiles(self) -> None:
        try:
            os.makedirs(os.path.dirname(self.whisk_profiles_path), exist_ok=True)
            if os.path.exists(self.whisk_profiles_path):
                with open(self.whisk_profiles_path, 'r', encoding='utf-8') as f:
                    self.whisk_profiles = json.load(f)
            else:
                self.whisk_profiles = {}
        except Exception:
            self.whisk_profiles = {}

    def _save_profiles(self) -> None:
        try:
            os.makedirs(os.path.dirname(self.whisk_profiles_path), exist_ok=True)
            with open(self.whisk_profiles_path, 'w', encoding='utf-8') as f:
                json.dump(self.whisk_profiles, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _refresh_profiles_list(self) -> None:
        try:
            self.profiles_list.delete(0, tk.END)
            items = sorted(self.whisk_profiles.items(), key=lambda kv: kv[1].get("last_login", 0), reverse=True)
            for email_addr, meta in items:
                ts = meta.get("last_login")
                time_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts)) if ts else ""
                self.profiles_list.insert(tk.END, f"{email_addr}  |  {time_str}")
        except Exception:
            pass

    # ===================== Login =====================
    def _login_whisk(self) -> None:
        email = (self.email_entry.get() or "").strip()
        if not email:
            messagebox.showerror("Lỗi", "Vui lòng nhập email!")
            return
        self.current_email = email
        password = (self.password_entry.get() or "").strip()
        if not password:
            messagebox.showerror("Lỗi", "Vui lòng nhập mật khẩu!")
            return
        threading.Thread(target=self._login_password_thread, args=(email, password), daemon=True).start()

    def _build_chrome(self, cache_key: str, existing_cache_dir: str | None = None) -> webdriver.Chrome:
        chrome_options = Options()
        # Prefer existing cache dir if available
        cache_dir = None
        if existing_cache_dir and os.path.isdir(existing_cache_dir):
            cache_dir = existing_cache_dir
        else:
            safe_key = re.sub(r'[^a-zA-Z0-9_.-]', '_', cache_key) or "default"
            cache_dir = os.path.join(os.getcwd(), "chrome_cache", f"whisk_{safe_key}")
            os.makedirs(cache_dir, exist_ok=True)
        self.current_cache_dir = cache_dir

        # Simple lock detection for single instance of profile
        try:
            lock_file = os.path.join(cache_dir, "SingletonLock")
            if os.path.exists(lock_file):
                raise Exception("Profile cache đang được sử dụng bởi phiên Chrome khác. Hãy đóng trình duyệt rồi thử lại.")
        except Exception:
            pass

        chrome_options.add_argument(f"--user-data-dir={cache_dir}")
        chrome_options.add_argument("--profile-directory=Default")
        ua = random.choice(self.user_agents)
        self.current_user_agent = ua
        chrome_options.add_argument(f"--user-agent={ua}")
        chrome_options.add_argument("--lang=vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7")
        chrome_options.add_argument("--start-maximized")
        # Set consistent window size for better compatibility
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_experimental_option("detach", True)
        # Further reduce automation fingerprints
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")

        try:
            chrome_options.set_capability('goog:loggingPrefs', {'performance': 'ALL'})
        except Exception:
            pass

        service = Service(ChromeDriverManager().install())
        
        # Retry logic for Chrome crashes
        max_retries = 3
        for attempt in range(max_retries):
            try:
                driver = webdriver.Chrome(service=service, options=chrome_options)
                try:
                    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
                except Exception:
                    pass
                try:
                    driver.execute_cdp_cmd('Network.enable', {})
                except Exception:
                    pass
                return driver
            except Exception as e:
                error_msg = str(e).lower()
                if "chrome failed to start" in error_msg or "crashed" in error_msg or "session not created" in error_msg:
                    if attempt < max_retries - 1:
                        try:
                            self._append_log(f"[SYSTEM] {time.strftime('%H:%M:%S')} | Chrome crash detected, đóng browser và thử lại... (lần {attempt + 1}/{max_retries})\n")
                        except Exception:
                            pass
                        # Kill any remaining Chrome processes
                        self._kill_chrome_processes()
                        time.sleep(2)  # Wait before retry
                        continue
                    else:
                        try:
                            self._append_log(f"[ERROR] {time.strftime('%H:%M:%S')} | Chrome crash sau {max_retries} lần thử: {e}\n")
                        except Exception:
                            pass
                        raise e
                else:
                    # Re-raise non-crash errors immediately
                    raise e
        
        # Should never reach here, but just in case
        raise Exception("Không thể khởi tạo Chrome sau nhiều lần thử")

    def _login_password_thread(self, email_addr: str, password: str) -> None:
        try:
            self._set_status("Đang mở trình duyệt...", "orange")
            self.login_btn.config(state="disabled")
            meta = self.whisk_profiles.get(email_addr)
            exist_dir = meta.get("cache_dir") if meta else None
            self.driver = self._build_chrome(email_addr, existing_cache_dir=exist_dir)

            self.driver.get("https://accounts.google.com/signin")
            self._human_delay(1.5, 3.0)
            self._human_warm_up_page(self.driver)
            self._google_type_email_then_password(self.driver, email_addr, password)

            ok = self._wait_signin_success(self.driver, timeout=240)
            if not ok:
                # Retry via AccountChooser as a fallback for "This browser or app may not be secure"
                try:
                    self._set_status("Thử lại đăng nhập qua AccountChooser...", "orange")
                    self.driver.get("https://accounts.google.com/AccountChooser?continue=https://labs.google/fx/vi/tools/whisk")
                    self._human_delay(1.0, 2.0)
                    self._human_warm_up_page(self.driver)
                    self._google_type_email_then_password(self.driver, email_addr, password)
                    ok = self._wait_signin_success(self.driver, timeout=180)
                except Exception:
                    ok = False
                if not ok:
                    raise Exception("Không thể đăng nhập tài khoản Google")

            # After signed in, open Whisk
            self.driver.get(WHISK_URL)
            self._wait_until(lambda: "labs.google" in (self.driver.current_url or ""), timeout=120)

            # Close welcome modal if it appears
            try:
                self._close_welcome_modal(self.driver)
            except Exception:
                pass

            # If Whisk has a final "Sign in with Google" gate, click it
            try:
                self._click_whisk_google_signin(self.driver)
            except Exception:
                pass

            self._remember_profile(email_addr, self.current_cache_dir, self.current_user_agent)
            self.login_success = True
            self._set_status("Đăng nhập Whisk thành công", "green")
            messagebox.showinfo("Thành công", "Đăng nhập Google Whisk thành công!")
        except Exception as ex:
            self.login_success = False
            self._set_status("Đăng nhập thất bại", "red")
            messagebox.showerror("Lỗi", f"Lỗi đăng nhập: {ex}")
        finally:
            try:
                if self.driver is not None:
                    self.driver.quit()
            except Exception:
                pass
            self.driver = None
            self.login_btn.config(state="normal")

    # ===================== Selenium helpers =====================
    def _close_welcome_modal(self, driver: webdriver.Chrome, timeout: int = 10) -> bool:
        """Close the welcome modal if it appears. Returns True if modal was found and closed."""
        try:
            # Look for the close button with aria-label "Close this modal"
            close_button_xpaths = [
                "//button[@aria-label='Close this modal']",
                "//button[contains(@aria-label, 'Close this modal')]",
                "//button[.//i[normalize-space(text())='close']]",
                "//button[contains(@class, 'kzGuMg')]",
                "//button[contains(@class, 'sc-') and .//i[normalize-space(text())='close']]",
            ]
            
            end_time = time.time() + timeout
            while time.time() < end_time:
                for xp in close_button_xpaths:
                    try:
                        el = driver.find_element(By.XPATH, xp)
                        if el.is_displayed() and el.is_enabled():
                            self._human_click_el(driver, el)
                            try:
                                self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Đã đóng welcome modal\n")
                            except Exception:
                                pass
                            return True
                    except Exception:
                        continue
                time.sleep(0.5)
            return False
        except Exception:
            return False

    def _click_whisk_google_signin(self, driver: webdriver.Chrome) -> None:
        candidates = [
            "//button[.//span[normalize-space()='Sign in with Google']]",
            "//button[contains(., 'Sign in with Google')]",
            "//button[contains(@class,'sc-') and contains(., 'Sign in with Google')]",
            "//button[.//span[normalize-space()='Đăng nhập bằng Google']]",
            "//button[contains(., 'Đăng nhập bằng Google')]",
        ]
        self._human_delay(0.5, 1.2)
        for xp in candidates:
            try:
                el = driver.find_element(By.XPATH, xp)
                try:
                    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xp)))
                except Exception:
                    pass
                self._human_click_el(driver, el)
                time.sleep(0.4)
                return
            except Exception:
                continue

    def _google_type_email_then_password(self, driver: webdriver.Chrome, email_addr: str, password: str) -> None:
        wait = WebDriverWait(driver, 30)
        wait.until(EC.presence_of_element_located((By.NAME, "identifier")))
        email_input = driver.find_element(By.NAME, "identifier")
        self._human_click_el(driver, email_input)
        email_input.clear()
        self._human_type_el(email_input, email_addr)
        driver.find_element(By.ID, "identifierNext").click()
        self._human_delay(1.5, 3.0)
        try:
            wait.until(EC.presence_of_element_located((By.NAME, "Passwd")))
            pw_input = driver.find_element(By.NAME, "Passwd")
        except TimeoutException:
            pw_input = driver.find_element(By.CSS_SELECTOR, "input[type='password']")
        self._human_click_el(driver, pw_input)
        pw_input.clear()
        self._human_type_el(pw_input, password)
        self._human_delay(0.4, 1.0)
        try:
            driver.find_element(By.ID, "passwordNext").click()
        except Exception:
            pw_input.send_keys(Keys.ENTER)

    def _wait_signin_success(self, driver: webdriver.Chrome, timeout: int = 180) -> bool:
        start = time.time()
        while time.time() - start < timeout:
            try:
                if self._is_google_signed_in(driver):
                    return True
            except Exception:
                pass
            time.sleep(1)
        return False

    def _is_google_signed_in(self, driver: webdriver.Chrome) -> bool:
        try:
            url = driver.current_url or ""
            if url.startswith("https://myaccount.google.com/"):
                return True
            selectors = [
                "a[aria-label^='Google Account:' i]",
                "a[aria-label^='Tài khoản Google:' i]",
                "img.gb_P.gbii",
            ]
            for sel in selectors:
                if driver.find_elements(By.CSS_SELECTOR, sel):
                    return True
        except Exception:
            return False
        return False

    def _human_delay(self, min_seconds: float = 1.0, max_seconds: float = 2.0) -> None:
        time.sleep(random.uniform(min_seconds, max_seconds))

    def _human_warm_up_page(self, driver: webdriver.Chrome) -> None:
        try:
            # random scrolls
            for _ in range(random.randint(1, 3)):
                dy = random.randint(100, 400) * (1 if random.random() < 0.5 else -1)
                driver.execute_script("window.scrollBy(0, arguments[0]);", dy)
                time.sleep(random.uniform(0.1, 0.4))
            # slight mouse moves over body if possible
            try:
                body = driver.find_element(By.TAG_NAME, 'body')
                ActionChains(driver).move_to_element_with_offset(body, random.randint(5, 50), random.randint(5, 30)).pause(random.uniform(0.05, 0.2)).perform()
            except Exception:
                pass
        except Exception:
            pass

    def _human_click_el(self, driver: webdriver.Chrome, element) -> None:
        try:
            ActionChains(driver).move_to_element(element).pause(random.uniform(0.05, 0.2)).click().perform()
        except Exception:
            try:
                element.click()
            except Exception:
                pass

    def _human_type_el(self, element, text: str) -> None:
        try:
            # Replace newlines with spaces to prevent accidental form submission
            safe_text = text.replace('\n', ' ').replace('\r', ' ')
            # Type the full text at once for immediate parsing (no per-char delay)
            element.send_keys(safe_text)
        except Exception:
            # Best-effort fallback: ignore if element refuses bulk input
            try:
                safe_text = text.replace('\n', ' ').replace('\r', ' ')
                element.send_keys(safe_text)
            except Exception:
                pass

    def _type_prompt_into_any_textarea(self, driver: webdriver.Chrome, text: str, timeout: int = 15) -> bool:
        """Type text into a visible <textarea> (do not rely on class selectors)."""
        if not text:
            return False
        try:
            # Prefer any visible textarea
            wait = WebDriverWait(driver, timeout)
            # Try a straightforward textarea
            area = None
            try:
                area = wait.until(EC.presence_of_element_located((By.TAG_NAME, 'textarea')))
            except Exception:
                area = None
            # If multiple, pick the first visible and enabled
            if area is None:
                candidates = driver.find_elements(By.TAG_NAME, 'textarea')
            else:
                candidates = [area] + [el for el in driver.find_elements(By.TAG_NAME, 'textarea') if el is not area]
            for el in candidates:
                try:
                    if not el.is_displayed() or not el.is_enabled():
                        continue
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                    time.sleep(0.2)
                    self._human_click_el(driver, el)
                    try:
                        el.clear()
                        time.sleep(0.05)
                    except Exception:
                        pass
                    self._human_type_el(el, text)
                    return True
                except Exception:
                    continue
        except Exception:
            pass
        try:
            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không tìm thấy textarea hiển thị để nhập văn bản\n")
        except Exception:
            pass
        return False

    def _click_add_image_button(self, driver: webdriver.Chrome, timeout: int = 5) -> bool:
        """Click the button that contains text 'Thêm hình ảnh' without relying on classes."""
        xpaths = [
            "//button[.//span[normalize-space(text())='Thêm hình ảnh']]",
            "//button[normalize-space(.)='Thêm hình ảnh']",
            "//button[contains(., 'Thêm hình ảnh')]",
            "//*[@role='button'][.//span[normalize-space(text())='Thêm hình ảnh']]",
            "//*[@role='button'][contains(., 'Thêm hình ảnh')]",
        ]
        end_time = time.time() + timeout
        while time.time() < end_time:
            for xp in xpaths:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    if not el.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        time.sleep(0.1)
                    self._human_click_el(driver, el)
                    return True
                except Exception:
                    continue
            time.sleep(0.2)
        try:
            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không tìm thấy nút 'Thêm hình ảnh'\n")
        except Exception:
            pass
        return False

    def _click_button_by_text(self, driver: webdriver.Chrome, text: str, timeout: int = 8) -> bool:
        """Click a button by its visible text without relying on classes."""
        if not text:
            return False
        xpaths = [
            f"//button[normalize-space(.)='{text}']",
            f"//button[.//span[normalize-space(text())='{text}']]",
            f"//button[contains(., '{text}')]",
            f"//*[@role='button' and normalize-space(.)='{text}']",
            f"//*[@role='button'][.//span[normalize-space(text())='{text}']]",
            f"//*[normalize-space(text())='{text}']/ancestor::*[self::button or @role='button'][1]",
        ]
        end_time = time.time() + timeout
        while time.time() < end_time:
            for xp in xpaths:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    if not el.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        time.sleep(0.1)
                    self._human_click_el(driver, el)
                    return True
                except Exception:
                    continue
            time.sleep(0.25)
        try:
            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không tìm thấy nút '{text}'\n")
        except Exception:
            pass
        return False

    def _click_button_by_aria_label(self, driver: webdriver.Chrome, aria_label: str, timeout: int = 6) -> bool:
        """Click a button by its aria-label attribute; fallback to icon text if needed."""
        if not aria_label:
            return False
        xpaths = [
            f"//button[@aria-label='{aria_label}']",
            f"//*[@role='button' and @aria-label='{aria_label}']",
        ]
        end_time = time.time() + timeout
        while time.time() < end_time:
            for xp in xpaths:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    if not el.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        time.sleep(0.1)
                    self._human_click_el(driver, el)
                    return True
                except Exception:
                    continue
            time.sleep(0.2)
        try:
            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không tìm thấy nút aria-label='{aria_label}'\n")
        except Exception:
            pass
        return False

    def _open_aspect_ratio_menu(self, driver: webdriver.Chrome, timeout: int = 6) -> bool:
        """Open the aspect ratio popover by clicking the trigger button with the aspect_ratio icon."""
        xpaths = [
            "//button[@aria-haspopup='dialog' and .//i[normalize-space(text())='aspect_ratio']]",
            "//*[@role='button' and @aria-haspopup='dialog' and .//i[normalize-space(text())='aspect_ratio']]",
            "//button[.//i[normalize-space(text())='aspect_ratio']]",
        ]
        end_time = time.time() + timeout
        while time.time() < end_time:
            for xp in xpaths:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    if not el.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        time.sleep(0.1)
                    self._human_click_el(driver, el)
                    return True
                except Exception:
                    continue
            time.sleep(0.2)
        try:
            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không tìm thấy nút mở chọn Tỷ lệ khung hình\n")
        except Exception:
            pass
        return False

    def _normalize_aspect_ratio(self, size_text: str) -> str:
        """Normalize aspect ratio format: 16:09 -> 16:9, 09:16 -> 9:16, 16:09:00 -> 16:9, etc."""
        if not size_text:
            return size_text
        
        try:
            # Handle time format like "16:09:00" -> "16:9"
            if ':' in size_text:
                parts = size_text.split(':')
                if len(parts) >= 2:
                    # Take only first two parts (hours:minutes), ignore seconds
                    left = str(int(parts[0])) if parts[0].isdigit() else parts[0]
                    right = str(int(parts[1])) if parts[1].isdigit() else parts[1]
                    return f"{left}:{right}"
        except Exception:
            pass
        
        return size_text

    def _pick_aspect_ratio(self, driver: webdriver.Chrome, size_text: str, timeout: int = 6) -> bool:
        """Pick the aspect ratio option by visible text like '1:1', '9:16', '16:9', '3:4', '4:3'."""
        if not (size_text or '').strip():
            return False
        
        # Normalize size format: convert 16:09 -> 16:9, 09:16 -> 9:16, etc.
        label = (size_text or '').strip()
        normalized_label = self._normalize_aspect_ratio(label)
        
        # Try both original and normalized formats
        search_labels = [label, normalized_label]
        if label != normalized_label:
            search_labels = [normalized_label, label]  # Try normalized first
        
        xpaths = []
        for search_label in search_labels:
            # Generic XPaths that don't rely on specific CSS classes
            xpaths.extend([
                # Button containing span with exact text
                f"//button[.//span[normalize-space(text())='{search_label}']]",
                # Any element with role button containing span with text
                f"//*[@role='button' and .//span[normalize-space(text())='{search_label}']]",
                # Button containing any element with exact text
                f"//button[.//*[normalize-space(text())='{search_label}']]",
                # Generic fallbacks
                f"//div[@role='dialog']//button[.//span[normalize-space(text())='{search_label}']]",
                f"//div[@role='dialog']//span[normalize-space(text())='{search_label}']/ancestor::button[1]",
                # Clickable element containing the text
                f"//*[normalize-space(text())='{search_label}']/ancestor::*[self::button or @role='button'][1]",
            ])
        
        end_time = time.time() + timeout
        while time.time() < end_time:
            for xp in xpaths:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    if not el.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        time.sleep(0.05)
                    self._human_click_el(driver, el)
                    try:
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Đã chọn size: {search_label}\n")
                    except Exception:
                        pass
                    return True
                except Exception:
                    continue
            time.sleep(0.2)
        try:
            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không tìm thấy tùy chọn size '{label}' (đã thử: {', '.join(search_labels)})\n")
        except Exception:
            pass
        return False

    def _open_tune_menu(self, driver: webdriver.Chrome, timeout: int = 6) -> bool:
        """Open the tune/settings popover by clicking the trigger button with the tune icon."""
        xpaths = [
            "//button[@aria-haspopup='dialog' and .//i[normalize-space(text())='tune']]",
            "//*[@role='button' and @aria-haspopup='dialog' and .//i[normalize-space(text())='tune']]",
            "//button[.//i[normalize-space(text())='tune']]",
        ]
        end_time = time.time() + timeout
        while time.time() < end_time:
            for xp in xpaths:
                try:
                    el = driver.find_element(By.XPATH, xp)
                    if not el.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                        time.sleep(0.1)
                    self._human_click_el(driver, el)
                    return True
                except Exception:
                    continue
            time.sleep(0.2)
        try:
            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không tìm thấy nút mở tune/settings\n")
        except Exception:
            pass
        return False

    def _set_random_seed(self, driver: webdriver.Chrome, timeout: int = 6) -> bool:
        """Set a random 6-digit number into the seed input (id=whisk-seed-input)."""
        try:
            seed_val = str(random.randint(100000, 999999))
            end_time = time.time() + timeout
            while time.time() < end_time:
                try:
                    inp = driver.find_element(By.ID, 'whisk-seed-input')
                    if not inp.is_displayed():
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
                        time.sleep(0.05)
                    self._human_click_el(driver, inp)
                    try:
                        inp.clear()
                    except Exception:
                        pass
                    inp.send_keys(seed_val)
                    try:
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Đặt seed: {seed_val}\n")
                    except Exception:
                        pass
                    return True
                except Exception:
                    time.sleep(0.2)
        except Exception:
            pass
        try:
            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không tìm thấy input seed (whisk-seed-input)\n")
        except Exception:
            pass
        return False

    def _click_submit_execute(self, driver: webdriver.Chrome, timeout: int = 10) -> bool:
        """Click the final execute/submit button labeled 'Gửi câu lệnh' when it becomes enabled."""
        end_time = time.time() + timeout
        while time.time() < end_time:
            try:
                # Multiple strategies to find the submit button
                candidates = []
                
                # Strategy 1: Look for the specific button structure with aria-label and arrow_forward icon
                try:
                    specific_buttons = driver.find_elements(By.XPATH,
                        "//button[@type='submit' and @aria-label='Gửi câu lệnh' and .//i[normalize-space(text())='arrow_forward']]"
                    )
                    candidates.extend(specific_buttons)
                except Exception:
                    pass
                
                # Strategy 2: Look for button with aria-label and submit type
                try:
                    aria_buttons = driver.find_elements(By.XPATH,
                        "//button[@type='submit' and @aria-label='Gửi câu lệnh']"
                    )
                    candidates.extend(aria_buttons)
                except Exception:
                    pass
                
                # Strategy 3: Look for button containing arrow_forward icon
                try:
                    icon_buttons = driver.find_elements(By.XPATH,
                        "//button[@type='submit' and .//i[normalize-space(text())='arrow_forward']]"
                    )
                    candidates.extend(icon_buttons)
                except Exception:
                    pass
                
                # Strategy 4: Look for button containing 'Gửi câu lệnh' text
                try:
                    text_buttons = driver.find_elements(By.XPATH,
                        "//button[@type='submit' and contains(., 'Gửi câu lệnh')]"
                    )
                    candidates.extend(text_buttons)
                except Exception:
                    pass
                
                # Strategy 5: Look for any button with submit type and role button
                try:
                    generic_buttons = driver.find_elements(By.XPATH,
                        "//*[@role='button' and @aria-label='Gửi câu lệnh'] | "
                        "//button[contains(., 'Gửi câu lệnh')]"
                    )
                    candidates.extend(generic_buttons)
                except Exception:
                    pass
                
                # Remove duplicates and check each candidate
                seen_elements = set()
                for btn in candidates:
                    try:
                        # Create a unique identifier for this element
                        element_id = id(btn)
                        if element_id in seen_elements:
                            continue
                        seen_elements.add(element_id)
                        
                        # Check if button is enabled and clickable
                        disabled_attr = btn.get_attribute('disabled')
                        data_state = btn.get_attribute('data-state')
                        is_enabled = (disabled_attr is None) and btn.is_enabled()
                        
                        # Also check if button is visible and interactable
                        if not btn.is_displayed():
                            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                            time.sleep(0.1)
                        
                        # Try to click if enabled
                        if is_enabled:
                            self._human_click_el(driver, btn)
                            try:
                                self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Đã nhấn 'Gửi câu lệnh'\n")
                            except Exception:
                                pass
                            return True
                        else:
                            # Log why button is not clickable
                            try:
                                self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Nút 'Gửi câu lệnh' chưa sẵn sàng (disabled={disabled_attr}, enabled={btn.is_enabled()})\n")
                            except Exception:
                                pass
                    except Exception:
                        continue
            except Exception:
                pass
            time.sleep(0.25)
        try:
            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không thể nhấn 'Gửi câu lệnh' sau {timeout}s (có thể đang disabled hoặc không tìm thấy)\n")
        except Exception:
            pass
        return False

    def _sanitize_filename(self, text: str, max_length: int = 50) -> str:
        """Convert prompt text to a safe filename by removing/sanitizing special characters."""
        if not text:
            return "no_prompt"
        
        # Remove or replace problematic characters
        import re
        # Keep only alphanumeric, spaces, hyphens, underscores
        sanitized = re.sub(r'[^\w\s\-_]', '', text)
        # Replace multiple spaces with single space
        sanitized = re.sub(r'\s+', ' ', sanitized)
        # Replace spaces with underscores
        sanitized = sanitized.replace(' ', '_')
        # Remove leading/trailing underscores
        sanitized = sanitized.strip('_')
        # Limit length
        if len(sanitized) > max_length:
            sanitized = sanitized[:max_length].rstrip('_')
        # Ensure it's not empty
        if not sanitized:
            sanitized = "prompt"
            
        return sanitized

    def _kill_chrome_processes(self) -> None:
        """Kill any remaining Chrome processes to resolve crashes."""
        try:
            import subprocess
            import platform
            system = platform.system().lower()
            if system == "windows":
                subprocess.run(["taskkill", "/f", "/im", "chrome.exe"], 
                             capture_output=True, timeout=5)
                subprocess.run(["taskkill", "/f", "/im", "chromedriver.exe"], 
                             capture_output=True, timeout=5)
            elif system in ["linux", "darwin"]:
                subprocess.run(["pkill", "-f", "chrome"], 
                             capture_output=True, timeout=5)
                subprocess.run(["pkill", "-f", "chromedriver"], 
                             capture_output=True, timeout=5)
        except Exception:
            pass

    def _take_error_screenshot(self, driver: webdriver.Chrome, prompt_text: str = "", error_reason: str = "") -> str | None:
        """Take a screenshot and save it to whisk_errors folder for debugging."""
        try:
            # Create error directory
            error_dir = os.path.join(os.getcwd(), 'whisk_errors')
            os.makedirs(error_dir, exist_ok=True)
            
            # Generate filename
            base_filename = self._sanitize_filename(prompt_text, max_length=30)
            timestamp = int(time.time())
            error_reason_clean = self._sanitize_filename(error_reason, max_length=20)
            filename = f"error_{base_filename}_{error_reason_clean}_{timestamp}.png"
            file_path = os.path.join(error_dir, filename)
            
            # Take screenshot
            driver.save_screenshot(file_path)
            
            # Log the screenshot
            try:
                self._append_exec_log(f"[ERROR] {time.strftime('%H:%M:%S')} | Đã chụp screenshot lỗi: {filename}\n")
            except Exception:
                pass
                
            return file_path
        except Exception as e:
            try:
                self._append_exec_log(f"[ERROR] {time.strftime('%H:%M:%S')} | Không thể chụp screenshot lỗi: {str(e)}\n")
            except Exception:
                pass
            return None

    def _download_result_images(self, driver: webdriver.Chrome, wait_seconds: int = 30, max_images: int = 10, skip_count: int = 0, prompt_text: str = "") -> int:
        """Wait then find <img src="blob:https://labs.google/..."> and download images to local folder. Returns count saved."""
        screenshot_taken = False  # Flag to prevent duplicate screenshots
        
        try:
            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Đợi {wait_seconds}s để tải ảnh kết quả...\n")
            time.sleep(max(0, wait_seconds))
            
            # Look for blob URLs specifically from labs.google
            imgs = driver.find_elements(By.CSS_SELECTOR, "img[src^='blob:https://labs.google']")
            
            # Fallback to any blob URL if labs.google specific ones not found
            if not imgs:
                imgs = driver.find_elements(By.CSS_SELECTOR, "img[src^='blob:']")
                
        except Exception:
            imgs = []
        if not imgs:
            try:
                self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không tìm thấy ảnh kết quả (blob:)\n")
                # Take screenshot for debugging
                if not screenshot_taken:
                    self._take_error_screenshot(driver, prompt_text, "no_images_found")
                    screenshot_taken = True
            except Exception:
                pass
            return 0
        
        # Check if we have enough new images (at least 2 more than uploaded)
        max_retries = 3
        retry_count = 0
        wait_time = 10
        
        while retry_count <= max_retries:
            try:
                total_found = len(imgs)
                new_images_count = max(0, total_found - skip_count)
                
                if skip_count > 0:
                    self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Tìm thấy {total_found} ảnh blob, bỏ qua {skip_count} ảnh đã upload, có {new_images_count} ảnh mới...\n")
                else:
                    self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Tìm thấy {total_found} ảnh blob, bắt đầu tải...\n")
                
                # Check if we have at least 2 new images (or no uploaded images)
                if skip_count == 0 or new_images_count >= 2:
                    if skip_count > 0:
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | ✅ Đủ ảnh mới ({new_images_count}), bắt đầu tải...\n")
                    break
                else:
                    if retry_count < max_retries:
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | ⚠️ Chưa đủ ảnh mới ({new_images_count} < 2), đợi thêm {wait_time}s... (lần {retry_count + 1}/{max_retries})\n")
                        time.sleep(wait_time)
                        
                        # Re-scan for images
                        try:
                            imgs = driver.find_elements(By.CSS_SELECTOR, "img[src^='blob:https://labs.google']")
                            if not imgs:
                                imgs = driver.find_elements(By.CSS_SELECTOR, "img[src^='blob:']")
                        except Exception:
                            imgs = []
                        
                        retry_count += 1
                    else:
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | ❌ Hết số lần thử ({max_retries}), chỉ có {new_images_count} ảnh mới, tiếp tục tải...\n")
                        # Take screenshot for debugging insufficient images
                        if not screenshot_taken:
                            self._take_error_screenshot(driver, prompt_text, f"insufficient_images_{new_images_count}")
                            screenshot_taken = True
                        break
                        
            except Exception:
                break
        # Ensure output directory
        out_dir = os.path.join(os.getcwd(), 'whisk_downloads')
        try:
            os.makedirs(out_dir, exist_ok=True)
        except Exception:
            pass
        saved = 0
        for i, img in enumerate(imgs):
            # Skip the first skip_count images (these are uploaded images)
            if i < skip_count:
                try:
                    src = img.get_attribute('src') or ''
                    self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | ⏭️ Bỏ qua ảnh {i+1} (đã upload): {src[:50]}...\n")
                except Exception:
                    pass
                continue
                
            if (i - skip_count) >= max_images:
                break
                
            try:
                src = img.get_attribute('src') or ''
                if not src.startswith('blob:'):
                    continue
                    
                # Log the blob URL being processed
                try:
                    self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Xử lý ảnh {i+1}: {src[:50]}...\n")
                except Exception:
                    pass
                
                data_url = self._fetch_blob_data_url(driver, src)
                if not data_url:
                    try:
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không thể fetch blob data cho ảnh {i+1}\n")
                    except Exception:
                        pass
                    continue
                    
                # Parse data URL
                if not data_url.startswith('data:'):
                    try:
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Data URL không hợp lệ cho ảnh {i+1}\n")
                    except Exception:
                        pass
                    continue
                    
                header, b64 = data_url.split(',', 1)
                ext = 'png'
                if 'image/' in header:
                    try:
                        ext = header.split('image/')[1].split(';')[0]
                    except Exception:
                        ext = 'png'
                
                # Generate filename based on prompt
                base_filename = self._sanitize_filename(prompt_text)
                timestamp = int(time.time())
                file_index = i - skip_count + 1  # Start from 1 for generated images
                file_name = f"{base_filename}_{timestamp}_{file_index}.{ext}"
                file_path = os.path.join(out_dir, file_name)
                
                with open(file_path, 'wb') as f:
                    f.write(base64.b64decode(b64))
                saved += 1
                
                try:
                    self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | ✅ Đã lưu ảnh {i+1}: {file_path}\n")
                except Exception:
                    pass
                    
            except Exception as e:
                try:
                    self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | ❌ Lỗi xử lý ảnh {i+1}: {str(e)}\n")
                except Exception:
                    pass
                continue
        if saved == 0:
            try:
                self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không thể tải về ảnh blob\n")
                # Take screenshot for debugging download failure
                if not screenshot_taken:
                    self._take_error_screenshot(driver, prompt_text, "download_failed")
                    screenshot_taken = True
            except Exception:
                pass
        return saved

    def _fetch_blob_data_url(self, driver: webdriver.Chrome, blob_url: str, timeout: int = 10) -> str | None:
        """Fetch a blob: URL in page context and return a data URL (base64)."""
        try:
            script = (
                "var url = arguments[0]; var cb = arguments[1];"
                "fetch(url).then(r=>r.blob()).then(b=>{var fr=new FileReader();"
                "fr.onload=function(){cb(fr.result)}; fr.onerror=function(){cb(null)}; fr.readAsDataURL(b);}).catch(()=>cb(null));"
            )
            driver.set_script_timeout(timeout)
            result = driver.execute_async_script(script, blob_url)
            return result
        except Exception:
            return None

    # ===== Containers: Nhập văn bản / Tải hình ảnh lên =====
    def _find_prompt_image_containers(self, driver: webdriver.Chrome) -> list:
        """Return a list of container nodes that include both 'Nhập văn bản' and 'Tải hình ảnh lên'."""
        containers = []
        try:
            nodes = driver.find_elements(By.XPATH, 
                "//*[.//label[normalize-space(text())='Nhập văn bản'] or .//button[.//label[normalize-space(text())='Nhập văn bản']] or .//*[normalize-space(text())='Nhập văn bản']]"
            )
            for node in nodes:
                try:
                    has_upload = len(node.find_elements(By.XPATH, 
                        ".//label[normalize-space(text())='Tải hình ảnh lên'] | .//button[.//label[normalize-space(text())='Tải hình ảnh lên']] | .//*[normalize-space(text())='Tải hình ảnh lên']"
                    )) > 0
                    if has_upload:
                        containers.append(node)
                except Exception:
                    continue
        except Exception:
            pass
        # Deduplicate by element id if any
        result = []
        seen = set()
        for c in containers:
            try:
                key = c.id if hasattr(c, 'id') else c
                if key in seen:
                    continue
                seen.add(key)
                result.append(c)
            except Exception:
                result.append(c)
        if not result:
            try:
                self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không tìm thấy container có 'Nhập văn bản' và 'Tải hình ảnh lên'\n")
            except Exception:
                pass
        return result

    def _fill_prompt_and_images_at_index(self, driver: webdriver.Chrome, index: int, prompt_text: str | None, image_paths: list[str] | None) -> bool:
        """Find the container at index and fill prompt and upload images using its controls.
        Resilient to hidden inputs; will send paths to descendant input[type='file'] if present.
        """
        try:
            containers = self._find_prompt_image_containers(driver)
            if not containers:
                return False
            if index < 0 or index >= len(containers):
                try:
                    self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Index {index} vượt ngoài số container hiện có ({len(containers)})\n")
                except Exception:
                    pass
                return False
            container = containers[index]
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", container)
            time.sleep(0.2)

            # Type prompt
            if prompt_text:
                # Prefer textarea inside container
                area = None
                try:
                    areas = container.find_elements(By.TAG_NAME, 'textarea')
                    areas = [a for a in areas if a.is_displayed() and a.is_enabled()]
                    if areas:
                        area = areas[0]
                except Exception:
                    area = None
                if area is None:
                    # Click the "Nhập văn bản" label/button to reveal input, then wait 1s for modal
                    try:
                        text_btn = None
                        candidates = container.find_elements(By.XPATH, 
                            ".//button[.//label[normalize-space(text())='Nhập văn bản']] | .//*[normalize-space(text())='Nhập văn bản']/ancestor::*[self::button or @role='button'][1] | .//*[normalize-space(text())='Nhập văn bản']"
                        )
                        if candidates:
                            text_btn = candidates[0]
                        if text_btn is not None:
                            self._human_click_el(driver, text_btn)
                            time.sleep(1.0)
                    except Exception:
                        pass
                    try:
                        areas = container.find_elements(By.TAG_NAME, 'textarea')
                        areas = [a for a in areas if a.is_displayed() and a.is_enabled()]
                        if areas:
                            area = areas[0]
                    except Exception:
                        area = None
                if area is not None:
                    try:
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", area)
                        time.sleep(0.1)
                        self._human_click_el(driver, area)
                        try:
                            area.clear()
                        except Exception:
                            pass
                        self._human_type_el(area, prompt_text)
                    except Exception:
                        # Fallback: type into any visible textarea on the page
                        self._type_prompt_into_any_textarea(driver, prompt_text)
                else:
                    # Fallback: type into any visible textarea on the page
                    ok = self._type_prompt_into_any_textarea(driver, prompt_text)
                    if not ok:
                        try:
                            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không tìm thấy textarea trong slot {index}\n")
                        except Exception:
                            pass

            # Upload images AFTER typing prompt
            if image_paths:
                uploaded_any = False
                # Click the upload button by label first to mirror user flow
                try:
                    upload_btn = None
                    candidates = container.find_elements(By.XPATH, 
                        ".//button[.//label[normalize-space(text())='Tải hình ảnh lên']] | .//*[normalize-space(text())='Tải hình ảnh lên']/ancestor::*[self::button or @role='button'][1] | .//*[normalize-space(text())='Tải hình ảnh lên']"
                    )
                    if candidates:
                        upload_btn = candidates[0]
                    if upload_btn is not None:
                        self._human_click_el(driver, upload_btn)
                        time.sleep(0.3)
                except Exception:
                    pass
                # Then try direct file input inside container
                try:
                    file_inputs = container.find_elements(By.CSS_SELECTOR, "input[type='file']")
                except Exception:
                    file_inputs = []
                if file_inputs:
                    for idx_path, path in enumerate(image_paths):
                        try:
                            # Ensure the input is interactable
                            try:
                                driver.execute_script("arguments[0].style.display='block'; arguments[0].style.visibility='visible';", file_inputs[0])
                            except Exception:
                                pass
                            file_inputs[0].send_keys(path)
                            # Wait 3 seconds between each upload as requested
                            time.sleep(3.0)
                            uploaded_any = True
                        except Exception:
                            pass
                if not uploaded_any:
                    try:
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không thể upload ảnh tại slot {index}\n")
                    except Exception:
                        pass
            return True
        except Exception:
            return False

    def _upload_image_via_toolbar(self, driver: webdriver.Chrome, image_path: str, timeout: int = 8) -> bool:
        """Upload a single image using the top toolbar that has buttons 'Nhập văn bản' and 'Tải hình ảnh lên'."""
        if not (image_path or '').strip():
            return False
        end_time = time.time() + timeout
        while time.time() < end_time:
            try:
                # Find a toolbar container that has both labels
                toolbars = driver.find_elements(By.XPATH,
                    "//div[.//*[normalize-space(text())='Nhập văn bản'] and .//*[normalize-space(text())='Tải hình ảnh lên']]"
                )
                for tb in toolbars:
                    try:
                        # Click the upload button by label first to mirror user action
                        try:
                            candidates = tb.find_elements(By.XPATH,
                                ".//button[.//label[normalize-space(text())='Tải hình ảnh lên']] | .//*[normalize-space(text())='Tải hình ảnh lên']/ancestor::*[self::button or @role='button'][1] | .//*[normalize-space(text())='Tải hình ảnh lên']"
                            )
                            if candidates:
                                self._human_click_el(driver, candidates[0])
                                time.sleep(0.3)
                        except Exception:
                            pass
                        # Find input[type=file] inside toolbar
                        inputs = tb.find_elements(By.CSS_SELECTOR, "input[type='file']")
                        if inputs:
                            try:
                                driver.execute_script("arguments[0].style.display='block'; arguments[0].style.visibility='visible';", inputs[0])
                            except Exception:
                                pass
                            inputs[0].send_keys(image_path)
                            # Wait 3 seconds after upload
                            time.sleep(3.0)
                            return True
                        # If no input found, click the upload button by label inside this toolbar
                        candidates = tb.find_elements(By.XPATH,
                            ".//button[.//label[normalize-space(text())='Tải hình ảnh lên']] | .//*[normalize-space(text())='Tải hình ảnh lên']/ancestor::*[self::button or @role='button'][1] | .//*[normalize-space(text())='Tải hình ảnh lên']"
                        )
                        if candidates:
                            self._human_click_el(driver, candidates[0])
                            time.sleep(0.3)
                            inputs = tb.find_elements(By.CSS_SELECTOR, "input[type='file']")
                            if inputs:
                                try:
                                    driver.execute_script("arguments[0].style.display='block'; arguments[0].style.visibility='visible';", inputs[0])
                                except Exception:
                                    pass
                                inputs[0].send_keys(image_path)
                                # Wait 3 seconds after upload
                                time.sleep(3.0)
                                return True
                    except Exception:
                        continue
            except Exception:
                pass
            time.sleep(0.25)
        try:
            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không thể upload ảnh qua toolbar (không tìm thấy input hoặc nút)\n")
        except Exception:
            pass
        return False

    def _wait_until(self, predicate, timeout: int = 60, interval: float = 1.0) -> bool:
        start = time.time()
        while time.time() - start < timeout:
            try:
                if predicate():
                    return True
            except Exception:
                pass
            time.sleep(interval)
        return False

    def _is_driver_alive(self, driver: webdriver.Chrome | None) -> bool:
        try:
            if driver is None:
                return False
            # A simple no-op command to verify session
            driver.current_url  # access property to trigger
            return True
        except Exception:
            return False

    def _open_new_tab_and_close_current(self, driver: webdriver.Chrome, url: str) -> None:
        """Open a new tab to the given URL, then close the previous tab and switch to the new one."""
        try:
            old_handle = driver.current_window_handle
            # Open new tab
            try:
                driver.switch_to.new_window('tab')
            except Exception:
                # Fallback
                driver.execute_script("window.open('about:blank','_blank');")
                handles = driver.window_handles
                driver.switch_to.window(handles[-1])
            # Navigate in new tab
            driver.get(url)
            # Close the old tab if it still exists
            try:
                driver.switch_to.window(old_handle)
                driver.close()
            except Exception:
                pass
            # Switch back to the new tab (last handle)
            try:
                driver.switch_to.window(driver.window_handles[-1])
            except Exception:
                pass
        except Exception:
            # Fallback: just navigate in the same tab
            try:
                driver.get(url)
            except Exception:
                pass

    def _remember_profile(self, email_addr: str, cache_dir: str, user_agent: str) -> None:
        if not email_addr:
            return
        self.whisk_profiles[email_addr] = {
            "cache_dir": cache_dir,
            "user_agent": user_agent,
            "last_login": int(time.time()),
        }
        self._save_profiles()
        self._refresh_profiles_list()

    # ===================== Profile actions =====================
    def _open_selected_profile(self) -> None:
        try:
            sel = self.profiles_list.curselection()
            if not sel:
                messagebox.showinfo("Thông báo", "Vui lòng chọn một tài khoản trong danh sách!")
                return
            line = self.profiles_list.get(sel[0])
            email_addr = line.split("  |  ")[0].strip()
            meta = self.whisk_profiles.get(email_addr)
            if not meta:
                messagebox.showerror("Lỗi", "Không tìm thấy cache!")
                return
            threading.Thread(target=self._open_profile_thread, args=(email_addr, meta), daemon=True).start()
        except Exception as ex:
            messagebox.showerror("Lỗi", f"Không thể mở profile: {ex}")

    def _open_profile_thread(self, email_addr: str, meta: dict) -> None:
        try:
            drv = self._open_profile_driver(meta)
            # Open the Whisk project page directly as requested
            drv.get(WHISK_PROJECT_URL)
            self._wait_until(lambda: "labs.google" in (drv.current_url or ""), timeout=120)
            
            # Close welcome modal if it appears
            try:
                self._close_welcome_modal(drv)
            except Exception:
                pass
                
            self._set_status(f"Đã mở Whisk cho {email_addr}", "green")
            self._set_exec_status(f"Đã mở Whisk cho {email_addr}", "green")
        except Exception as ex:
            self._set_status(f"Lỗi mở Whisk: {ex}", "red")
            self._set_exec_status(f"Lỗi mở Whisk: {ex}", "red")

    def _open_profile_driver(self, meta: dict) -> webdriver.Chrome:
        chrome_options = Options()
        cache_dir = meta.get("cache_dir")
        if cache_dir and os.path.isdir(cache_dir):
            chrome_options.add_argument(f"--user-data-dir={cache_dir}")
            chrome_options.add_argument("--profile-directory=Default")
        if meta.get("user_agent"):
            chrome_options.add_argument(f"--user-agent={meta['user_agent']}")
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--lang=vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7")
        # Headless toggle for execute flow
        try:
            if self.headless_mode.get():
                chrome_options.add_argument("--headless")
                chrome_options.add_argument("--no-sandbox")
                chrome_options.add_argument("--disable-dev-shm-usage")
                # Set window size for headless mode (16-inch laptop resolution)
                chrome_options.add_argument("--window-size=1920,1080")
                chrome_options.add_argument("--start-maximized")
                # Additional options for better headless rendering
                chrome_options.add_argument("--disable-gpu")
                chrome_options.add_argument("--disable-extensions")
                chrome_options.add_argument("--disable-plugins")
                # Note: We don't disable images as we need them for downloads
        except Exception:
            pass
        try:
            chrome_options.set_capability('goog:loggingPrefs', {'performance': 'ALL'})
        except Exception:
            pass
        service = Service(ChromeDriverManager().install())
        
        # Retry logic for Chrome crashes
        max_retries = 3
        for attempt in range(max_retries):
            try:
                drv = webdriver.Chrome(service=service, options=chrome_options)
                try:
                    drv.execute_cdp_cmd('Network.enable', {})
                except Exception:
                    pass
                return drv
            except Exception as e:
                error_msg = str(e).lower()
                if "chrome failed to start" in error_msg or "crashed" in error_msg or "session not created" in error_msg:
                    if attempt < max_retries - 1:
                        try:
                            self._append_exec_log(f"[SYSTEM] {time.strftime('%H:%M:%S')} | Chrome crash detected, đóng browser và thử lại... (lần {attempt + 1}/{max_retries})\n")
                        except Exception:
                            pass
                        # Kill any remaining Chrome processes
                        self._kill_chrome_processes()
                        time.sleep(2)  # Wait before retry
                        continue
                    else:
                        try:
                            self._append_exec_log(f"[ERROR] {time.strftime('%H:%M:%S')} | Chrome crash sau {max_retries} lần thử: {e}\n")
                        except Exception:
                            pass
                        raise e
                else:
                    # Re-raise non-crash errors immediately
                    raise e
        
        # Should never reach here, but just in case
        raise Exception("Không thể khởi tạo Chrome sau nhiều lần thử")

    # ===== Job queue & execution =====
    def _enqueue_or_start_account_job(self, email_addr: str, meta: dict, row: dict) -> None:
        # Assign a monotonically increasing job id for clearer logs
        self.job_counter = getattr(self, 'job_counter', 0) + 1
        job_id = self.job_counter
        job = {"email": email_addr, "meta": meta, "row": row, "job_id": job_id}
        st = self.account_states.get(email_addr)
        if st is None:
            st = {'queue': [], 'running': False, 'lock': threading.Lock()}
            self.account_states[email_addr] = st
        with st['lock']:
            if st['running']:
                st['queue'].append(job)
                action = "queue"
            else:
                st['running'] = True
                action = "start"
        if action == "queue":
            self._append_exec_log(f"[JOB #{job_id}] {time.strftime('%H:%M:%S')} | Queued for {email_addr}\n")
            try:
                self._update_exec_counters()
            except Exception:
                pass
        else:
            self._append_exec_log(f"[JOB #{job_id}] {time.strftime('%H:%M:%S')} | Started for {email_addr}\n")
            threading.Thread(target=self._execute_row_thread, args=(email_addr, meta, row, job_id), daemon=True).start()

    def _execute_row_thread(self, email_addr: str, meta: dict, row: dict, job_id: int | None = None) -> None:
        try:
            # Reuse existing driver if possible; otherwise create a new one
            drv = self.exec_drivers.get(email_addr)
            if not self._is_driver_alive(drv):
                drv = self._open_profile_driver(meta)
                self.exec_drivers[email_addr] = drv
            self.stop_exec = False
            self._set_exec_status(f"Mở Whisk cho {email_addr}...", 'orange')
            # Prepare a concise summary for logs
            try:
                main_prompt_preview = (row.get('main promt') or '')[:60]
            except Exception:
                main_prompt_preview = ''
            try:
                size_preview = (row.get('size') or '')
            except Exception:
                size_preview = ''
            self._append_exec_log(f"[JOB #{job_id}] {time.strftime('%H:%M:%S')} | START email={email_addr} size={size_preview} main='{main_prompt_preview}'\n")
            # Open new tab for this execution and close the previous one
            self._open_new_tab_and_close_current(drv, WHISK_PROJECT_URL)
            self._wait_until(lambda: "labs.google" in (drv.current_url or ""), timeout=120)
            
            # Close welcome modal if it appears
            try:
                self._close_welcome_modal(drv)
            except Exception:
                pass
                
            # Retry từ đầu nếu không đủ 2 ảnh kết quả (tối đa 2 lần retry)
            attempt_success = False
            max_attempts = 3  # 1 lần đầu + 2 lần retry
            for attempt in range(1, max_attempts + 1):
                time.sleep(3)
                # Click "Nhập công cụ" nếu có
                try:
                    if self._click_button_by_text(drv, "Nhập công cụ", timeout=1):
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Đã nhấn 'Nhập công cụ' (lần {attempt}/{max_attempts})\n")
                except Exception:
                    pass
                # Nhập main promt và chọn size
                try:
                    main_prompt = (row.get('main promt') or '').strip()
                    if main_prompt:
                        self._type_prompt_into_any_textarea(drv, main_prompt)
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Đã nhập main promt vào textarea (lần {attempt}/{max_attempts})\n")
                    size_val = (row.get('size') or '').strip()
                    if size_val:
                        if self._open_aspect_ratio_menu(drv):
                            time.sleep(1)
                            self._pick_aspect_ratio(drv, size_val)
                except Exception as _ex:
                    self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Lỗi nhập promt: {_ex}\n")

                # Upload ảnh nếu có
                uploaded_image_count = 0
                try:
                    has_image = any([
                        (row.get('image_1') or '').strip(),
                        (row.get('image_2') or '').strip(),
                        (row.get('kind_image') or '').strip(),
                        (row.get('screen_image') or '').strip(),
                    ])
                    if has_image:
                        try:
                            if self._click_add_image_button(drv):
                                self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Đã nhấn 'Thêm hình ảnh' (lần {attempt}/{max_attempts})\n")
                                time.sleep(1)
                                if self._click_button_by_aria_label(drv, "Thêm danh mục mới", timeout=4):
                                    self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Đã nhấn 'Thêm danh mục mới'\n")
                                    time.sleep(1)
                        except Exception:
                            pass
                        field_sets = [
                            ('', [p for p in [row.get('image_1')] if (p or '').strip()]),
                            ('', [p for p in [row.get('image_2')] if (p or '').strip()]),
                            ('', [p for p in [row.get('screen_image')] if (p or '').strip()]),
                            ('', [p for p in [row.get('kind_image')] if (p or '').strip()]),
                        ]
                        for idx, (ptext, imgs) in enumerate(field_sets):
                            # Only keep images that actually exist on disk
                            imgs_to_use = []
                            for _p in imgs:
                                try:
                                    if os.path.isfile(_p):
                                        imgs_to_use.append(_p)
                                    else:
                                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | ⚠️ Ảnh không tồn tại, bỏ qua: {_p}\n")
                                except Exception:
                                    continue
                            if imgs_to_use:
                                uploaded_image_count += len(imgs_to_use)
                            if ptext or imgs_to_use:
                                filled = self._fill_prompt_and_images_at_index(drv, idx, ptext, imgs_to_use)
                                if filled:
                                    self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Slot {idx}: đã tải ảnh/nhập văn bản\n")
                                else:
                                    self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Slot {idx}: không tìm thấy container\n")
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Tổng cộng đã upload {uploaded_image_count} ảnh\n")
                        # Đợi 20 giây sau khi upload ảnh xong trước khi tiếp tục các step tiếp theo
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Đợi 20 giây sau khi upload ảnh...\n")
                        time.sleep(20.0)
                    else:
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không có hình ảnh để upload, tiếp tục với text prompt\n")
                    # Chọn size sau upload (nếu cần)
                    try:
                        if self._open_aspect_ratio_menu(drv):
                            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Đã mở menu Tỷ lệ khung hình\n")
                            time.sleep(1)
                            size_val_after = (row.get('size') or '').strip()
                            if size_val_after:
                                self._pick_aspect_ratio(drv, size_val_after)
                    except Exception:
                        pass
                    # Tune + seed
                    try:
                        if self._open_tune_menu(drv):
                            time.sleep(1)
                            self._set_random_seed(drv)
                    except Exception:
                        pass
                    # Gửi câu lệnh
                    try:
                        self._click_submit_execute(drv)
                    except Exception:
                        pass
                    # Tải ảnh kết quả
                    try:
                        main_prompt = (row.get('main promt') or '').strip()
                        saved_count = self._download_result_images(drv, wait_seconds=30, max_images=10, skip_count=uploaded_image_count, prompt_text=main_prompt)
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Đã tải {saved_count} ảnh kết quả mới\n")
                    except Exception:
                        saved_count = 0
                except Exception as _ex:
                    self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Lỗi xử lý hình ảnh: {_ex}\n")
                    saved_count = 0

                # Kiểm tra điều kiện đủ ảnh
                if saved_count >= 2:
                    attempt_success = True
                    break
                else:
                    if attempt < max_attempts:
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Không đủ 2 ảnh (có {saved_count}). Thử lại từ đầu ({attempt+1}/{max_attempts})...\n")
                        try:
                            self._open_new_tab_and_close_current(drv, WHISK_PROJECT_URL)
                            self._wait_until(lambda: "labs.google" in (drv.current_url or ""), timeout=120)
                            time.sleep(2)
                            try:
                                self._close_welcome_modal(drv)
                            except Exception:
                                pass
                        except Exception:
                            pass
                    else:
                        self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Hết số lần retry. Vẫn không đủ 2 ảnh.\n")

            # Tổng kết và cập nhật counters theo kết quả
            summary = (
                f"size={row.get('size') or ''} | main='" + (row.get('main promt') or '')[:60] + "' | "
                f"img1={os.path.basename(row.get('image_1') or '')} | img2={os.path.basename(row.get('kind_image') or row.get('image_2') or '')}"
            )
            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Executing: {summary}\n")
            time.sleep(1)
            if attempt_success:
                self._append_exec_log(f"[JOB #{job_id}] {time.strftime('%H:%M:%S')} | SUCCESS email={email_addr} (đủ ≥2 ảnh)\n")
                try:
                    self.exec_success_count = getattr(self, 'exec_success_count', 0) + 1
                    self._update_exec_counters()
                except Exception:
                    pass
            else:
                self._append_exec_log(f"[JOB #{job_id}] {time.strftime('%H:%M:%S')} | FAILED email={email_addr} (không đủ ảnh kết quả)\n")
                try:
                    self.exec_error_count = getattr(self, 'exec_error_count', 0) + 1
                    self._update_exec_counters()
                except Exception:
                    pass
        except Exception as ex:
            self._set_exec_status(f"Lỗi execute: {ex}", 'red')
            try:
                self.exec_error_count = getattr(self, 'exec_error_count', 0) + 1
                self._update_exec_counters()
            except Exception:
                pass
            try:
                self._append_exec_log(f"[JOB #{job_id}] {time.strftime('%H:%M:%S')} | FAILED email={email_addr} error={str(ex)}\n")
            except Exception:
                pass
        finally:
            try:
                # Keep driver alive for subsequent jobs; do not quit here
                local_drv = self.exec_drivers.get(email_addr)
                if not self._is_driver_alive(local_drv):
                    # Clean up reference if it died unexpectedly
                    if email_addr in self.exec_drivers:
                        del self.exec_drivers[email_addr]
            except Exception:
                pass
            try:
                # Do not remove alive driver; it will be reused for next job
                pass
            except Exception:
                pass
            # Auto-run next job if available
            try:
                st = self.account_states.get(email_addr)
                next_job = None
                if st is None:
                    st = {'queue': [], 'running': False, 'lock': threading.Lock()}
                    self.account_states[email_addr] = st
                with st['lock']:
                    if st['queue']:
                        next_job = st['queue'].pop(0)
                    else:
                        st['running'] = False
                if next_job is not None:
                    self._append_exec_log(f"[JOB #{next_job.get('job_id')}] {time.strftime('%H:%M:%S')} | Started for {next_job.get('email')}\n")
                    threading.Thread(target=self._execute_row_thread, args=(next_job['email'], next_job['meta'], next_job['row'], next_job.get('job_id')), daemon=True).start()
                else:
                    self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | All jobs completed for {email_addr}\n")
                    # Close the browser for this account when there are no more queued jobs
                    try:
                        drv_to_close = self.exec_drivers.get(email_addr)
                        if drv_to_close is not None:
                            try:
                                drv_to_close.quit()
                            except Exception:
                                pass
                        # Remove reference so a new session is created next time
                        self.exec_drivers.pop(email_addr, None)
                        try:
                            self._append_exec_log(f"[SYSTEM] {time.strftime('%H:%M:%S')} | Đã đóng browser cho {email_addr} (hết queue)\n")
                        except Exception:
                            pass
                    except Exception:
                        pass
                try:
                    self._update_exec_counters()
                except Exception:
                    pass
            except Exception as ex:
                try:
                    st = self.account_states.get(email_addr)
                    if st:
                        with st['lock']:
                            st['running'] = False
                except Exception:
                    pass
                self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Queue scheduling error: {ex}\n")

    def _delete_selected_profile(self) -> None:
        try:
            sel = self.profiles_list.curselection()
            if not sel:
                messagebox.showinfo("Thông báo", "Vui lòng chọn một tài khoản trong danh sách!")
                return
            line = self.profiles_list.get(sel[0])
            email_addr = line.split("  |  ")[0].strip()
            meta = self.whisk_profiles.get(email_addr)
            if not meta:
                return
            cache_dir = meta.get("cache_dir")
            try:
                if cache_dir and os.path.isdir(cache_dir):
                    import shutil
                    shutil.rmtree(cache_dir, ignore_errors=True)
            except Exception:
                pass
            self.whisk_profiles.pop(email_addr, None)
            self._save_profiles()
            self._refresh_profiles_list()
            self._set_status(f"Đã xóa cache của {email_addr}", "green")
        except Exception as ex:
            messagebox.showerror("Lỗi", f"Không thể xóa cache: {ex}")

    def _duplicate_selected_profile(self) -> None:
        try:
            sel = self.profiles_list.curselection()
            if not sel:
                messagebox.showinfo("Thông báo", "Vui lòng chọn một tài khoản trong danh sách!")
                return
            line = self.profiles_list.get(sel[0])
            source_key = line.split("  |  ")[0].strip()
            meta = self.whisk_profiles.get(source_key)
            if not meta:
                messagebox.showerror("Lỗi", "Không tìm thấy cache nguồn!")
                return

            src_cache = meta.get("cache_dir")
            if not src_cache or not os.path.isdir(src_cache):
                messagebox.showerror("Lỗi", "Thư mục cache nguồn không tồn tại!")
                return

            # Auto-generate new profile label using timestamp
            timestamp_label = time.strftime('%Y%m%d_%H%M%S')
            new_key = f"{source_key}_{timestamp_label}"
            # Ensure uniqueness among profile keys
            if new_key in self.whisk_profiles:
                new_key = f"{new_key}_{random.randint(1000,9999)}"

            # Prepare destination cache dir
            safe_key = re.sub(r'[^a-zA-Z0-9_.-]', '_', new_key) or "clone"
            base_dst_cache = os.path.join(os.getcwd(), "chrome_cache", f"whisk_{safe_key}")
            dst_cache = base_dst_cache
            # If destination exists, add a random suffix to avoid collision
            if os.path.exists(dst_cache):
                dst_cache = f"{base_dst_cache}_{random.randint(1000,9999)}"

            # Copy cache directory excluding lock files
            try:
                import shutil
                def _ignore(dir, names):
                    ignores = {"SingletonLock", "LOCK", "lockfile"}
                    return [n for n in names if n in ignores]
                shutil.copytree(src_cache, dst_cache, ignore=_ignore)
            except Exception as ex:
                messagebox.showerror("Lỗi", f"Không thể nhân bản cache: {ex}")
                return

            # Register new profile
            self.whisk_profiles[new_key] = {
                "cache_dir": dst_cache,
                "user_agent": meta.get("user_agent") or random.choice(self.user_agents),
                "last_login": int(time.time()),
            }
            self._save_profiles()
            self._refresh_profiles_list()
            self._set_status(f"Đã nhân bản cache từ '{source_key}' → '{new_key}'", "green")
            try:
                self._append_log(f"[SYSTEM] {time.strftime('%H:%M:%S')} | Duplicated cache to {dst_cache}\n")
            except Exception:
                pass
        except Exception as ex:
            messagebox.showerror("Lỗi", f"Không thể nhân bản cache: {ex}")

    # ===================== UX helpers =====================
    def _append_log(self, text: str) -> None:
        try:
            if hasattr(self, 'log'):
                self.log.configure(state='normal')
                self.log.insert(tk.END, text)
                self.log.see(tk.END)
                self.log.configure(state='disabled')
        except Exception:
            pass

    # Exec tab helpers (logging/status for execute tab)
    def _append_exec_log(self, text: str) -> None:
        try:
            if hasattr(self, 'exec_log'):
                self.exec_log.configure(state='normal')
                self.exec_log.insert(tk.END, text)
                self.exec_log.see(tk.END)
                self.exec_log.configure(state='disabled')
        except Exception:
            pass

    def _set_exec_status(self, text: str, color: str) -> None:
        try:
            if getattr(self, 'use_tk_ui', True) and hasattr(self, 'exec_status'):
                style_map = {
                    'blue': 'Info.TLabel',
                    'green': 'Success.TLabel',
                    'red': 'Error.TLabel',
                    'orange': 'Warning.TLabel'
                }
                style = style_map.get(color, 'Info.TLabel')
                self.exec_status.config(text=text, style=style)
        except Exception:
            pass

    def _update_exec_counters(self) -> None:
        try:
            # Queued total across accounts
            total_q = 0
            running = 0
            for _, st in self.account_states.items():
                try:
                    with st['lock']:
                        total_q += len(st['queue'])
                        running += 1 if st.get('running') else 0
                except Exception:
                    continue
            if hasattr(self, 'lbl_queue'):
                self.lbl_queue.config(text=str(total_q))
            if hasattr(self, 'lbl_exec'):
                self.lbl_exec.config(text=str(running))
            if hasattr(self, 'lbl_ok'):
                self.lbl_ok.config(text=str(getattr(self, 'exec_success_count', 0)))
            if hasattr(self, 'lbl_err'):
                self.lbl_err.config(text=str(getattr(self, 'exec_error_count', 0)))
        except Exception:
            pass

    # ===== Excel Template + Import =====
    def _download_excel_template_whisk(self) -> None:
        try:
            from tkinter import filedialog
            try:
                from openpyxl import Workbook
                from openpyxl.styles import NamedStyle
            except Exception:
                messagebox.showerror("Thiếu thư viện", "Cần cài openpyxl để tạo template: pip install openpyxl")
                return
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                initialfile="whisk_template.xlsx",
                title="Lưu template Excel"
            )
            if not path:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "Tasks"
            # Headers per request (keep the exact keys as provided)
            headers = [
                "main promt",
                "image_1",
                "image_2",
                "screen_image",
                "kind_image",
                "size",  # allowed: 1:1, 9:16, 16:9, 3:4, 4:3
            ]
            ws.append(headers)
            
            # Set size column (column F) to text format to preserve aspect ratio strings
            try:
                # Create a text style
                text_style = NamedStyle(name="text_style")
                text_style.number_format = '@'  # Text format
                
                # Apply text format to the entire size column (column F)
                for row in range(2, 100):  # Apply to first 98 data rows
                    cell = ws[f'F{row}']
                    cell.number_format = '@'
            except Exception:
                pass
            
            # Sample rows with explicit string formatting for size column
            ws.append(["A hero shot product on clean background", "C:/images/p1.jpg", "C:/images/p2.jpg", "C:/images/screen.jpg", "C:/images/variantA.jpg", "1:1"]) 
            ws.append(["Street fashion look", "C:/images/a.jpg", "C:/images/b.jpg", "", "", "9:16"]) 
            
            # Ensure sample data size values are treated as text
            try:
                ws['F2'].number_format = '@'  # 1:1
                ws['F3'].number_format = '@'  # 9:16
            except Exception:
                pass
                
            wb.save(path)
            self._append_exec_log(f"[EXEC] {time.strftime('%H:%M:%S')} | Đã lưu template: {path}\n")
        except Exception as ex:
            messagebox.showerror("Lỗi", f"Không thể tạo template: {ex}")

    def _import_excel_whisk(self) -> None:
        try:
            from tkinter import filedialog
            try:
                from openpyxl import load_workbook
            except Exception:
                messagebox.showerror("Thiếu thư viện", "Cần cài openpyxl để import: pip install openpyxl")
                return
            path = filedialog.askopenfilename(title="Chọn file Excel", filetypes=[("Excel", "*.xlsx")])
            if not path:
                return
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            ws = wb.active

            def _cell_to_str(val):
                try:
                    if val is None:
                        return ''
                    return str(val).strip()
                except Exception:
                    return ''

            rows = []
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if row is None:
                    continue
                if i == 0:
                    headers = [(_cell_to_str(c) or '').lower() for c in row]
                    continue
                # Map by header name; tolerate minor case differences and spaces
                def get(col_name: str):
                    try:
                        idx = headers.index(col_name)
                        return _cell_to_str(row[idx])
                    except Exception:
                        return ''
                item = {
                    'main promt': get('main promt'),
                    'image_1': get('image_1'),
                    'image_2': get('image_2'),
                    'screen_image': get('screen_image'),
                    'kind_image': get('kind_image'),
                    'size': get('size'),
                }
                # Chỉ nhận các hàng có main promt (không rỗng)
                if (item.get('main promt') or '').strip():
                    rows.append(item)

            if not rows:
                messagebox.showerror("Lỗi", "Không có dữ liệu hợp lệ trong file Excel!")
                return

            # Dispatch jobs round-robin across cached accounts
            emails = list(self.whisk_profiles.keys())
            if not emails:
                try:
                    self._set_exec_status("Bạn cần đăng nhập ít nhất một tài khoản trước khi import.", 'orange')
                except Exception:
                    pass
                try:
                    messagebox.showwarning("Cần đăng nhập", "Bạn cần đăng nhập ít nhất một tài khoản trước khi import.")
                except Exception:
                    pass
                return
            self._set_exec_status(f"Imported {len(rows)} row(s) - dispatching...", 'orange')
            idx_email = 0
            for r in rows:
                target_email = emails[idx_email % len(emails)]
                idx_email += 1
                meta = self.whisk_profiles.get(target_email)
                if not meta:
                    continue
                self._enqueue_or_start_account_job(target_email, meta, r)
            self._set_exec_status("Jobs enqueued/started.", 'green')
        except Exception as ex:
            messagebox.showerror("Lỗi", f"Không thể import Excel: {ex}")

    def _close_all_browsers(self) -> None:
        """Close all browser instances (main driver and execution drivers)."""
        try:
            closed_count = 0
            
            # Close main driver
            if hasattr(self, 'driver') and self.driver is not None:
                try:
                    self.driver.quit()
                    self.driver = None
                    closed_count += 1
                except Exception:
                    pass
            
            # Close all execution drivers
            if hasattr(self, 'exec_drivers') and self.exec_drivers:
                for email, driver in list(self.exec_drivers.items()):
                    try:
                        if driver is not None:
                            driver.quit()
                            closed_count += 1
                    except Exception:
                        pass
                self.exec_drivers.clear()
            
            # Log the cleanup
            if closed_count > 0:
                try:
                    self._append_exec_log(f"[SYSTEM] {time.strftime('%H:%M:%S')} | Đã đóng {closed_count} browser instances\n")
                except Exception:
                    pass
                
        except Exception:
            pass

    def _set_status(self, text: str, color: str) -> None:
        try:
            if hasattr(self, 'ui_callbacks') and self.ui_callbacks.get('on_status'):
                try:
                    self.ui_callbacks['on_status'](text, color)
                except Exception:
                    pass
            if getattr(self, 'use_tk_ui', True) and hasattr(self, 'status_label'):
                style_map = {
                    'blue': 'Info.TLabel',
                    'green': 'Success.TLabel',
                    'red': 'Error.TLabel',
                    'orange': 'Warning.TLabel'
                }
                style = style_map.get(color, 'Info.TLabel')
                self.status_label.config(text=text, style=style)
        except Exception:
            pass


def main() -> None:
    # Use ttkbootstrap Window when available for a nicer UI
    if _HAS_TTKBOOTSTRAP and TtkbWindow is not None:
        root = TtkbWindow(themename='superhero')
    else:
        root = tk.Tk()
    app = WhiskBrowserTool(root, use_tk_ui=True)

    def _on_app_close():
        try:
            # Close all browser instances
            app._close_all_browsers()
        except Exception:
            pass
        try:
            root.destroy()
        except Exception:
            pass

    try:
        root.protocol("WM_DELETE_WINDOW", _on_app_close)
    except Exception:
        pass
    root.mainloop()


if __name__ == "__main__":
    main()


